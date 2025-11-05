"""
Test with debug output to see what's being parsed and how distinctiveness is extracted.
"""

import io
import pandas as pd
import openpyxl
from metric_reader import parse_metric_requirements, normalise_requirements, open_metric_workbook


class MockUploadedFile:
    """Mock file object for testing"""
    def __init__(self, buffer):
        self.buffer = buffer
        self.name = "test_metric.xlsx"
    
    def read(self):
        return self.buffer.read()


def test_with_debug():
    """
    Test with debug output
    """
    wb = openpyxl.Workbook()
    
    # Create Headline Results sheet  
    ws_headline = wb.create_sheet("Headline Results")
    
    headers = ["Unit Type", "Target", "Baseline Units", "Units Required", "Unit Deficit"]
    for col, header in enumerate(headers, start=1):
        ws_headline.cell(row=5, column=col, value=header)
    
    data = [
        ["Habitat units", "10.00%", 10.00, 11.00, 0.00],
        ["Hedgerow units", "10.00%", 0.00, 0.00, 0.00],
        ["Watercourse units", "10.00%", 0.00, 0.00, 0.00],
    ]
    
    for row_idx, row_data in enumerate(data, start=6):
        for col_idx, value in enumerate(row_data, start=1):
            ws_headline.cell(row=row_idx, column=col_idx, value=value)
    
    # Create Trading Summary Hedgerows sheet
    ws_hedge = wb.create_sheet("Trading Summary Hedgerows")
    
    row = 1
    
    # Medium Distinctiveness section
    ws_hedge.cell(row=row, column=1, value="Medium Distinctiveness")
    row += 1
    ws_hedge.cell(row=row, column=1, value="Habitat group")
    ws_hedge.cell(row=row, column=2, value="On-site unit change")
    ws_hedge.cell(row=row, column=3, value="Off-site unit change")
    ws_hedge.cell(row=row, column=4, value="Project-wide unit change")
    row += 1
    ws_hedge.cell(row=row, column=1, value="Species-rich native hedgerow")
    ws_hedge.cell(row=row, column=2, value=1.47)
    ws_hedge.cell(row=row, column=3, value=0.00)
    ws_hedge.cell(row=row, column=4, value=1.47)
    row += 1
    ws_hedge.cell(row=row, column=1, value="")  # Total row
    ws_hedge.cell(row=row, column=2, value=1.47)
    ws_hedge.cell(row=row, column=3, value=0.00)
    ws_hedge.cell(row=row, column=4, value=1.47)
    row += 2
    
    # Low Distinctiveness section
    ws_hedge.cell(row=row, column=1, value="Low Distinctiveness")
    row += 1
    ws_hedge.cell(row=row, column=1, value="Habitat group")
    ws_hedge.cell(row=row, column=2, value="On-site unit change")
    ws_hedge.cell(row=row, column=3, value="Off-site unit change")
    ws_hedge.cell(row=row, column=4, value="Project-wide unit change")
    row += 1
    ws_hedge.cell(row=row, column=1, value="Native hedgerow")
    ws_hedge.cell(row=row, column=2, value=-0.70)
    ws_hedge.cell(row=row, column=3, value=0.00)
    ws_hedge.cell(row=row, column=4, value=-0.70)
    row += 1
    ws_hedge.cell(row=row, column=1, value="")  # Total row
    ws_hedge.cell(row=row, column=2, value=-0.70)
    ws_hedge.cell(row=row, column=3, value=0.00)
    ws_hedge.cell(row=row, column=4, value=-0.70)
    
    # Remove default sheet
    if "Sheet" in wb.sheetnames:
        wb.remove(wb["Sheet"])
    
    # Save to BytesIO
    excel_buffer = io.BytesIO()
    wb.save(excel_buffer)
    excel_buffer.seek(0)
    
    # Create mock uploaded file
    mock_file = MockUploadedFile(excel_buffer)
    
    # Parse with normalise_requirements directly to see what's parsed
    print("\n" + "="*80)
    print("DEBUG: Direct normalise_requirements call")
    print("="*80)
    
    xls = open_metric_workbook(mock_file)
    HEDGE_SHEETS = [
        "Trading Summary Hedgerows",
        "Hedgerows Trading Summary",
        "Hedgerow Trading Summary",
        "Trading Summary (Hedgerows)"
    ]
    
    hedge_norm, colmap, sheet = normalise_requirements(xls, HEDGE_SHEETS, "Hedgerows")
    
    print("\nParsed hedgerow data:")
    print(hedge_norm)
    
    print("\nColumn mapping:")
    for k, v in colmap.items():
        print(f"  {k}: {v}")
    
    print("\nDistinctiveness values:")
    for idx, row in hedge_norm.iterrows():
        print(f"  {row['habitat']}: distinctiveness='{row['distinctiveness']}'")
    
    # Now call parse_metric_requirements
    excel_buffer.seek(0)
    mock_file2 = MockUploadedFile(excel_buffer)
    
    print("\n" + "="*80)
    print("Full parse_metric_requirements call")
    print("="*80)
    
    requirements = parse_metric_requirements(mock_file2)
    
    hedge_req_df = requirements["hedgerows"]
    
    print("\nFinal requirements:")
    if hedge_req_df.empty:
        print("  (empty - all requirements covered)")
    else:
        print(hedge_req_df)
    
    return True


if __name__ == "__main__":
    test_with_debug()
