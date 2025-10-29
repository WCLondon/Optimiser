"""
Comprehensive test for metric_reader with trading rules
"""

import io
import pandas as pd
import openpyxl
from metric_reader import parse_metric_requirements


def test_with_offsets():
    """Test that on-site offsets are applied"""
    # Create a mock Excel file with deficits AND surpluses
    wb = openpyxl.Workbook()
    
    # Create Trading Summary Area Habitats sheet
    ws = wb.create_sheet("Trading Summary Area Habitats")
    
    # Add very high distinctiveness header
    ws.cell(row=1, column=1, value="Very High Distinctiveness")
    ws.cell(row=2, column=1, value="")
    
    # Add medium distinctiveness header
    ws.cell(row=8, column=1, value="Medium Distinctiveness")
    ws.cell(row=9, column=1, value="")
    
    # Add headers for data table
    headers = ["Habitat", "Broad habitat", "Distinctiveness", "Project-wide unit change", "On-site unit change"]
    for col, header in enumerate(headers, start=1):
        ws.cell(row=15, column=col, value=header)
    
    # Add test data with deficit and matching surplus
    # Medium distinctiveness, same broad group - should offset
    data = [
        ["Grassland A", "Grassland and marsh", "Medium", -5.0, 0.0],  # Deficit
        ["Grassland B", "Grassland and marsh", "Medium", 3.0, 0.0],   # Surplus - can offset 3 units
        # Result: Grassland A should have 2.0 units unmet (5.0 - 3.0)
    ]
    
    for row_idx, row_data in enumerate(data, start=16):
        for col_idx, value in enumerate(row_data, start=1):
            ws.cell(row=row_idx, column=col_idx, value=value)
    
    # Create Headline Results sheet with target
    ws_headline = wb.create_sheet("Headline Results")
    ws_headline.cell(row=1, column=1, value="Headline Results")
    ws_headline.cell(row=5, column=1, value="Unit Type")
    ws_headline.cell(row=5, column=2, value="Baseline")
    ws_headline.cell(row=5, column=3, value="Target %")
    ws_headline.cell(row=6, column=1, value="Area habitat units")
    ws_headline.cell(row=6, column=2, value=100.0)  # baseline
    ws_headline.cell(row=6, column=3, value="10 %")  # 10% target
    
    # Remove default sheet
    if "Sheet" in wb.sheetnames:
        wb.remove(wb["Sheet"])
    
    # Save to BytesIO
    excel_buffer = io.BytesIO()
    wb.save(excel_buffer)
    excel_buffer.seek(0)
    
    # Create mock uploaded file
    class MockUploadedFile:
        def __init__(self, buffer):
            self.buffer = buffer
            self.name = "test_metric_offsets.xlsx"
        
        def read(self):
            return self.buffer.read()
    
    mock_file = MockUploadedFile(excel_buffer)
    
    # Test parsing
    try:
        requirements = parse_metric_requirements(mock_file)
        
        print("✅ parse_metric_requirements with offsets succeeded")
        print(f"   Area habitats: {len(requirements['area'])} rows")
        
        area_df = requirements['area']
        if not area_df.empty:
            print("\nArea requirements (after offsets):")
            for _, row in area_df.iterrows():
                print(f"   - {row['habitat']}: {row['units']:.2f} units")
            
            # Check that offsets were applied
            # We expect:
            # 1. Grassland A residual: 2.0 units (5.0 - 3.0 from Grassland B)
            # 2. Headline requirement: 10.0 units (10% of 100 baseline, no surplus to cover it)
            
            total_units = area_df['units'].sum()
            print(f"\nTotal off-site mitigation needed: {total_units:.2f} units")
            print(f"Expected: ~12.0 units (2.0 habitat residual + 10.0 headline)")
            
            # Allow some tolerance for floating point
            assert 11.5 < total_units < 12.5, f"Expected ~12 total units, got {total_units:.2f}"
            
            print("\n✅ Offsets applied correctly!")
            print("   - On-site surplus reduced habitat deficit")
            print("   - Headline Net Gain requirement calculated")
            print("   - Combined off-site mitigation calculated")
            
        return True
        
    except Exception as e:
        print(f"❌ Test failed: {e}")
        import traceback
        traceback.print_exc()
        return False


if __name__ == "__main__":
    success = test_with_offsets()
    exit(0 if success else 1)
