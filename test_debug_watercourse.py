"""
Debug test to understand why surpluses aren't offsetting Net Gain in user's case.
"""

import io
import pandas as pd
import openpyxl
from metric_reader import parse_metric_requirements


class MockUploadedFile:
    def __init__(self, buffer):
        self.buffer = buffer
        self.name = "test.xlsx"
    
    def read(self):
        return self.buffer.read()


def test_debug():
    wb = openpyxl.Workbook()
    
    # Headline Results
    ws_headline = wb.create_sheet("Headline Results")
    headers = ["Unit Type", "Target", "Baseline Units", "Units Required", "Unit Deficit"]
    for col, header in enumerate(headers, start=1):
        ws_headline.cell(row=5, column=col, value=header)
    
    data = [
        ["Habitat units", "10.00%", 7.71, 8.48, 3.38],
        ["Hedgerow units", "10.00%", 3.54, 3.89, 0.00],
        ["Watercourse units", "10.00%", 0.18, 0.19, 0.00],
    ]
    
    for row_idx, row_data in enumerate(data, start=6):
        for col_idx, value in enumerate(row_data, start=1):
            ws_headline.cell(row=row_idx, column=col_idx, value=value)
    
    # Trading Summary WaterCs
    ws_water = wb.create_sheet("Trading Summary WaterCs")
    headers_row = ["Habitat group", "On-site unit change", "Off-site unit change", "Project-wide unit change"]
    for col, header in enumerate(headers_row, start=1):
        ws_water.cell(row=1, column=col, value=header)
    
    current_row = 2
    
    # Medium Distinctiveness section
    ws_water.cell(row=current_row, column=1, value="Medium Distinctiveness")
    current_row += 1
    ws_water.cell(row=current_row, column=1, value="Habitat group")
    ws_water.cell(row=current_row, column=4, value="Project wide unit change")
    current_row += 1
    ws_water.cell(row=current_row, column=1, value="Ditches")
    ws_water.cell(row=current_row, column=2, value=0.06)
    ws_water.cell(row=current_row, column=3, value=0.00)
    ws_water.cell(row=current_row, column=4, value=0.06)
    current_row += 2
    
    # Low Distinctiveness section
    ws_water.cell(row=current_row, column=1, value="Low Distinctiveness")
    current_row += 1
    ws_water.cell(row=current_row, column=1, value="Habitat group")
    ws_water.cell(row=current_row, column=4, value="Project wide unit change")
    current_row += 1
    ws_water.cell(row=current_row, column=1, value="Culvert")
    ws_water.cell(row=current_row, column=2, value=0.01)
    ws_water.cell(row=current_row, column=3, value=0.00)
    ws_water.cell(row=current_row, column=4, value=0.01)
    
    if "Sheet" in wb.sheetnames:
        wb.remove(wb["Sheet"])
    
    excel_buffer = io.BytesIO()
    wb.save(excel_buffer)
    excel_buffer.seek(0)
    
    mock_file = MockUploadedFile(excel_buffer)
    
    print("=" * 80)
    print("DEBUG TEST")
    print("=" * 80)
    print("\nExpected:")
    print("  - Medium surplus: 0.06")
    print("  - Low surplus: 0.01")
    print("  - Net Gain needed: 0.018")
    print("  - Expected result: No requirements (covered by surplus)")
    print()
    
    requirements = parse_metric_requirements(mock_file)
    
    water_req = requirements["watercourses"]
    
    print("Actual result:")
    print(f"  - Watercourse requirements count: {len(water_req)}")
    
    if not water_req.empty:
        print("\n  Requirements:")
        for _, row in water_req.iterrows():
            print(f"    - {row['habitat']}: {row['units']:.6f} units")
        
        has_net_gain = any("Net Gain" in str(h) for h in water_req["habitat"])
        if has_net_gain:
            net_gain_row = water_req[water_req["habitat"].str.contains("Net Gain", na=False)]
            net_gain_units = net_gain_row["units"].iloc[0] if not net_gain_row.empty else 0.0
            print(f"\n  ❌ Net Gain NOT fully offset: {net_gain_units:.6f} units remaining")
            print(f"     (Expected: 0.00, Got: {net_gain_units:.6f})")
            return False
    else:
        print("    (empty - all covered)")
        print("\n  ✅ Net Gain fully offset")
        return True
    
    print("=" * 80)


if __name__ == "__main__":
    success = test_debug()
    if not success:
        print("\n❌ ISSUE REPRODUCED")
        exit(1)
    else:
        print("\n✅ WORKING CORRECTLY")
