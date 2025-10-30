"""
Demonstration script showing how to use the new baseline reading functionality
"""

import io
import openpyxl
from metric_reader import parse_metric_requirements


def create_demo_metric():
    """Create a demo BNG metric file with all three unit types"""
    wb = openpyxl.Workbook()
    
    # Create Headline Results sheet
    ws_headline = wb.create_sheet("Headline Results")
    
    # Add some header text
    ws_headline.cell(row=1, column=1, value="BNG Metric - Headline Results")
    ws_headline.cell(row=2, column=1, value="")
    ws_headline.cell(row=3, column=1, value="Scroll down for final results ⚠")
    ws_headline.cell(row=4, column=1, value="")
    
    # Add the target table header (starting at row 5)
    headers = ["Unit Type", "Target", "Baseline Units", "Units Required", "Unit Deficit"]
    for col, header in enumerate(headers, start=1):
        ws_headline.cell(row=5, column=col, value=header)
    
    # Add realistic test data for all three unit types
    # Format: [Unit Type, Target %, Baseline Units, Units Required, Unit Deficit]
    data = [
        ["Habitat units", "10.00%", 0.71, 0.78, 0.72],
        ["Hedgerow units", "10.00%", 2.50, 2.75, 0.25],
        ["Watercourse units", "10.00%", 1.20, 1.32, 0.12],
    ]
    
    for row_idx, row_data in enumerate(data, start=6):
        for col_idx, value in enumerate(row_data, start=1):
            ws_headline.cell(row=row_idx, column=col_idx, value=value)
    
    # Create a Trading Summary Area Habitats sheet
    ws_area = wb.create_sheet("Trading Summary Area Habitats")
    area_headers = ["Habitat", "Broad habitat", "Distinctiveness", "Project-wide unit change", "On-site unit change"]
    for col, header in enumerate(area_headers, start=1):
        ws_area.cell(row=1, column=col, value=header)
    
    # Sample area habitat data with deficits (negative values indicate loss)
    # Format: [Habitat name, Broad habitat, Distinctiveness, Project-wide change, On-site change]
    area_data = [
        ["Grassland", "Grassland and marsh", "Medium", -3.5, -1.0],  # 3.5 unit deficit
        ["Woodland", "Woodland and forest", "High", -2.0, -0.5],     # 2.0 unit deficit
    ]
    
    for row_idx, row_data in enumerate(area_data, start=2):
        for col_idx, value in enumerate(row_data, start=1):
            ws_area.cell(row=row_idx, column=col_idx, value=value)
    
    # Create Trading Summary Hedgerows sheet
    ws_hedge = wb.create_sheet("Trading Summary Hedgerows")
    hedge_headers = ["Habitat", "Broad habitat", "Distinctiveness", "Project-wide unit change"]
    for col, header in enumerate(hedge_headers, start=1):
        ws_hedge.cell(row=1, column=col, value=header)
    
    hedge_data = [
        ["Native hedgerow", "Hedgerow", "Medium", -0.5],
    ]
    
    for row_idx, row_data in enumerate(hedge_data, start=2):
        for col_idx, value in enumerate(row_data, start=1):
            ws_hedge.cell(row=row_idx, column=col_idx, value=value)
    
    # Remove default sheet
    if "Sheet" in wb.sheetnames:
        wb.remove(wb["Sheet"])
    
    # Save to BytesIO
    excel_buffer = io.BytesIO()
    wb.save(excel_buffer)
    excel_buffer.seek(0)
    
    return excel_buffer


def demo_baseline_reading():
    """Demonstrate how to access baseline information from the parsed metric"""
    
    print("=" * 70)
    print("BNG METRIC BASELINE READING DEMONSTRATION")
    print("=" * 70)
    
    # Create demo metric file
    metric_file = create_demo_metric()
    
    # Parse the metric
    print("\n📄 Parsing BNG metric file...")
    requirements = parse_metric_requirements(metric_file)
    
    # Access the baseline information
    baseline_info = requirements["baseline_info"]
    
    print("\n✅ Metric parsed successfully!")
    print("\n" + "=" * 70)
    print("BASELINE INFORMATION FROM HEADLINE RESULTS")
    print("=" * 70)
    
    for unit_type_key, display_name in [
        ("habitat", "Habitat Units"),
        ("hedgerow", "Hedgerow Units"),
        ("watercourse", "Watercourse Units")
    ]:
        info = baseline_info[unit_type_key]
        print(f"\n{display_name}:")
        print(f"  Target:           {info['target_percent']:.1%}")
        print(f"  Baseline Units:   {info['baseline_units']:.2f}")
        print(f"  Units Required:   {info['units_required']:.2f}")
        print(f"  Unit Deficit:     {info['unit_deficit']:.2f}")
    
    print("\n" + "=" * 70)
    print("REQUIREMENTS FROM TRADING SUMMARIES")
    print("=" * 70)
    
    print("\nArea Habitat Requirements:")
    if not requirements["area"].empty:
        for _, row in requirements["area"].iterrows():
            print(f"  - {row['habitat']}: {row['units']:.2f} units")
    else:
        print("  (none)")
    
    print("\nHedgerow Requirements:")
    if not requirements["hedgerows"].empty:
        for _, row in requirements["hedgerows"].iterrows():
            print(f"  - {row['habitat']}: {row['units']:.2f} units")
    else:
        print("  (none)")
    
    print("\nWatercourse Requirements:")
    if not requirements["watercourses"].empty:
        for _, row in requirements["watercourses"].iterrows():
            print(f"  - {row['habitat']}: {row['units']:.2f} units")
    else:
        print("  (none)")
    
    print("\n" + "=" * 70)
    print("EXAMPLE: Using baseline info for calculations")
    print("=" * 70)
    
    # Example calculation using baseline info
    for unit_type_key, display_name in [
        ("habitat", "Habitat"),
        ("hedgerow", "Hedgerow"),
        ("watercourse", "Watercourse")
    ]:
        info = baseline_info[unit_type_key]
        baseline = info["baseline_units"]
        target_pct = info["target_percent"]
        
        if baseline > 0:
            # Net Gain calculation: baseline units × target percentage
            # This represents the additional biodiversity units required beyond baseline
            net_gain_target = baseline * target_pct
            print(f"\n{display_name} Net Gain Target Calculation:")
            print(f"  Baseline: {baseline:.2f} units")
            print(f"  Target: {target_pct:.1%}")
            print(f"  Net Gain Required: {net_gain_target:.2f} units")
    
    print("\n" + "=" * 70)
    print("✅ Demonstration complete!")
    print("=" * 70)


if __name__ == "__main__":
    demo_baseline_reading()
