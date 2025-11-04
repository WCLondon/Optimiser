"""
Test to verify the exact scenario from the GitHub issue is resolved.

Issue description:
"Hedgerow surpluses are not being read by the metric reader on parsing and therefore 
not mitigating for downstream deficits. Please see example below where the optimiser 
included non-native and ornamental hedgerow even though it can be mitigated for upstream"

Example data:
- Very High: Species-rich native hedgerow with trees - associated with bank or ditch: 0.00
- High: Species-rich native hedgerow with trees: 0.37 ✓
- Medium: Species-rich native hedgerow: 0.13 ✓
- Low: Native hedgerow: 0.00
- Very Low: Non-native and ornamental hedgerow: -0.03 ⚠

Expected: The -0.03 deficit should be offset by the 0.37 + 0.13 = 0.50 surplus
"""

import io
import pandas as pd
import openpyxl
from metric_reader import parse_metric_requirements


class MockUploadedFile:
    """Mock file object for testing"""
    def __init__(self, buffer):
        self.buffer = buffer
        self.name = "test_metric.xlsx"
    
    def read(self):
        return self.buffer.read()


def test_exact_issue_scenario():
    """
    Test the exact scenario from the GitHub issue.
    
    The issue shows that when there are hedgerow surpluses (High: 0.37, Medium: 0.13),
    the Very Low deficit (Non-native and ornamental hedgerow: -0.03) should be offset
    by the surpluses, but it wasn't happening.
    """
    
    wb = openpyxl.Workbook()
    
    # Create Headline Results sheet
    ws_headline = wb.create_sheet("Headline Results")
    
    headers = ["Unit Type", "Target", "Baseline Units", "Units Required", "Unit Deficit"]
    for col, header in enumerate(headers, start=1):
        ws_headline.cell(row=5, column=col, value=header)
    
    # Using data that matches the issue (approximate baseline to get similar net gain)
    data = [
        ["Habitat units", "10.00%", 1.81, 2.00, 1.44],
        ["Hedgerow units", "10.00%", 0.94, 1.04, 1.04],
        ["Watercourse units", "10.00%", 0.00, 0.00, 0.00],
    ]
    
    for row_idx, row_data in enumerate(data, start=6):
        for col_idx, value in enumerate(row_data, start=1):
            ws_headline.cell(row=row_idx, column=col_idx, value=value)
    
    # Create Trading Summary Hedgerows sheet matching the issue data
    ws_hedge = wb.create_sheet("Trading Summary Hedgerows")
    
    hedge_headers = ["Habitat", "Distinctiveness", "Project-wide unit change"]
    for col, header in enumerate(hedge_headers, start=1):
        ws_hedge.cell(row=1, column=col, value=header)
    
    # Exact data from the issue
    hedge_data = [
        ["Species-rich native hedgerow with trees - associated with bank or ditch", "Very High", 0.00],
        ["Species-rich native hedgerow with trees", "High", 0.37],
        ["Species-rich native hedgerow", "Medium", 0.13],
        ["Native hedgerow", "Low", 0.00],
        ["Non-native and ornamental hedgerow", "Very Low", -0.03],
    ]
    
    for row_idx, row_data in enumerate(hedge_data, start=2):
        for col_idx, value in enumerate(row_data, start=1):
            ws_hedge.cell(row=row_idx, column=col_idx, value=value)
    
    # Remove default sheet
    if "Sheet" in wb.sheetnames:
        wb.remove(wb["Sheet"])
    
    # Save to BytesIO
    excel_buffer = io.BytesIO()
    wb.save(excel_buffer)
    excel_buffer.seek(0)
    
    # Create mock uploaded file
    mock_file = MockUploadedFile(excel_buffer)
    
    # Parse
    requirements = parse_metric_requirements(mock_file)
    
    # Display results
    print("\n" + "="*70)
    print("GitHub Issue Scenario Test")
    print("="*70)
    
    print("\nInput (from issue):")
    print("  Very High: Species-rich native hedgerow with trees - bank/ditch: 0.00")
    print("  High: Species-rich native hedgerow with trees: +0.37 ✓")
    print("  Medium: Species-rich native hedgerow: +0.13 ✓")
    print("  Low: Native hedgerow: 0.00")
    print("  Very Low: Non-native and ornamental hedgerow: -0.03 ⚠")
    print("\n  Total surplus: 0.37 + 0.13 = 0.50 units")
    print("  Total deficit: 0.03 units")
    print("  Net gain requirement: 0.94 × 10% = 0.094 units")
    
    hedge_req_df = requirements["hedgerows"]
    
    print("\nActual requirements:")
    if hedge_req_df.empty:
        print("  (empty - all requirements covered by on-site surplus)")
    else:
        print(hedge_req_df.to_string(index=False))
    
    print("\n" + "-"*70)
    print("Verification:")
    print("-"*70)
    
    # Check 1: Non-native and ornamental hedgerow should NOT be in requirements
    deficit_rows = hedge_req_df[hedge_req_df["habitat"].str.contains("Non-native", case=False, na=False)]
    if deficit_rows.empty:
        print("✅ Non-native and ornamental hedgerow deficit correctly offset")
        print("   (was showing as ⚠ in issue, now properly mitigated by surplus)")
    else:
        print("❌ ISSUE NOT FIXED: Non-native hedgerow still in requirements")
        print(f"   Units required: {deficit_rows.iloc[0]['units']}")
        return False
    
    # Check 2: Total surplus (0.50) should cover deficit (0.03) and net gain (0.094)
    total_surplus = 0.50
    total_need = 0.03 + 0.094  # 0.124
    
    if total_surplus >= total_need:
        if hedge_req_df.empty or hedge_req_df["units"].sum() < 0.001:
            print("✅ All hedgerow requirements covered by on-site surplus")
            print(f"   Surplus (0.50) > Need (0.124), no off-site requirements needed")
        else:
            print("❌ Unexpected requirements found")
            return False
    else:
        remaining_need = total_need - total_surplus
        total_req = hedge_req_df["units"].sum()
        if abs(total_req - remaining_need) < 0.001:
            print(f"✅ Remaining requirements correctly calculated: {total_req:.3f} units")
        else:
            print(f"❌ Requirements mismatch: expected {remaining_need:.3f}, got {total_req:.3f}")
            return False
    
    # Check 3: Verify the fix addresses the core issue
    print("✅ Core issue resolved: Hedgerow surpluses ARE being read and ARE")
    print("   mitigating downstream deficits as expected")
    
    print("\n" + "="*70)
    print("✅ ISSUE FIXED: Hedgerow surplus offsetting working correctly!")
    print("="*70)
    
    return True


if __name__ == "__main__":
    success = test_exact_issue_scenario()
    exit(0 if success else 1)
