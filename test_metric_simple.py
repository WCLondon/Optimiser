"""
Simple test - check if parsing works at all
"""

import io
import pandas as pd
import openpyxl
from metric_reader import parse_metric_requirements, normalise_requirements, build_band_map_from_raw


def test_basic_parsing():
    """Test basic parsing without complex logic"""
    wb = openpyxl.Workbook()
    
    # Trading Summary sheet
    ws = wb.create_sheet("Trading Summary Area Habitats")
    
    # Create section with distinctiveness header
    ws.cell(row=1, column=1, value="Medium distinctiveness habitat")
    ws.cell(row=2, column=1, value="")
    
    # Add data headers on row 3
    ws.cell(row=3, column=1, value="Habitat")
    ws.cell(row=3, column=2, value="Broad habitat")
    ws.cell(row=3, column=3, value="Project-wide unit change")
    ws.cell(row=3, column=4, value="On-site unit change")
    
    # Add one deficit row
    ws.cell(row=4, column=1, value="Grassland")
    ws.cell(row=4, column=2, value="Grassland and marsh")
    ws.cell(row=4, column=3, value=-5.0)
    ws.cell(row=4, column=4, value=0.0)
    
    # Headline Results sheet
    ws_h = wb.create_sheet("Headline Results")
    ws_h.cell(row=1, column=1, value="Unit Type")
    ws_h.cell(row=1, column=2, value="Baseline")
    ws_h.cell(row=1, column=3, value="Target")
    ws_h.cell(row=2, column=1, value="Area habitat units")
    ws_h.cell(row=2, column=2, value=50.0)
    ws_h.cell(row=2, column=3, value="10 %")
    
    if "Sheet" in wb.sheetnames:
        wb.remove(wb["Sheet"])
    
    excel_buffer = io.BytesIO()
    wb.save(excel_buffer)
    excel_buffer.seek(0)
    
    class MockFile:
        def __init__(self, buf):
            self.buffer = buf
            self.name = "test.xlsx"
        def read(self):
            return self.buffer.read()
    
    try:
        result = parse_metric_requirements(MockFile(excel_buffer))
        print(f"✅ Parsing succeeded")
        print(f"Area rows: {len(result['area'])}")
        print(f"Hedgerows rows: {len(result['hedgerows'])}")
        print(f"Watercourses rows: {len(result['watercourses'])}")
        
        if not result['area'].empty:
            print("\nArea requirements:")
            print(result['area'])
        
        return True
    except Exception as e:
        print(f"❌ Failed: {e}")
        import traceback
        traceback.print_exc()
        return False


if __name__ == "__main__":
    test_basic_parsing()
