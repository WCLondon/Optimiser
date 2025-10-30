"""
Test for baseline reading functionality from Headline Results sheet
"""

import io
import pandas as pd
import openpyxl
from metric_reader import parse_metric_requirements, parse_headline_all_unit_types, open_metric_workbook


def test_parse_headline_all_unit_types():
    """Test that parse_headline_all_unit_types reads all three unit types correctly"""
    
    # Create a simple mock Excel file in memory
    wb = openpyxl.Workbook()
    
    # Create Headline Results sheet with the target table
    ws = wb.create_sheet("Headline Results")
    
    # Add some content before the table (to simulate real metrics)
    ws.cell(row=1, column=1, value="Headline Results")
    ws.cell(row=2, column=1, value="")
    ws.cell(row=3, column=1, value="Scroll down for final results")
    ws.cell(row=4, column=1, value="")
    
    # Add the target table header (starting at row 5)
    headers = ["Unit Type", "Target", "Baseline Units", "Units Required", "Unit Deficit"]
    for col, header in enumerate(headers, start=1):
        ws.cell(row=5, column=col, value=header)
    
    # Add test data for all three unit types
    data = [
        ["Habitat units", "10.00%", 0.71, 0.78, 0.72],
        ["Hedgerow units", "10.00%", 0.00, 0.00, 0.00],
        ["Watercourse units", "10.00%", 0.00, 0.00, 0.00],
    ]
    
    for row_idx, row_data in enumerate(data, start=6):
        for col_idx, value in enumerate(row_data, start=1):
            ws.cell(row=row_idx, column=col_idx, value=value)
    
    # Remove default sheet
    if "Sheet" in wb.sheetnames:
        wb.remove(wb["Sheet"])
    
    # Save to BytesIO
    excel_buffer = io.BytesIO()
    wb.save(excel_buffer)
    excel_buffer.seek(0)
    
    # Open as ExcelFile
    xls = open_metric_workbook(excel_buffer)
    
    # Test parsing
    result = parse_headline_all_unit_types(xls)
    
    print("✅ parse_headline_all_unit_types succeeded")
    print("\nParsed baseline information:")
    for unit_type in ["habitat", "hedgerow", "watercourse"]:
        info = result[unit_type]
        print(f"\n{unit_type.title()}:")
        print(f"  Target: {info['target_percent']:.2%}")
        print(f"  Baseline Units: {info['baseline_units']:.2f}")
        print(f"  Units Required: {info['units_required']:.2f}")
        print(f"  Unit Deficit: {info['unit_deficit']:.2f}")
    
    # Verify the values
    assert result["habitat"]["target_percent"] == 0.10, f"Expected 0.10, got {result['habitat']['target_percent']}"
    assert result["habitat"]["baseline_units"] == 0.71, f"Expected 0.71, got {result['habitat']['baseline_units']}"
    assert result["habitat"]["units_required"] == 0.78, f"Expected 0.78, got {result['habitat']['units_required']}"
    assert result["habitat"]["unit_deficit"] == 0.72, f"Expected 0.72, got {result['habitat']['unit_deficit']}"
    
    assert result["hedgerow"]["target_percent"] == 0.10, f"Expected 0.10, got {result['hedgerow']['target_percent']}"
    assert result["hedgerow"]["baseline_units"] == 0.00, f"Expected 0.00, got {result['hedgerow']['baseline_units']}"
    
    assert result["watercourse"]["target_percent"] == 0.10, f"Expected 0.10, got {result['watercourse']['target_percent']}"
    assert result["watercourse"]["baseline_units"] == 0.00, f"Expected 0.00, got {result['watercourse']['baseline_units']}"
    
    print("\n✅ All baseline values verified correctly!")
    return True


def test_parse_metric_requirements_includes_baseline():
    """Test that parse_metric_requirements returns baseline_info"""
    
    # Create a mock Excel file with both Headline Results and Trading Summary
    wb = openpyxl.Workbook()
    
    # Create Headline Results sheet
    ws_headline = wb.create_sheet("Headline Results")
    
    # Add the target table header (starting at row 5)
    headers = ["Unit Type", "Target", "Baseline Units", "Units Required", "Unit Deficit"]
    for col, header in enumerate(headers, start=1):
        ws_headline.cell(row=5, column=col, value=header)
    
    # Add test data for all three unit types
    data = [
        ["Habitat units", "10.00%", 0.71, 0.78, 0.72],
        ["Hedgerow units", "10.00%", 0.50, 0.55, 0.05],
        ["Watercourse units", "10.00%", 0.30, 0.33, 0.03],
    ]
    
    for row_idx, row_data in enumerate(data, start=6):
        for col_idx, value in enumerate(row_data, start=1):
            ws_headline.cell(row=row_idx, column=col_idx, value=value)
    
    # Create a Trading Summary Area Habitats sheet
    ws_area = wb.create_sheet("Trading Summary Area Habitats")
    
    # Add headers
    area_headers = ["Habitat", "Broad habitat", "Distinctiveness", "Project-wide unit change", "On-site unit change"]
    for col, header in enumerate(area_headers, start=1):
        ws_area.cell(row=1, column=col, value=header)
    
    # Add some test data (deficits = negative values)
    area_data = [
        ["Grassland", "Grassland and marsh", "Medium", -2.0, 0.0],
    ]
    
    for row_idx, row_data in enumerate(area_data, start=2):
        for col_idx, value in enumerate(row_data, start=1):
            ws_area.cell(row=row_idx, column=col_idx, value=value)
    
    # Remove default sheet
    if "Sheet" in wb.sheetnames:
        wb.remove(wb["Sheet"])
    
    # Save to BytesIO
    excel_buffer = io.BytesIO()
    wb.save(excel_buffer)
    excel_buffer.seek(0)
    
    # Create a mock uploaded file object
    class MockUploadedFile:
        def __init__(self, buffer):
            self.buffer = buffer
            self.name = "test_metric.xlsx"
        
        def read(self):
            return self.buffer.read()
    
    mock_file = MockUploadedFile(excel_buffer)
    
    # Test parsing
    requirements = parse_metric_requirements(mock_file)
    
    print("✅ parse_metric_requirements succeeded")
    
    # Verify baseline_info is included
    assert "baseline_info" in requirements, "baseline_info key not found in results"
    
    baseline_info = requirements["baseline_info"]
    
    print("\nBaseline info returned:")
    for unit_type in ["habitat", "hedgerow", "watercourse"]:
        info = baseline_info[unit_type]
        print(f"\n{unit_type.title()}:")
        print(f"  Target: {info['target_percent']:.2%}")
        print(f"  Baseline Units: {info['baseline_units']:.2f}")
        print(f"  Units Required: {info['units_required']:.2f}")
        print(f"  Unit Deficit: {info['unit_deficit']:.2f}")
    
    # Verify the baseline values were read correctly
    assert baseline_info["habitat"]["baseline_units"] == 0.71, f"Expected 0.71, got {baseline_info['habitat']['baseline_units']}"
    assert baseline_info["hedgerow"]["baseline_units"] == 0.50, f"Expected 0.50, got {baseline_info['hedgerow']['baseline_units']}"
    assert baseline_info["watercourse"]["baseline_units"] == 0.30, f"Expected 0.30, got {baseline_info['watercourse']['baseline_units']}"
    
    print("\n✅ All baseline_info values verified correctly!")
    return True


if __name__ == "__main__":
    print("=" * 60)
    print("Test 1: parse_headline_all_unit_types")
    print("=" * 60)
    success1 = test_parse_headline_all_unit_types()
    
    print("\n" + "=" * 60)
    print("Test 2: parse_metric_requirements includes baseline_info")
    print("=" * 60)
    success2 = test_parse_metric_requirements_includes_baseline()
    
    print("\n" + "=" * 60)
    if success1 and success2:
        print("✅ ALL TESTS PASSED!")
    else:
        print("❌ SOME TESTS FAILED")
    print("=" * 60)
    
    exit(0 if (success1 and success2) else 1)
