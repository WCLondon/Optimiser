"""
Test parsing a metric file that has the summary columns showing trading is already satisfied.

The issue: The metric file shows:
- High surplus: 2.70 with "High Distinctiveness Units available to offset lower distinctiveness deficit" = 2.70 ✓
- Medium deficit: -1.59 with "Cumulative availability of units" = 1.11 ✓ (meaning it's covered)
- Very Low deficit: -0.25 with "Cumulative availability of units" = 0.87 ✓ (meaning it's covered)

But we're parsing the individual habitat deficits from the "Project-wide unit change" column,
which shows -0.55, -1.04, -0.25 with warnings.

The question: Should we trust the summary columns that say trading is satisfied,
or should we read the raw deficits and apply our own trading logic?

Current behavior: We read raw deficits and apply our own trading logic (which works correctly).
The issue might be a DISPLAY problem, not a calculation problem.
"""

import io
import pandas as pd
import openpyxl
from metric_reader import parse_metric_requirements


class MockUploadedFile:
    """Mock file object for testing"""
    def __init__(self, buffer):
        self.buffer = buffer
        self.name = "test_metric.xlsx"
    
    def read(self):
        return self.buffer.read()


def test_metric_with_summary_columns():
    """
    Test that we correctly handle a metric file that has summary columns.
    
    The metric file format shown by the user has extra columns like:
    - "High Distinctiveness Units available to offset lower distinctiveness deficit"
    - "Cumulative availability of units"
    
    These columns show that the metric itself has already calculated that deficits are covered.
    But we should still be reading the individual habitat values and applying our own logic.
    """
    
    wb = openpyxl.Workbook()
    
    # Create Headline Results sheet
    ws_headline = wb.create_sheet("Headline Results")
    
    headers = ["Unit Type", "Target", "Baseline Units", "Units Required", "Unit Deficit"]
    for col, header in enumerate(headers, start=1):
        ws_headline.cell(row=5, column=col, value=header)
    
    data = [
        ["Habitat units", "10.00%", 1.81, 2.00, 0.00],
        ["Hedgerow units", "10.00%", 1.84, 2.03, 0.00],
        ["Watercourse units", "10.00%", 0.00, 0.00, 0.00],
    ]
    
    for row_idx, row_data in enumerate(data, start=6):
        for col_idx, value in enumerate(row_data, start=1):
            ws_headline.cell(row=row_idx, column=col_idx, value=value)
    
    # Create Trading Summary Hedgerows sheet matching the user's format
    # with EXTRA summary columns that we should ignore
    ws_hedge = wb.create_sheet("Trading Summary Hedgerows")
    
    # Headers including summary columns
    hedge_headers = [
        "Habitat group",
        "On-site unit change", 
        "Off-site unit change",
        "Project-wide unit change",
        "",  # Empty column
        "High Distinctiveness Units available to offset lower distinctiveness deficit",
        "Cumulative availability of units"
    ]
    
    for col, header in enumerate(hedge_headers, start=1):
        ws_hedge.cell(row=1, column=col, value=header)
    
    # Section: Very High Distinctiveness
    ws_hedge.cell(row=2, column=1, value="Very High Distinctiveness")
    ws_hedge.cell(row=3, column=1, value="Habitat group")
    ws_hedge.cell(row=3, column=4, value="Project-wide unit change")
    ws_hedge.cell(row=4, column=1, value="Species-rich native hedgerow with trees - associated with bank or ditch")
    ws_hedge.cell(row=4, column=4, value=0.00)
    ws_hedge.cell(row=4, column=6, value=0.00)  # Summary column
    
    # Section: High Distinctiveness  
    ws_hedge.cell(row=6, column=1, value="High Distinctiveness")
    ws_hedge.cell(row=7, column=1, value="Habitat group")
    ws_hedge.cell(row=7, column=4, value="Project-wide unit change")
    ws_hedge.cell(row=8, column=1, value="Species-rich native hedgerow with trees")
    ws_hedge.cell(row=8, column=4, value=2.70)
    ws_hedge.cell(row=8, column=6, value=2.70)  # Summary: surplus available
    
    # Section: Medium Distinctiveness
    ws_hedge.cell(row=10, column=1, value="Medium Distinctiveness")
    ws_hedge.cell(row=11, column=1, value="Habitat group")
    ws_hedge.cell(row=11, column=4, value="Project-wide unit change")
    ws_hedge.cell(row=12, column=1, value="Species-rich native hedgerow")
    ws_hedge.cell(row=12, column=4, value=-0.55)
    ws_hedge.cell(row=13, column=1, value="Native hedgerow with trees")
    ws_hedge.cell(row=13, column=4, value=-1.04)
    ws_hedge.cell(row=13, column=7, value=1.11)  # Summary: cumulative availability (2.70 - 1.59 = 1.11)
    
    # Section: Very Low Distinctiveness
    ws_hedge.cell(row=15, column=1, value="Very Low Distinctiveness")
    ws_hedge.cell(row=16, column=1, value="Habitat group")
    ws_hedge.cell(row=16, column=4, value="Project-wide unit change")
    ws_hedge.cell(row=17, column=1, value="Non-native and ornamental hedgerow")
    ws_hedge.cell(row=17, column=4, value=-0.25)
    ws_hedge.cell(row=17, column=7, value=0.87)  # Summary: cumulative availability (1.11 - 0.25 = 0.87 + some rounding)
    
    # Remove default sheet
    if "Sheet" in wb.sheetnames:
        wb.remove(wb["Sheet"])
    
    # Save to BytesIO
    excel_buffer = io.BytesIO()
    wb.save(excel_buffer)
    excel_buffer.seek(0)
    
    # Create mock uploaded file
    mock_file = MockUploadedFile(excel_buffer)
    
    # Parse
    requirements = parse_metric_requirements(mock_file)
    
    print("\n" + "="*90)
    print("TEST: METRIC FILE WITH SUMMARY COLUMNS (SECTION HEADER FORMAT)")
    print("="*90)
    
    print("\n📊 METRIC FILE STRUCTURE:")
    print("-" * 90)
    print("The metric file has:")
    print("  1. Individual habitat rows with raw deficits (-0.55, -1.04, -0.25)")
    print("  2. Summary columns showing 'Cumulative availability' (1.11, 0.87)")
    print("  3. Section headers for distinctiveness (Very High, High, Medium, Very Low)")
    print()
    print("The summary columns indicate the metric itself calculated that deficits are covered.")
    print("But we parse individual habitat values and apply our own trading logic.")
    
    print("\n📋 EXPECTED BEHAVIOR:")
    print("-" * 90)
    print("Our parser should:")
    print("  1. Read individual habitat values from 'Project-wide unit change' column")
    print("  2. Extract distinctiveness from section headers")
    print("  3. Apply trading rules using apply_hedgerow_offsets()")
    print("  4. High surplus (2.70) should offset Medium deficits (-1.59) and Very Low deficit (-0.25)")
    print("  5. Result: requirements['hedgerows'] should be EMPTY or minimal")
    
    hedge_req_df = requirements["hedgerows"]
    
    print("\n📤 ACTUAL RESULT:")
    print("-" * 90)
    if hedge_req_df.empty:
        print("✅ requirements['hedgerows'] is EMPTY")
        print()
        print("SUCCESS: Our parser correctly handled the metric file with summary columns!")
        print("         - Ignored summary columns")
        print("         - Read individual habitat values")
        print("         - Applied our own trading logic")
        print("         - All deficits offset by surplus")
        success = True
    else:
        print("❌ requirements['hedgerows'] is NOT EMPTY:")
        print()
        print(hedge_req_df.to_string(index=False))
        print()
        print("ISSUE: Deficits are showing up despite summary columns saying they're covered.")
        success = False
    
    print("\n" + "="*90)
    
    return success


if __name__ == "__main__":
    success = test_metric_with_summary_columns()
    exit(0 if success else 1)
