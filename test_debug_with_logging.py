"""
Add debug logging to metric_reader to understand watercourse surplus allocation.
"""

import io
import pandas as pd
import openpyxl


# Patch metric_reader to add debug output
import metric_reader

original_parse = metric_reader.parse_metric_requirements


def debug_parse_metric_requirements(uploaded_file):
    """Wrapper that adds debug output"""
    result = original_parse(uploaded_file)
    
    # Check if we have watercourse data
    if "watercourses" in result and not result["watercourses"].empty:
        print("\n" + "=" * 80)
        print("DEBUG: Watercourse Parsing")
        print("=" * 80)
        
        water_req = result["watercourses"]
        baseline_info = result.get("baseline_info", {})
        
        if "watercourse" in baseline_info:
            wc_info = baseline_info["watercourse"]
            print(f"\nBaseline info:")
            print(f"  - Baseline units: {wc_info.get('baseline_units', 0):.6f}")
            print(f"  - Target percent: {wc_info.get('target_percent', 0):.6f}")
            print(f"  - Net Gain required: {wc_info.get('baseline_units', 0) * wc_info.get('target_percent', 0):.6f}")
        
        print(f"\nWatercourse requirements:")
        for _, row in water_req.iterrows():
            print(f"  - {row['habitat']}: {row['units']:.6f} units")
        
        print("=" * 80 + "\n")
    
    return result


# Monkey patch for debugging
metric_reader.parse_metric_requirements = debug_parse_metric_requirements


class MockUploadedFile:
    def __init__(self, buffer):
        self.buffer = buffer
        self.name = "test.xlsx"
    
    def read(self):
        return self.buffer.read()


def create_test_metric():
    """Create test metric matching user's exact format"""
    wb = openpyxl.Workbook()
    
    # Headline Results
    ws_headline = wb.create_sheet("Headline Results")
    headers = ["Unit Type", "Target", "Baseline Units", "Units Required", "Unit Deficit"]
    for col, header in enumerate(headers, start=1):
        ws_headline.cell(row=5, column=col, value=header)
    
    data = [
        ["Habitat units", "10.00%", 7.71, 8.48, 3.38],
        ["Hedgerow units", "10.00%", 3.54, 3.89, 0.00],
        ["Watercourse units", "10.00%", 0.18, 0.19, 0.00],
    ]
    
    for row_idx, row_data in enumerate(data, start=6):
        for col_idx, value in enumerate(row_data, start=1):
            ws_headline.cell(row=row_idx, column=col_idx, value=value)
    
    # Trading Summary WaterCs - exact user format
    ws_water = wb.create_sheet("Trading Summary WaterCs")
    
    # Row 1: Headers
    ws_water.cell(row=1, column=1, value="Habitat group")
    ws_water.cell(row=1, column=2, value="On-site unit change")
    ws_water.cell(row=1, column=3, value="Off-site unit change")
    ws_water.cell(row=1, column=4, value="Project-wide unit change")
    
    # Medium Distinctiveness section
    ws_water.cell(row=2, column=1, value="Medium Distinctiveness")
    ws_water.cell(row=3, column=1, value="Habitat group")
    ws_water.cell(row=3, column=4, value="Project wide unit change")
    ws_water.cell(row=4, column=1, value="Ditches")
    ws_water.cell(row=4, column=2, value=0.06)
    ws_water.cell(row=4, column=3, value=0.00)
    ws_water.cell(row=4, column=4, value=0.06)
    
    # Low Distinctiveness section
    ws_water.cell(row=6, column=1, value="Low Distinctiveness")
    ws_water.cell(row=7, column=1, value="Habitat group")
    ws_water.cell(row=7, column=4, value="Project wide unit change")
    ws_water.cell(row=8, column=1, value="Culvert")
    ws_water.cell(row=8, column=2, value=0.01)
    ws_water.cell(row=8, column=3, value=0.00)
    ws_water.cell(row=8, column=4, value=0.01)
    
    if "Sheet" in wb.sheetnames:
        wb.remove(wb["Sheet"])
    
    excel_buffer = io.BytesIO()
    wb.save(excel_buffer)
    excel_buffer.seek(0)
    
    return MockUploadedFile(excel_buffer)


if __name__ == "__main__":
    print("\nTesting with user's exact format...")
    mock_file = create_test_metric()
    
    requirements = metric_reader.parse_metric_requirements(mock_file)
    
    water_req = requirements["watercourses"]
    
    if not water_req.empty:
        has_net_gain = any("Net Gain" in str(h) for h in water_req["habitat"])
        if has_net_gain:
            print("\n❌ ISSUE: Net Gain (Watercourses) still in requirements!")
            print("   Surpluses are NOT offsetting Net Gain properly.")
        else:
            print("\n✅ No Net Gain in requirements (but other deficits remain)")
    else:
        print("\n✅ All requirements covered (empty)")
