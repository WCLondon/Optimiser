"""
Debug test to check what distinctiveness values are being extracted
"""

import io
import pandas as pd
import openpyxl
from metric_reader import parse_metric_requirements, normalise_requirements, open_metric_workbook


class MockUploadedFile:
    """Mock file object for testing"""
    def __init__(self, buffer):
        self.buffer = buffer
        self.name = "test_metric.xlsx"
    
    def read(self):
        return self.buffer.read()


def test_distinctiveness_extraction():
    """
    Debug: Check what distinctiveness values are extracted
    """
    
    wb = openpyxl.Workbook()
    
    # Create Headline Results sheet
    ws_headline = wb.create_sheet("Headline Results")
    
    headers = ["Unit Type", "Target", "Baseline Units", "Units Required", "Unit Deficit"]
    for col, header in enumerate(headers, start=1):
        ws_headline.cell(row=5, column=col, value=header)
    
    data = [
        ["Habitat units", "10.00%", 1.81, 2.00, 0.00],
        ["Hedgerow units", "10.00%", 1.84, 2.03, 0.00],
        ["Watercourse units", "10.00%", 0.00, 0.00, 0.00],
    ]
    
    for row_idx, row_data in enumerate(data, start=6):
        for col_idx, value in enumerate(row_data, start=1):
            ws_headline.cell(row=row_idx, column=col_idx, value=value)
    
    # Create Trading Summary Hedgerows sheet with exact data from issue
    ws_hedge = wb.create_sheet("Trading Summary Hedgerows")
    
    hedge_headers = ["Habitat", "Distinctiveness", "Project-wide unit change"]
    for col, header in enumerate(hedge_headers, start=1):
        ws_hedge.cell(row=1, column=col, value=header)
    
    # Exact data from the issue
    hedge_data = [
        ["Species-rich native hedgerow with trees - associated with bank or ditch", "Very High", 0.00],
        ["Species-rich native hedgerow with trees", "High", 2.70],
        ["Species-rich native hedgerow", "Medium", -0.55],
        ["Native hedgerow with trees", "Medium", -1.04],
        ["Non-native and ornamental hedgerow", "Very Low", -0.25],
    ]
    
    for row_idx, row_data in enumerate(hedge_data, start=2):
        for col_idx, value in enumerate(row_data, start=1):
            ws_hedge.cell(row=row_idx, column=col_idx, value=value)
    
    # Remove default sheet
    if "Sheet" in wb.sheetnames:
        wb.remove(wb["Sheet"])
    
    # Save to BytesIO
    excel_buffer = io.BytesIO()
    wb.save(excel_buffer)
    excel_buffer.seek(0)
    
    # Create mock uploaded file
    mock_file = MockUploadedFile(excel_buffer)
    
    # Open workbook and parse normalised data
    xls = open_metric_workbook(mock_file)
    
    HEDGE_SHEETS = [
        "Trading Summary Hedgerows",
        "Hedgerows Trading Summary",
        "Hedgerow Trading Summary",
        "Trading Summary (Hedgerows)"
    ]
    
    hedge_norm, colmap, sheet_name = normalise_requirements(xls, HEDGE_SHEETS, "Hedgerows")
    
    print("\n" + "="*80)
    print("DISTINCTIVENESS EXTRACTION DEBUG")
    print("="*80)
    
    print(f"\nFound sheet: {sheet_name}")
    print(f"Column mapping: {colmap}")
    
    print("\nNormalized hedgerow data:")
    print(hedge_norm[["habitat", "distinctiveness", "project_wide_change"]])
    
    print("\nDistinctiveness values extracted:")
    for idx, row in hedge_norm.iterrows():
        habitat = row["habitat"]
        dist = row["distinctiveness"]
        change = row["project_wide_change"]
        print(f"  {habitat[:50]:50s} | {str(dist):12s} | {change:+7.2f}")
    
    # Check for NA values
    na_count = hedge_norm["distinctiveness"].isna().sum()
    print(f"\nNumber of rows with NA distinctiveness: {na_count}")
    
    if na_count > 0:
        print("❌ WARNING: Some rows have NA distinctiveness!")
        print("Rows with NA distinctiveness:")
        na_rows = hedge_norm[hedge_norm["distinctiveness"].isna()]
        print(na_rows[["habitat", "distinctiveness", "project_wide_change"]])
    else:
        print("✅ All rows have distinctiveness values")
    
    print("\n" + "="*80)


if __name__ == "__main__":
    test_distinctiveness_extraction()
