"""
Test with the EXACT format from the GitHub issue.

The issue shows a Trading Summary Hedgerows tab with this structure:
- Section headers like "Medium Distinctiveness"
- Column headers: "Habitat group", "On-site unit change", "Off-site unit change", "Project-wide unit change"
- Multiple habitat rows under each distinctiveness section
- Most are 0.00 except:
  - Medium: Species-rich native hedgerow: 1.47
  - Low: Native hedgerow: -0.70

The key insight: The column is "Habitat group" not just "Habitat"!
This might affect parsing.
"""

import io
import pandas as pd
import openpyxl
from metric_reader import parse_metric_requirements


class MockUploadedFile:
    """Mock file object for testing"""
    def __init__(self, buffer):
        self.buffer = buffer
        self.name = "test_metric.xlsx"
    
    def read(self):
        return self.buffer.read()


def test_exact_issue_format():
    """
    Test with exact format from issue - note "Habitat group" column name
    """
    wb = openpyxl.Workbook()
    
    # Create Headline Results sheet  
    ws_headline = wb.create_sheet("Headline Results")
    
    headers = ["Unit Type", "Target", "Baseline Units", "Units Required", "Unit Deficit"]
    for col, header in enumerate(headers, start=1):
        ws_headline.cell(row=5, column=col, value=header)
    
    data = [
        ["Habitat units", "10.00%", 10.00, 11.00, 0.00],
        ["Hedgerow units", "10.00%", 0.00, 0.00, 0.00],
        ["Watercourse units", "10.00%", 0.00, 0.00, 0.00],
    ]
    
    for row_idx, row_data in enumerate(data, start=6):
        for col_idx, value in enumerate(row_data, start=1):
            ws_headline.cell(row=row_idx, column=col_idx, value=value)
    
    # Create Trading Summary Hedgerows sheet with EXACT format from issue
    ws_hedge = wb.create_sheet("Trading Summary Hedgerows")
    
    row = 1
    
    # Very High Distinctiveness section
    ws_hedge.cell(row=row, column=1, value="Very High Distinctiveness")
    row += 1
    ws_hedge.cell(row=row, column=1, value="Habitat group")  # Note: "Habitat group" not "Habitat"
    ws_hedge.cell(row=row, column=2, value="On-site unit change")
    ws_hedge.cell(row=row, column=3, value="Off-site unit change")
    ws_hedge.cell(row=row, column=4, value="Project-wide unit change")
    row += 1
    ws_hedge.cell(row=row, column=1, value="Species-rich native hedgerow with trees - associated with bank or ditch")
    ws_hedge.cell(row=row, column=2, value=0.00)
    ws_hedge.cell(row=row, column=3, value=0.00)
    ws_hedge.cell(row=row, column=4, value=0.00)
    row += 1
    ws_hedge.cell(row=row, column=1, value="")  # Total row
    ws_hedge.cell(row=row, column=2, value=0.00)
    ws_hedge.cell(row=row, column=3, value=0.00)
    ws_hedge.cell(row=row, column=4, value=0.00)
    row += 2
    
    # High Distinctiveness section
    ws_hedge.cell(row=row, column=1, value="High Distinctiveness")
    row += 1
    ws_hedge.cell(row=row, column=1, value="Habitat group")
    ws_hedge.cell(row=row, column=2, value="On-site unit change")
    ws_hedge.cell(row=row, column=3, value="Off-site unit change")
    ws_hedge.cell(row=row, column=4, value="Project-wide unit change")
    row += 1
    ws_hedge.cell(row=row, column=1, value="Species-rich native hedgerow with trees")
    ws_hedge.cell(row=row, column=2, value=0.00)
    ws_hedge.cell(row=row, column=3, value=0.00)
    ws_hedge.cell(row=row, column=4, value=0.00)
    row += 1
    ws_hedge.cell(row=row, column=1, value="Species-rich native hedgerow - associated with bank or ditch")
    ws_hedge.cell(row=row, column=2, value=0.00)
    ws_hedge.cell(row=row, column=3, value=0.00)
    ws_hedge.cell(row=row, column=4, value=0.00)
    row += 1
    ws_hedge.cell(row=row, column=1, value="Native hedgerow with trees - associated with bank or ditch")
    ws_hedge.cell(row=row, column=2, value=0.00)
    ws_hedge.cell(row=row, column=3, value=0.00)
    ws_hedge.cell(row=row, column=4, value=0.00)
    row += 1
    ws_hedge.cell(row=row, column=1, value="")  # Total row
    ws_hedge.cell(row=row, column=2, value=0.00)
    ws_hedge.cell(row=row, column=3, value=0.00)
    ws_hedge.cell(row=row, column=4, value=0.00)
    row += 2
    
    # Medium Distinctiveness section - THIS IS THE KEY SURPLUS
    ws_hedge.cell(row=row, column=1, value="Medium Distinctiveness")
    row += 1
    ws_hedge.cell(row=row, column=1, value="Habitat group")
    ws_hedge.cell(row=row, column=2, value="On-site unit change")
    ws_hedge.cell(row=row, column=3, value="Off-site unit change")
    ws_hedge.cell(row=row, column=4, value="Project-wide unit change")
    row += 1
    ws_hedge.cell(row=row, column=1, value="Species-rich native hedgerow")
    ws_hedge.cell(row=row, column=2, value=1.47)
    ws_hedge.cell(row=row, column=3, value=0.00)
    ws_hedge.cell(row=row, column=4, value=1.47)
    row += 1
    ws_hedge.cell(row=row, column=1, value="Native hedgerow - associated with bank or ditch")
    ws_hedge.cell(row=row, column=2, value=0.00)
    ws_hedge.cell(row=row, column=3, value=0.00)
    ws_hedge.cell(row=row, column=4, value=0.00)
    row += 1
    ws_hedge.cell(row=row, column=1, value="Native hedgerow with trees")
    ws_hedge.cell(row=row, column=2, value=0.00)
    ws_hedge.cell(row=row, column=3, value=0.00)
    ws_hedge.cell(row=row, column=4, value=0.00)
    row += 1
    ws_hedge.cell(row=row, column=1, value="Ecologically valuable line of trees")
    ws_hedge.cell(row=row, column=2, value=0.00)
    ws_hedge.cell(row=row, column=3, value=0.00)
    ws_hedge.cell(row=row, column=4, value=0.00)
    row += 1
    ws_hedge.cell(row=row, column=1, value="Ecologically valuable line of trees - associated with bank or ditch")
    ws_hedge.cell(row=row, column=2, value=0.00)
    ws_hedge.cell(row=row, column=3, value=0.00)
    ws_hedge.cell(row=row, column=4, value=0.00)
    row += 1
    ws_hedge.cell(row=row, column=1, value="")  # Total row
    ws_hedge.cell(row=row, column=2, value=1.47)
    ws_hedge.cell(row=row, column=3, value=0.00)
    ws_hedge.cell(row=row, column=4, value=1.47)
    row += 2
    
    # Low Distinctiveness section - THIS IS THE KEY DEFICIT
    ws_hedge.cell(row=row, column=1, value="Low Distinctiveness")
    row += 1
    ws_hedge.cell(row=row, column=1, value="Habitat group")
    ws_hedge.cell(row=row, column=2, value="On-site unit change")
    ws_hedge.cell(row=row, column=3, value="Off-site unit change")
    ws_hedge.cell(row=row, column=4, value="Project-wide unit change")
    row += 1
    ws_hedge.cell(row=row, column=1, value="Native hedgerow")
    ws_hedge.cell(row=row, column=2, value=-0.70)
    ws_hedge.cell(row=row, column=3, value=0.00)
    ws_hedge.cell(row=row, column=4, value=-0.70)
    row += 1
    ws_hedge.cell(row=row, column=1, value="Line of trees")
    ws_hedge.cell(row=row, column=2, value=0.00)
    ws_hedge.cell(row=row, column=3, value=0.00)
    ws_hedge.cell(row=row, column=4, value=0.00)
    row += 1
    ws_hedge.cell(row=row, column=1, value="Line of trees - associated with bank or ditch")
    ws_hedge.cell(row=row, column=2, value=0.00)
    ws_hedge.cell(row=row, column=3, value=0.00)
    ws_hedge.cell(row=row, column=4, value=0.00)
    row += 1
    ws_hedge.cell(row=row, column=1, value="")  # Total row
    ws_hedge.cell(row=row, column=2, value=-0.70)
    ws_hedge.cell(row=row, column=3, value=0.00)
    ws_hedge.cell(row=row, column=4, value=-0.70)
    row += 2
    
    # Very Low Distinctiveness section
    ws_hedge.cell(row=row, column=1, value="Very Low Distinctiveness")
    row += 1
    ws_hedge.cell(row=row, column=1, value="Habitat group")
    ws_hedge.cell(row=row, column=2, value="On-site unit change")
    ws_hedge.cell(row=row, column=3, value="Off-site unit change")
    ws_hedge.cell(row=row, column=4, value="Project-wide unit change")
    row += 1
    ws_hedge.cell(row=row, column=1, value="Non-native and ornamental hedgerow")
    ws_hedge.cell(row=row, column=2, value=0.00)
    ws_hedge.cell(row=row, column=3, value=0.00)
    ws_hedge.cell(row=row, column=4, value=0.00)
    row += 1
    ws_hedge.cell(row=row, column=1, value="")  # Total row
    ws_hedge.cell(row=row, column=2, value=0.00)
    ws_hedge.cell(row=row, column=3, value=0.00)
    ws_hedge.cell(row=row, column=4, value=0.00)
    
    # Remove default sheet
    if "Sheet" in wb.sheetnames:
        wb.remove(wb["Sheet"])
    
    # Save to BytesIO
    excel_buffer = io.BytesIO()
    wb.save(excel_buffer)
    excel_buffer.seek(0)
    
    # Create mock uploaded file
    mock_file = MockUploadedFile(excel_buffer)
    
    # Parse
    print("\n" + "="*80)
    print("PARSING WITH EXACT FORMAT FROM ISSUE")
    print("="*80)
    print("\nKey details:")
    print("  - Column name: 'Habitat group' (not 'Habitat')")
    print("  - Section headers: 'Medium Distinctiveness', 'Low Distinctiveness', etc.")
    print("  - Total rows with empty habitat name after each section")
    print("\nExpected behavior:")
    print("  - Medium: Species-rich native hedgerow: +1.47 (surplus)")
    print("  - Low: Native hedgerow: -0.70 (deficit)")
    print("  - Medium surplus SHOULD offset Low deficit")
    print("  - Result: NO Native hedgerow in requirements")
    
    requirements = parse_metric_requirements(mock_file)
    
    # Verify results
    hedge_req_df = requirements["hedgerows"]
    
    print("\n" + "-"*80)
    print("PARSED REQUIREMENTS:")
    print("-"*80)
    if hedge_req_df.empty:
        print("  (empty - all requirements covered by on-site surplus)")
    else:
        print(hedge_req_df.to_string(index=False))
        print(f"\n  Total units: {hedge_req_df['units'].sum():.6f}")
    
    # Check if Native hedgerow appears
    native_rows = hedge_req_df[
        hedge_req_df["habitat"].str.contains("Native hedgerow", case=False, na=False) & 
        ~hedge_req_df["habitat"].str.contains("Species-rich", case=False, na=False) &
        ~hedge_req_df["habitat"].str.contains("with trees", case=False, na=False) &
        ~hedge_req_df["habitat"].str.contains("associated", case=False, na=False)
    ]
    
    print("\n" + "="*80)
    if not native_rows.empty:
        print("❌ ISSUE REPRODUCED!")
        print(f"   'Native hedgerow' appears with {native_rows.iloc[0]['units']:.6f} units")
        print("   This is the bug - it should be offset by Medium surplus!")
        return False
    else:
        print("✅ WORKING CORRECTLY!")
        print("   'Native hedgerow' properly offset by Medium surplus")
        return True


if __name__ == "__main__":
    success = test_exact_issue_format()
    exit(0 if success else 1)
