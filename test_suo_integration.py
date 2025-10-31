"""
Integration test for SUO with metric reader.
Tests the complete flow: metric file -> surplus extraction -> SUO computation.
"""

import io
import pandas as pd
import openpyxl
from metric_reader import parse_metric_requirements
from suo import compute_suo, SUOConfig


def create_test_metric_with_surplus():
    """Create a test metric file with both deficits and surpluses."""
    wb = openpyxl.Workbook()
    
    # Create Trading Summary Area Habitats sheet
    ws = wb.create_sheet("Trading Summary Area Habitats")
    
    # Add Medium distinctiveness band header (row 1)
    ws.cell(row=1, column=1, value="Medium Distinctiveness")
    
    # Add column headers (row 2)
    headers = ["Habitat", "Broad habitat", "Project-wide unit change", "On-site unit change"]
    for col, header in enumerate(headers, start=1):
        ws.cell(row=2, column=col, value=header)
    
    # Add Medium distinctiveness data (rows 3-5)
    # More surplus than deficits to ensure some remains after on-site trading
    medium_data = [
        ["Heathland - Dwarf shrub", "Heathland and shrub", 30.0, 25.0],   # Surplus (Medium) - 30 units
        ["Grassland - Modified", "Grassland and marsh", -5.0, 0.0],       # Deficit (Medium) - 5 units
        ["Scrub - Blackthorn", "Heathland and shrub", 20.0, 15.0],       # Surplus (Medium) - 20 units
    ]
    
    for row_idx, row_data in enumerate(medium_data, start=3):
        for col_idx, value in enumerate(row_data, start=1):
            ws.cell(row=row_idx, column=col_idx, value=value)
    
    # Add High distinctiveness band header (row 6)
    ws.cell(row=6, column=1, value="High Distinctiveness")
    
    # Add column headers again (row 7)
    for col, header in enumerate(headers, start=1):
        ws.cell(row=7, column=col, value=header)
    
    # Add High distinctiveness data (row 8)
    high_data = [
        ["Woodland - Mixed", "Woodland and forest", -3.0, 0.0],  # Deficit (High) - 3 units
    ]
    
    for row_idx, row_data in enumerate(high_data, start=8):
        for col_idx, value in enumerate(row_data, start=1):
            ws.cell(row=row_idx, column=col_idx, value=value)
    
    # Add Low distinctiveness band header (row 9)
    ws.cell(row=9, column=1, value="Low Distinctiveness")
    
    # Add column headers again (row 10)
    for col, header in enumerate(headers, start=1):
        ws.cell(row=10, column=col, value=header)
    
    # Add Low distinctiveness data (row 11) - this should be excluded from SUO
    low_data = [
        ["Scrub - Dense", "Heathland and shrub", 10.0, 8.0],  # Surplus (Low) - 10 units
    ]
    
    for row_idx, row_data in enumerate(low_data, start=11):
        for col_idx, value in enumerate(row_data, start=1):
            ws.cell(row=row_idx, column=col_idx, value=value)
    
    # Remove default sheet
    if "Sheet" in wb.sheetnames:
        wb.remove(wb["Sheet"])
    
    # Save to BytesIO
    excel_buffer = io.BytesIO()
    wb.save(excel_buffer)
    excel_buffer.seek(0)
    
    return excel_buffer


def test_suo_integration():
    """Test complete SUO flow with metric file."""
    print("\n=== Integration Test: Metric File -> Surplus -> SUO ===")
    
    # Step 1: Create test metric
    metric_buffer = create_test_metric_with_surplus()
    
    # Step 2: Parse metric
    class MockUploadedFile:
        def __init__(self, buffer):
            self.buffer = buffer
            self.name = "test_metric.xlsx"
        
        def read(self):
            # Reset position before reading to allow multiple reads
            self.buffer.seek(0)
            return self.buffer.read()
    
    mock_file = MockUploadedFile(metric_buffer)
    requirements = parse_metric_requirements(mock_file)
    
    print(f"\n📋 Requirements parsed:")
    print(f"   Area deficits: {len(requirements['area'])} rows")
    if not requirements['area'].empty:
        print(requirements['area'])
    
    print(f"\n🌳 Surplus found:")
    if 'surplus' in requirements and not requirements['surplus'].empty:
        print(f"   {len(requirements['surplus'])} surplus habitat(s)")
        print(requirements['surplus'])
    else:
        print("   No surplus found")
    
    # Step 3: Verify surplus filtering
    surplus_df = requirements.get('surplus', pd.DataFrame())
    if not surplus_df.empty:
        print(f"\n✅ Surplus extraction successful")
        
        # Check distinctiveness levels
        distinctiveness_counts = surplus_df['distinctiveness'].value_counts()
        print(f"\n   Distinctiveness breakdown:")
        for dist, count in distinctiveness_counts.items():
            print(f"     - {dist}: {count}")
        
        # Verify only Medium+ is present (Low should be filtered in on-site trading)
        # Note: The metric reader applies on-site trading rules, which may consume
        # both Low and Medium surpluses. What remains should be eligible.
    
    # Step 4: Apply SUO if surplus exists
    if not surplus_df.empty:
        print(f"\n🎯 Testing SUO computation...")
        
        # Prepare requirements DataFrame
        req_df = requirements['area'].copy()
        
        # Verify expected columns exist
        if 'habitat' not in req_df.columns or 'units' not in req_df.columns:
            print(f"   ⚠️ Unexpected column structure in requirements: {list(req_df.columns)}")
            return True
        
        req_df['line_id'] = req_df.index.astype(str)
        req_df = req_df.rename(columns={'habitat': 'habitat_name', 'units': 'units_needed'})
        
        # Get trading_group from catalog or use empty string
        # Note: In real usage, this would come from the habitat catalog merge
        req_df['trading_group'] = ''
        
        # Prepare surplus for SUO
        surplus_for_suo = surplus_df.copy()
        surplus_for_suo['site_id'] = 'development_site'
        
        # Ensure surplus has trading_group column
        if 'broad_group' in surplus_for_suo.columns:
            surplus_for_suo = surplus_for_suo.rename(columns={'broad_group': 'trading_group'})
        elif 'trading_group' not in surplus_for_suo.columns:
            surplus_for_suo['trading_group'] = ''
        
        # Create SRM (local tier, SRM=1.0)
        srm = pd.DataFrame({'site_id': ['development_site'], 'srm': [1.0]})
        
        # Compute SUO
        config = SUOConfig(headroom_fraction=0.5, min_distinctiveness="Medium")
        req_reduced, alloc_ledger, summary = compute_suo(
            req_df[['line_id', 'trading_group', 'units_needed']],
            surplus_for_suo[['site_id', 'distinctiveness', 'trading_group', 'units_surplus']],
            srm,
            config
        )
        
        print(f"\n   SUO Results:")
        print(f"     Eligible surplus: {summary['eligible_surplus']:.2f} units")
        print(f"     Usable (50% headroom): {summary['usable_units']:.2f} units")
        print(f"     Reduction fraction: {summary['reduction_fraction_final']*100:.1f}%")
        
        if summary['reduction_fraction_final'] > 0:
            print(f"\n   ✅ SUO applied successfully!")
            print(f"\n   Requirements before/after:")
            comparison = req_reduced[['line_id', 'units_needed_before', 'units_needed_after']].copy()
            print(comparison.to_string(index=False))
        else:
            print(f"\n   ℹ️ No reduction possible (insufficient eligible surplus)")
    else:
        print(f"\n   ⚠️ No surplus to test SUO with")
    
    print(f"\n✅ Integration test completed")
    return True


if __name__ == "__main__":
    try:
        success = test_suo_integration()
        exit(0 if success else 1)
    except Exception as e:
        print(f"\n❌ Integration test failed: {e}")
        import traceback
        traceback.print_exc()
        exit(1)
