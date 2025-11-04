"""
Comprehensive test to verify the full workflow including demand population
"""

import io
import pandas as pd
import openpyxl
from metric_reader import parse_metric_requirements


class MockUploadedFile:
    """Mock file object for testing"""
    def __init__(self, buffer):
        self.buffer = buffer
        self.name = "test_metric.xlsx"
    
    def read(self):
        return self.buffer.read()


def test_full_workflow_with_high_surplus_offsetting_medium_and_very_low_deficits():
    """
    Test the EXACT scenario from the GitHub issue to verify surplus offsetting works end-to-end
    
    Scenario:
    - High: Species-rich native hedgerow with trees: +2.70 (surplus)
    - Medium: Species-rich native hedgerow: -0.55 (deficit)
    - Medium: Native hedgerow with trees: -1.04 (deficit)
    - Very Low: Non-native and ornamental hedgerow: -0.25 (deficit)
    
    Total surplus: 2.70
    Total deficits: 1.84
    Net gain: 10% of 1.84 = 0.184
    
    Expected: High surplus (2.70) should offset all deficits (1.84) and net gain (0.184)
    Result: No requirements should appear (or minimal remaining)
    """
    
    wb = openpyxl.Workbook()
    
    # Create Headline Results sheet
    ws_headline = wb.create_sheet("Headline Results")
    
    headers = ["Unit Type", "Target", "Baseline Units", "Units Required", "Unit Deficit"]
    for col, header in enumerate(headers, start=1):
        ws_headline.cell(row=5, column=col, value=header)
    
    data = [
        ["Habitat units", "10.00%", 1.81, 2.00, 0.00],
        ["Hedgerow units", "10.00%", 1.84, 2.03, 0.00],  # baseline=1.84, 10% net gain = 0.184
        ["Watercourse units", "10.00%", 0.00, 0.00, 0.00],
    ]
    
    for row_idx, row_data in enumerate(data, start=6):
        for col_idx, value in enumerate(row_data, start=1):
            ws_headline.cell(row=row_idx, column=col_idx, value=value)
    
    # Create Trading Summary Hedgerows sheet - EXACT data from GitHub issue
    ws_hedge = wb.create_sheet("Trading Summary Hedgerows")
    
    hedge_headers = ["Habitat", "Distinctiveness", "Project-wide unit change"]
    for col, header in enumerate(hedge_headers, start=1):
        ws_hedge.cell(row=1, column=col, value=header)
    
    hedge_data = [
        ["Species-rich native hedgerow with trees - associated with bank or ditch", "Very High", 0.00],
        ["Species-rich native hedgerow with trees", "High", 2.70],
        ["Species-rich native hedgerow", "Medium", -0.55],
        ["Native hedgerow with trees", "Medium", -1.04],
        ["Non-native and ornamental hedgerow", "Very Low", -0.25],
    ]
    
    for row_idx, row_data in enumerate(hedge_data, start=2):
        for col_idx, value in enumerate(row_data, start=1):
            ws_hedge.cell(row=row_idx, column=col_idx, value=value)
    
    # Remove default sheet
    if "Sheet" in wb.sheetnames:
        wb.remove(wb["Sheet"])
    
    # Save to BytesIO
    excel_buffer = io.BytesIO()
    wb.save(excel_buffer)
    excel_buffer.seek(0)
    
    # Create mock uploaded file
    mock_file = MockUploadedFile(excel_buffer)
    
    # Parse requirements
    requirements = parse_metric_requirements(mock_file)
    
    print("\n" + "="*90)
    print("FULL WORKFLOW TEST - HIGH SURPLUS OFFSETTING MEDIUM & VERY LOW DEFICITS")
    print("="*90)
    
    print("\n📥 INPUT (from Trading Summary):")
    print("-" * 90)
    print(f"{'Habitat':<60s} {'Distinctiveness':<15s} {'Units':>10s}")
    print("-" * 90)
    print(f"{'Species-rich native hedgerow with trees - bank/ditch':<60s} {'Very High':<15s} {0.00:>10.2f}")
    print(f"{'Species-rich native hedgerow with trees':<60s} {'High':<15s} {2.70:>10.2f} ✓")
    print(f"{'Species-rich native hedgerow':<60s} {'Medium':<15s} {-0.55:>10.2f} ⚠")
    print(f"{'Native hedgerow with trees':<60s} {'Medium':<15s} {-1.04:>10.2f} ⚠")
    print(f"{'Non-native and ornamental hedgerow':<60s} {'Very Low':<15s} {-0.25:>10.2f} ⚠")
    print("-" * 90)
    print(f"Total surplus:  2.70 units")
    print(f"Total deficits: 1.84 units (0.55 + 1.04 + 0.25)")
    print(f"Net gain req:   0.184 units (1.84 × 10%)")
    print(f"Total need:     2.024 units")
    
    print("\n📊 TRADING RULES:")
    print("-" * 90)
    print("✓ High (2.70) CAN offset Medium deficits (distinctiveness rank: High=3 > Medium=2)")
    print("✓ High (2.70) CAN offset Very Low deficits (distinctiveness rank: High=3 > Very Low=0)")
    print()
    print("Expected: Surplus (2.70) > Total need (2.024)")
    print("          → All requirements should be covered")
    print("          → requirements['hedgerows'] should be EMPTY")
    
    print("\n📤 OUTPUT (from parse_metric_requirements):")
    print("-" * 90)
    
    hedge_req_df = requirements["hedgerows"]
    
    if hedge_req_df.empty:
        print("✅ requirements['hedgerows'] is EMPTY")
        print()
        print("Result: ALL hedgerow requirements were correctly offset by on-site surplus!")
        print("        - High surplus (2.70) offset all Medium deficits (0.55 + 1.04 = 1.59)")
        print("        - High surplus (2.70) offset Very Low deficit (0.25)")
        print("        - Remaining surplus (2.70 - 1.84 = 0.86) covered net gain (0.184)")
        print("        - Still have 0.676 surplus remaining!")
        success = True
    else:
        print("❌ requirements['hedgerows'] is NOT EMPTY:")
        print()
        print(hedge_req_df.to_string(index=False))
        print()
        total_req = hedge_req_df["units"].sum()
        print(f"Total requirements: {total_req:.4f} units")
        print()
        print("PROBLEM: These requirements should have been offset by the High surplus (2.70 units)!")
        success = False
    
    print("\n" + "="*90)
    
    # Additional checks
    if not success:
        print("\n❌ TEST FAILED")
        print("The High surplus is NOT offsetting Medium and Very Low deficits as expected!")
        print()
        print("Possible causes:")
        print("1. Distinctiveness values not being parsed correctly")
        print("2. Trading rules not being applied correctly in apply_hedgerow_offsets()")
        print("3. Bug in can_offset_hedgerow() function")
    else:
        print("\n✅ TEST PASSED")
        print("Hedgerow surpluses ARE correctly offsetting deficits according to trading rules!")
    
    print("="*90 + "\n")
    
    return success


if __name__ == "__main__":
    success = test_full_workflow_with_high_surplus_offsetting_medium_and_very_low_deficits()
    exit(0 if success else 1)
