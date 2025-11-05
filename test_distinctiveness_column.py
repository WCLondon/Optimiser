"""
Test that the fix still works with files that have a proper "Distinctiveness" column.
This ensures we didn't break the existing functionality.
"""

import io
import pandas as pd
import openpyxl
from metric_reader import parse_metric_requirements


class MockUploadedFile:
    """Mock file object for testing"""
    def __init__(self, buffer):
        self.buffer = buffer
        self.name = "test_metric.xlsx"
    
    def read(self):
        return self.buffer.read()


def test_with_distinctiveness_column():
    """
    Test with an explicit Distinctiveness column (the newer format).
    This should still work after our fix.
    """
    wb = openpyxl.Workbook()
    
    # Create Headline Results sheet  
    ws_headline = wb.create_sheet("Headline Results")
    
    headers = ["Unit Type", "Target", "Baseline Units", "Units Required", "Unit Deficit"]
    for col, header in enumerate(headers, start=1):
        ws_headline.cell(row=5, column=col, value=header)
    
    data = [
        ["Habitat units", "10.00%", 10.00, 11.00, 0.00],
        ["Hedgerow units", "10.00%", 0.00, 0.00, 0.00],
        ["Watercourse units", "10.00%", 0.00, 0.00, 0.00],
    ]
    
    for row_idx, row_data in enumerate(data, start=6):
        for col_idx, value in enumerate(row_data, start=1):
            ws_headline.cell(row=row_idx, column=col_idx, value=value)
    
    # Create Trading Summary Hedgerows sheet with Distinctiveness column
    ws_hedge = wb.create_sheet("Trading Summary Hedgerows")
    
    # Headers
    ws_hedge.cell(row=1, column=1, value="Habitat")
    ws_hedge.cell(row=1, column=2, value="Distinctiveness")
    ws_hedge.cell(row=1, column=3, value="Project-wide unit change")
    
    # Data
    ws_hedge.cell(row=2, column=1, value="Species-rich native hedgerow")
    ws_hedge.cell(row=2, column=2, value="Medium")
    ws_hedge.cell(row=2, column=3, value=1.47)
    
    ws_hedge.cell(row=3, column=1, value="Native hedgerow")
    ws_hedge.cell(row=3, column=2, value="Low")
    ws_hedge.cell(row=3, column=3, value=-0.70)
    
    # Remove default sheet
    if "Sheet" in wb.sheetnames:
        wb.remove(wb["Sheet"])
    
    # Save to BytesIO
    excel_buffer = io.BytesIO()
    wb.save(excel_buffer)
    excel_buffer.seek(0)
    
    # Create mock uploaded file
    mock_file = MockUploadedFile(excel_buffer)
    
    # Parse
    print("\n" + "="*80)
    print("TEST: With explicit Distinctiveness column")
    print("="*80)
    print("\nThis tests the newer format where distinctiveness is in a column.")
    print("Expected: Medium surplus should offset Low deficit")
    
    requirements = parse_metric_requirements(mock_file)
    
    # Verify results
    hedge_req_df = requirements["hedgerows"]
    
    print("\n" + "-"*80)
    print("RESULT:")
    print("-"*80)
    if hedge_req_df.empty:
        print("  (empty - all requirements covered)")
    else:
        print(hedge_req_df.to_string(index=False))
    
    # Check if Native hedgerow appears
    native_rows = hedge_req_df[
        hedge_req_df["habitat"].str.contains("Native hedgerow", case=False, na=False) & 
        ~hedge_req_df["habitat"].str.contains("Species-rich", case=False, na=False)
    ]
    
    print("\n" + "="*80)
    if not native_rows.empty:
        print("❌ REGRESSION!")
        print("   Format with Distinctiveness column is broken by the fix")
        return False
    else:
        print("✅ STILL WORKING!")
        print("   Distinctiveness column format still works correctly")
        return True


if __name__ == "__main__":
    success = test_with_distinctiveness_column()
    exit(0 if success else 1)
