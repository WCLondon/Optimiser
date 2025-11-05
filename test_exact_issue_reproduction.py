"""
Test to reproduce the EXACT issue from GitHub:

Trading Summary shows:
- Medium: Species-rich native hedgerow: +1.47 units (surplus)
- Low: Native hedgerow: -0.70 units (deficit)

Expected readout: Should NOT show Native hedgerow (0.70 deficit offset by 1.47 surplus)
Actual readout (wrong): Native hedgerow 0.6996

The issue is that the Low Native hedgerow deficit (-0.70) is showing up in the
optimiser output despite there being a Medium Species-rich native hedgerow surplus (+1.47)
that should offset it according to hedgerow trading rules.
"""

import io
import pandas as pd
import openpyxl
from metric_reader import parse_metric_requirements


class MockUploadedFile:
    """Mock file object for testing"""
    def __init__(self, buffer):
        self.buffer = buffer
        self.name = "test_metric.xlsx"
    
    def read(self):
        return self.buffer.read()


def create_issue_metric():
    """
    Create a metric file matching the EXACT data from the issue:
    
    From Trading Summary Hedgerows:
    - Very High Distinctiveness: Species-rich native hedgerow with trees - associated with bank or ditch: 0.00
    - High Distinctiveness: 
        - Species-rich native hedgerow with trees: 0.00
        - Species-rich native hedgerow - associated with bank or ditch: 0.00
        - Native hedgerow with trees - associated with bank or ditch: 0.00
    - Medium Distinctiveness:
        - Species-rich native hedgerow: 1.47 ✓
        - Native hedgerow - associated with bank or ditch: 0.00
        - Native hedgerow with trees: 0.00
        - (others 0.00)
    - Low Distinctiveness:
        - Native hedgerow: -0.70 ⚠
        - Line of trees: 0.00
        - Line of trees - associated with bank or ditch: 0.00
    - Very Low Distinctiveness:
        - Non-native and ornamental hedgerow: 0.00
    """
    
    wb = openpyxl.Workbook()
    
    # Create Headline Results sheet
    ws_headline = wb.create_sheet("Headline Results")
    
    headers = ["Unit Type", "Target", "Baseline Units", "Units Required", "Unit Deficit"]
    for col, header in enumerate(headers, start=1):
        ws_headline.cell(row=5, column=col, value=header)
    
    # Baseline from issue: enough to calculate 0.77 cumulative availability
    # 0.77 = 1.47 (Medium surplus) - 0.70 (Low deficit)
    # Net gain requirement should be small
    data = [
        ["Habitat units", "10.00%", 10.00, 11.00, 0.00],
        ["Hedgerow units", "10.00%", 0.00, 0.00, 0.00],  # No net gain for simplicity
        ["Watercourse units", "10.00%", 0.00, 0.00, 0.00],
    ]
    
    for row_idx, row_data in enumerate(data, start=6):
        for col_idx, value in enumerate(row_data, start=1):
            ws_headline.cell(row=row_idx, column=col_idx, value=value)
    
    # Create Trading Summary Hedgerows sheet with SECTION HEADERS format
    # This is the format the issue likely used
    ws_hedge = wb.create_sheet("Trading Summary Hedgerows")
    
    row = 1
    
    # Very High Distinctiveness section
    ws_hedge.cell(row=row, column=1, value="Very High Distinctiveness")
    row += 1
    ws_hedge.cell(row=row, column=1, value="Habitat group")
    ws_hedge.cell(row=row, column=2, value="On-site unit change")
    ws_hedge.cell(row=row, column=3, value="Off-site unit change")
    ws_hedge.cell(row=row, column=4, value="Project-wide unit change")
    row += 1
    ws_hedge.cell(row=row, column=1, value="Species-rich native hedgerow with trees - associated with bank or ditch")
    ws_hedge.cell(row=row, column=2, value=0.00)
    ws_hedge.cell(row=row, column=3, value=0.00)
    ws_hedge.cell(row=row, column=4, value=0.00)
    row += 2
    
    # High Distinctiveness section
    ws_hedge.cell(row=row, column=1, value="High Distinctiveness")
    row += 1
    ws_hedge.cell(row=row, column=1, value="Habitat group")
    ws_hedge.cell(row=row, column=2, value="On-site unit change")
    ws_hedge.cell(row=row, column=3, value="Off-site unit change")
    ws_hedge.cell(row=row, column=4, value="Project-wide unit change")
    row += 1
    ws_hedge.cell(row=row, column=1, value="Species-rich native hedgerow with trees")
    ws_hedge.cell(row=row, column=2, value=0.00)
    ws_hedge.cell(row=row, column=3, value=0.00)
    ws_hedge.cell(row=row, column=4, value=0.00)
    row += 1
    ws_hedge.cell(row=row, column=1, value="Species-rich native hedgerow - associated with bank or ditch")
    ws_hedge.cell(row=row, column=2, value=0.00)
    ws_hedge.cell(row=row, column=3, value=0.00)
    ws_hedge.cell(row=row, column=4, value=0.00)
    row += 1
    ws_hedge.cell(row=row, column=1, value="Native hedgerow with trees - associated with bank or ditch")
    ws_hedge.cell(row=row, column=2, value=0.00)
    ws_hedge.cell(row=row, column=3, value=0.00)
    ws_hedge.cell(row=row, column=4, value=0.00)
    row += 2
    
    # Medium Distinctiveness section
    ws_hedge.cell(row=row, column=1, value="Medium Distinctiveness")
    row += 1
    ws_hedge.cell(row=row, column=1, value="Habitat group")
    ws_hedge.cell(row=row, column=2, value="On-site unit change")
    ws_hedge.cell(row=row, column=3, value="Off-site unit change")
    ws_hedge.cell(row=row, column=4, value="Project-wide unit change")
    row += 1
    ws_hedge.cell(row=row, column=1, value="Species-rich native hedgerow")
    ws_hedge.cell(row=row, column=2, value=1.47)
    ws_hedge.cell(row=row, column=3, value=0.00)
    ws_hedge.cell(row=row, column=4, value=1.47)
    row += 1
    ws_hedge.cell(row=row, column=1, value="Native hedgerow - associated with bank or ditch")
    ws_hedge.cell(row=row, column=2, value=0.00)
    ws_hedge.cell(row=row, column=3, value=0.00)
    ws_hedge.cell(row=row, column=4, value=0.00)
    row += 1
    ws_hedge.cell(row=row, column=1, value="Native hedgerow with trees")
    ws_hedge.cell(row=row, column=2, value=0.00)
    ws_hedge.cell(row=row, column=3, value=0.00)
    ws_hedge.cell(row=row, column=4, value=0.00)
    row += 2
    
    # Low Distinctiveness section
    ws_hedge.cell(row=row, column=1, value="Low Distinctiveness")
    row += 1
    ws_hedge.cell(row=row, column=1, value="Habitat group")
    ws_hedge.cell(row=row, column=2, value="On-site unit change")
    ws_hedge.cell(row=row, column=3, value="Off-site unit change")
    ws_hedge.cell(row=row, column=4, value="Project-wide unit change")
    row += 1
    ws_hedge.cell(row=row, column=1, value="Native hedgerow")
    ws_hedge.cell(row=row, column=2, value=-0.70)
    ws_hedge.cell(row=row, column=3, value=0.00)
    ws_hedge.cell(row=row, column=4, value=-0.70)
    row += 1
    ws_hedge.cell(row=row, column=1, value="Line of trees")
    ws_hedge.cell(row=row, column=2, value=0.00)
    ws_hedge.cell(row=row, column=3, value=0.00)
    ws_hedge.cell(row=row, column=4, value=0.00)
    row += 1
    ws_hedge.cell(row=row, column=1, value="Line of trees - associated with bank or ditch")
    ws_hedge.cell(row=row, column=2, value=0.00)
    ws_hedge.cell(row=row, column=3, value=0.00)
    ws_hedge.cell(row=row, column=4, value=0.00)
    row += 2
    
    # Very Low Distinctiveness section
    ws_hedge.cell(row=row, column=1, value="Very Low Distinctiveness")
    row += 1
    ws_hedge.cell(row=row, column=1, value="Habitat group")
    ws_hedge.cell(row=row, column=2, value="On-site unit change")
    ws_hedge.cell(row=row, column=3, value="Off-site unit change")
    ws_hedge.cell(row=row, column=4, value="Project-wide unit change")
    row += 1
    ws_hedge.cell(row=row, column=1, value="Non-native and ornamental hedgerow")
    ws_hedge.cell(row=row, column=2, value=0.00)
    ws_hedge.cell(row=row, column=3, value=0.00)
    ws_hedge.cell(row=row, column=4, value=0.00)
    
    # Remove default sheet
    if "Sheet" in wb.sheetnames:
        wb.remove(wb["Sheet"])
    
    return wb


def test_exact_issue():
    """
    Test with the exact data from the GitHub issue
    """
    wb = create_issue_metric()
    
    # Save to BytesIO
    excel_buffer = io.BytesIO()
    wb.save(excel_buffer)
    excel_buffer.seek(0)
    
    # Create mock uploaded file
    mock_file = MockUploadedFile(excel_buffer)
    
    # Parse
    requirements = parse_metric_requirements(mock_file)
    
    # Verify results
    hedge_req_df = requirements["hedgerows"]
    
    print("\n" + "="*80)
    print("EXACT REPRODUCTION OF GITHUB ISSUE")
    print("="*80)
    print("\nFrom Trading Summary Hedgerows:")
    print("  Medium Distinctiveness:")
    print("    - Species-rich native hedgerow: +1.47 units ✓")
    print("  Low Distinctiveness:")
    print("    - Native hedgerow: -0.70 units ⚠")
    print("\nHedgerow Trading Rules:")
    print("  - Medium, Low, Very Low: Same distinctiveness or better can offset")
    print("  - So Medium surplus (1.47) SHOULD offset Low deficit (0.70)")
    print("\nExpected behavior:")
    print("  ✓ Low Native hedgerow deficit offset by Medium surplus")
    print("  ✓ Remaining surplus: 1.47 - 0.70 = 0.77 units")
    print("  ✓ No 'Native hedgerow' in requirements")
    
    print("\n" + "-"*80)
    print("ACTUAL REQUIREMENTS FROM METRIC READER:")
    print("-"*80)
    if hedge_req_df.empty:
        print("  (empty - all requirements covered by on-site surplus)")
    else:
        print(hedge_req_df.to_string(index=False))
        print(f"\n  Total units: {hedge_req_df['units'].sum():.6f}")
    
    # Check if Native hedgerow appears in requirements
    native_rows = hedge_req_df[
        hedge_req_df["habitat"].str.contains("Native hedgerow", case=False, na=False) & 
        ~hedge_req_df["habitat"].str.contains("Species-rich", case=False, na=False) &
        ~hedge_req_df["habitat"].str.contains("with trees", case=False, na=False) &
        ~hedge_req_df["habitat"].str.contains("associated", case=False, na=False)
    ]
    
    print("\n" + "="*80)
    if not native_rows.empty:
        print("❌ ISSUE REPRODUCED!")
        print(f"   'Native hedgerow' appears with {native_rows.iloc[0]['units']:.6f} units")
        print("   Expected: 0 (should be fully offset by Medium Species-rich surplus)")
        return False
    else:
        print("✅ ISSUE RESOLVED!")
        print("   'Native hedgerow' correctly offset by surplus")
        print("   No Native hedgerow in requirements")
        return True


if __name__ == "__main__":
    success = test_exact_issue()
    exit(0 if success else 1)
