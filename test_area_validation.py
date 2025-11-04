"""
Test that area habitat validation warning works when distinctiveness extraction fails
"""

import io
import pandas as pd
import openpyxl
import warnings
from metric_reader import parse_metric_requirements


class MockUploadedFile:
    """Mock file object for testing"""
    def __init__(self, buffer):
        self.buffer = buffer
        self.name = "test_metric.xlsx"
    
    def read(self):
        return self.buffer.read()


def test_area_failed_distinctiveness():
    """Test that warning is issued when distinctiveness extraction fails for area habitats"""
    
    wb = openpyxl.Workbook()
    
    # Create Headline Results
    ws_headline = wb.create_sheet("Headline Results")
    headers = ["Unit Type", "Target", "Baseline Units", "Units Required", "Unit Deficit"]
    for col, header in enumerate(headers, start=1):
        ws_headline.cell(row=5, column=col, value=header)
    
    data = [
        ["Habitat units", "10.00%", 5.00, 5.50, 0.00],
        ["Hedgerow units", "10.00%", 0.00, 0.00, 0.00],
        ["Watercourse units", "10.00%", 0.00, 0.00, 0.00],
    ]
    
    for row_idx, row_data in enumerate(data, start=6):
        for col_idx, value in enumerate(row_data, start=1):
            ws_headline.cell(row=row_idx, column=col_idx, value=value)
    
    # Create Trading Summary WITHOUT proper distinctiveness headers
    ws_area = wb.create_sheet("Trading Summary Area Habitats")
    
    area_headers = ["Habitat group", "Group", "Project-wide unit change"]
    for col, header in enumerate(area_headers, start=1):
        ws_area.cell(row=1, column=col, value=header)
    
    area_data = [
        ["Modified grassland", "Grassland", -2.00],
        ["Mixed scrub", "Heathland", 1.50],
    ]
    
    for row_idx, row_data in enumerate(area_data, start=2):
        for col_idx, value in enumerate(row_data, start=1):
            ws_area.cell(row=row_idx, column=col_idx, value=value)
    
    if "Sheet" in wb.sheetnames:
        wb.remove(wb["Sheet"])
    
    excel_buffer = io.BytesIO()
    wb.save(excel_buffer)
    excel_buffer.seek(0)
    
    mock_file = MockUploadedFile(excel_buffer)
    
    # Parse with warnings captured
    with warnings.catch_warnings(record=True) as w:
        warnings.simplefilter("always")
        requirements = parse_metric_requirements(mock_file)
        
        # Check if warning was issued
        area_warnings = [warning for warning in w 
                        if "Area Habitats" in str(warning.message) 
                        and "distinctiveness" in str(warning.message).lower()]
        
        print("\n" + "="*80)
        print("Test: Area habitat validation warning")
        print("="*80)
        
        if area_warnings:
            print("\n✅ WARNING ISSUED AS EXPECTED:")
            for warning in area_warnings:
                print(f"  {warning.message}")
            success = True
        else:
            print("\n❌ NO WARNING ISSUED (unexpected)")
            success = False
        
        area_req_df = requirements["area"]
        if not area_req_df.empty:
            print("\n✅ Deficits appear in requirements (as expected when distinctiveness unknown)")
            print(area_req_df.to_string(index=False))
        
        print("\n" + "="*80)
        
        return success


if __name__ == "__main__":
    success = test_area_failed_distinctiveness()
    exit(0 if success else 1)
