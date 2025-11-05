"""
Test parsing the exact format provided by the user.

Trading Summary structure:
- Section headers: "Medium Distinctiveness", "Low Distinctiveness"
- No explicit Distinctiveness column
- Has summary columns like "Medium Distinctiveness Units available to offset Lower Distinctiveness Deficit"

Expected behavior:
- Parser should extract distinctiveness from section headers
- Medium surplus (0.06) and Low surplus (0.01) should offset Net Gain (0.018)
"""

import io
import pandas as pd
import openpyxl
from metric_reader import parse_metric_requirements


class MockUploadedFile:
    """Mock file object for testing"""
    def __init__(self, buffer):
        self.buffer = buffer
        self.name = "test_metric.xlsx"
    
    def read(self):
        return self.buffer.read()


def test_user_watercourse_format():
    """
    Test the exact format from the user's metric file.
    
    Headline Results:
    - Watercourse units: 10% target, 0.18 baseline
    - Net Gain needed: 0.018 units
    
    Trading Summary:
    - Medium Distinctiveness: Ditches +0.06
    - Low Distinctiveness: Culvert +0.01
    - Total: 0.07 surplus (should cover 0.018 Net Gain)
    """
    
    wb = openpyxl.Workbook()
    
    # Create Headline Results sheet
    ws_headline = wb.create_sheet("Headline Results")
    
    headers = ["Unit Type", "Target", "Baseline Units", "Units Required", "Unit Deficit"]
    for col, header in enumerate(headers, start=1):
        ws_headline.cell(row=5, column=col, value=header)
    
    data = [
        ["Habitat units", "10.00%", 7.71, 8.48, 3.38],
        ["Hedgerow units", "10.00%", 3.54, 3.89, 0.00],
        ["Watercourse units", "10.00%", 0.18, 0.19, 0.00],
    ]
    
    for row_idx, row_data in enumerate(data, start=6):
        for col_idx, value in enumerate(row_data, start=1):
            ws_headline.cell(row=row_idx, column=col_idx, value=value)
    
    # Create Trading Summary WaterC's sheet matching user's format
    ws_water = wb.create_sheet("Trading Summary WaterCs")
    
    # Headers with summary columns
    headers_row = [
        "Habitat group",
        "On-site unit change",
        "Off-site unit change",
        "Project-wide unit change",
        "Medium Distinctiveness Units available to offset Lower Distinctiveness Deficit"
    ]
    
    for col, header in enumerate(headers_row, start=1):
        ws_water.cell(row=1, column=col, value=header)
    
    current_row = 2
    
    # Medium Distinctiveness section
    ws_water.cell(row=current_row, column=1, value="Medium Distinctiveness")
    current_row += 1
    
    ws_water.cell(row=current_row, column=1, value="Habitat group")
    ws_water.cell(row=current_row, column=4, value="Project wide unit change")
    ws_water.cell(row=current_row, column=5, value="Medium Distinctiveness Units available to offset Lower Distinctiveness Deficit")
    current_row += 1
    
    # Ditches - Medium surplus
    ws_water.cell(row=current_row, column=1, value="Ditches")
    ws_water.cell(row=current_row, column=2, value=0.06)
    ws_water.cell(row=current_row, column=3, value=0.00)
    ws_water.cell(row=current_row, column=4, value=0.06)
    ws_water.cell(row=current_row, column=5, value=0.06)
    current_row += 1
    
    # Remaining losses row
    ws_water.cell(row=current_row, column=1, value="Remaining losses; Like for like not satisfied")
    ws_water.cell(row=current_row, column=4, value=0.00)
    current_row += 1
    
    # Canals - no change
    ws_water.cell(row=current_row, column=1, value="Canals")
    ws_water.cell(row=current_row, column=2, value=0.00)
    ws_water.cell(row=current_row, column=3, value=0.00)
    ws_water.cell(row=current_row, column=4, value=0.00)
    current_row += 2  # Skip a row
    
    # Low Distinctiveness section
    ws_water.cell(row=current_row, column=1, value="Low Distinctiveness")
    current_row += 1
    
    ws_water.cell(row=current_row, column=1, value="Habitat group")
    ws_water.cell(row=current_row, column=4, value="Project wide unit change")
    ws_water.cell(row=current_row, column=5, value="Low Distinctiveness net change in units")
    current_row += 1
    
    # Culvert - Low surplus
    ws_water.cell(row=current_row, column=1, value="Culvert")
    ws_water.cell(row=current_row, column=2, value=0.01)
    ws_water.cell(row=current_row, column=3, value=0.00)
    ws_water.cell(row=current_row, column=4, value=0.01)
    ws_water.cell(row=current_row, column=5, value=0.01)
    current_row += 1
    
    # Remove default sheet
    if "Sheet" in wb.sheetnames:
        wb.remove(wb["Sheet"])
    
    # Save to BytesIO
    excel_buffer = io.BytesIO()
    wb.save(excel_buffer)
    excel_buffer.seek(0)
    
    # Create mock uploaded file
    mock_file = MockUploadedFile(excel_buffer)
    
    # Parse
    print("=" * 90)
    print("TEST: USER'S WATERCOURSE FORMAT")
    print("=" * 90)
    print()
    print("📊 METRIC DATA (from user's comment):")
    print("-" * 90)
    print("Watercourse Trading Summary:")
    print("  - Medium Distinctiveness: Ditches +0.06 units")
    print("  - Low Distinctiveness: Culvert +0.01 units")
    print("  - Total surplus: 0.07 units")
    print()
    print("Headline Results:")
    print("  - Watercourse baseline: 0.18 units")
    print("  - Target: 10%")
    print("  - Net Gain requirement: 0.18 × 10% = 0.018 units")
    print()
    print("📋 EXPECTED BEHAVIOR:")
    print("-" * 90)
    print("  - Surplus (0.07) should offset Net Gain (0.018)")
    print("  - Result: No requirements (all covered)")
    print()
    
    try:
        requirements = parse_metric_requirements(mock_file)
        
        water_req_df = requirements["watercourses"]
        
        print("📤 ACTUAL RESULT:")
        print("-" * 90)
        print(f"Watercourse requirements count: {len(water_req_df)}")
        
        if not water_req_df.empty:
            print("\nRequirements:")
            total_req = 0.0
            for _, row in water_req_df.iterrows():
                print(f"  - {row['habitat']}: {row['units']:.4f} units")
                total_req += row['units']
            print(f"\nTotal requirements: {total_req:.4f} units")
            
            # Check if Net Gain is in requirements
            has_net_gain = any("Net Gain" in str(h) for h in water_req_df["habitat"])
            if has_net_gain:
                net_gain_row = water_req_df[water_req_df["habitat"].str.contains("Net Gain", na=False)]
                net_gain_units = net_gain_row["units"].iloc[0] if not net_gain_row.empty else 0.0
                print()
                print(f"❌ ISSUE REPRODUCED: Net Gain requirement remains: {net_gain_units:.4f} units")
                print(f"   Expected: 0.00 units (should be offset by Medium/Low surplus)")
                print()
                print("   This confirms the user's issue!")
                return False
        else:
            print("  (empty - all covered)")
            print()
            print("✅ SUCCESS: Net Gain fully offset by Medium/Low surplus!")
            return True
    
    except Exception as e:
        print(f"❌ ERROR during parsing: {e}")
        import traceback
        traceback.print_exc()
        return False
    
    print()
    print("=" * 90)


if __name__ == "__main__":
    print()
    success = test_user_watercourse_format()
    print()
    
    if success:
        print("✅ TEST PASSED - User's format works correctly")
    else:
        print("❌ TEST FAILED - Reproduced user's issue")
        exit(1)
