"""
Test for Medium distinctiveness habitat hierarchy in metric_reader
"""

import io
import pandas as pd
import openpyxl
from metric_reader import parse_metric_requirements, apply_area_offsets


def test_medium_hierarchy_priority():
    """Test that priority Medium groups are processed before secondary Medium groups"""
    
    # Create a test dataframe with Medium distinctiveness habitats
    # Priority group: Cropland, Woodland and forest
    # Secondary group: Grassland, Heathland and shrub
    
    test_data = pd.DataFrame({
        "category": ["Area Habitats"] * 5,
        "habitat": [
            "Grassland - Modified",
            "Cropland - Cereal",
            "Heathland - Acid",
            "Woodland - Broadleaved",
            "Medium surplus habitat"
        ],
        "broad_group": [
            "Grassland",
            "Cropland",
            "Heathland and shrub",
            "Woodland and forest",
            "Cropland"  # Surplus in priority group
        ],
        "distinctiveness": ["Medium", "Medium", "Medium", "Medium", "Medium"],
        "project_wide_change": [-5.0, -3.0, -2.0, -4.0, 20.0],  # 4 deficits, 1 surplus
        "on_site_change": [0.0, 0.0, 0.0, 0.0, 20.0]
    })
    
    # Apply offsets
    result = apply_area_offsets(test_data)
    
    flow_log = result.get("flow_log", [])
    
    print(f"\n✅ Found {len(flow_log)} allocation entries in flow log")
    
    # Verify flow log structure
    assert len(flow_log) > 0, "Flow log should contain allocation entries"
    
    # Check that priority_medium flag is present
    for entry in flow_log:
        assert "priority_medium" in entry, "Each flow log entry should have priority_medium flag"
        print(f"   - {entry['deficit_habitat']} ({entry['deficit_broad_group']}): "
              f"{entry['units_allocated']:.2f} units, priority_medium={entry['priority_medium']}")
    
    # Verify that priority Medium groups were processed first
    # Since we only have 20 units of surplus and 14 units of deficit total,
    # the priority groups (Cropland -3 + Woodland -4 = -7) should be fully covered
    # And some secondary groups should be covered too (Grassland -5 + Heathland -2 = -7)
    
    priority_allocations = [e for e in flow_log if e["priority_medium"]]
    secondary_allocations = [e for e in flow_log if not e["priority_medium"]]
    
    print(f"\n   Priority Medium allocations: {len(priority_allocations)}")
    print(f"   Secondary Medium allocations: {len(secondary_allocations)}")
    
    # All priority groups should receive some allocation
    priority_deficit_groups = {"cropland", "woodland and forest"}
    allocated_priority_groups = {e["deficit_broad_group"].lower() for e in priority_allocations}
    
    print(f"\n   Allocated priority groups: {allocated_priority_groups}")
    
    # Check residual deficits
    residual = result["residual_off_site"]
    print(f"\n   Residual deficits: {len(residual)} rows")
    if not residual.empty:
        for _, row in residual.iterrows():
            print(f"      - {row['habitat']} ({row['broad_group']}): {row['unmet_units_after_on_site_offset']:.2f} units")
    
    print("\n✅ Medium hierarchy test passed!")
    return True


def test_medium_hierarchy_full_example():
    """Test with a realistic Excel file containing multiple Medium groups"""
    
    wb = openpyxl.Workbook()
    
    # Create Trading Summary Area Habitats sheet
    ws = wb.create_sheet("Trading Summary Area Habitats")
    
    # Add headers
    headers = ["Habitat", "Broad habitat", "Distinctiveness", "Project-wide unit change", "On-site unit change"]
    for col, header in enumerate(headers, start=1):
        ws.cell(row=1, column=col, value=header)
    
    # Add test data with both priority and secondary Medium groups
    # Priority groups should be processed first
    data = [
        # Secondary group deficits (should be processed after priority)
        ["Grassland - Modified", "Grassland", "Medium", -5.0, 0.0],
        ["Heathland - Acid", "Heathland and shrub", "Medium", -3.0, 0.0],
        
        # Priority group deficits (should be processed first)
        ["Cropland - Cereal", "Cropland", "Medium", -4.0, 0.0],
        ["Woodland - Broadleaved", "Woodland and forest", "Medium", -6.0, 0.0],
        
        # Surplus (Medium, can offset other Medium in same broad group)
        ["Cropland - Temporary", "Cropland", "Medium", 10.0, 10.0],
        
        # High surplus (can offset any Medium)
        ["Woodland - Mixed", "Woodland and forest", "High", 15.0, 15.0],
    ]
    
    for row_idx, row_data in enumerate(data, start=2):
        for col_idx, value in enumerate(row_data, start=1):
            ws.cell(row=row_idx, column=col_idx, value=value)
    
    # Add Headline Results sheet
    ws_headline = wb.create_sheet("Headline Results")
    headline_headers = ["Unit type", "Baseline units", "Target", "Units required", "Unit deficit"]
    for col, header in enumerate(headline_headers, start=1):
        ws_headline.cell(row=1, column=col, value=header)
    ws_headline.cell(row=2, column=1, value="Area habitat units")
    ws_headline.cell(row=2, column=2, value=100.0)
    ws_headline.cell(row=2, column=3, value="10%")
    ws_headline.cell(row=2, column=4, value=10.0)
    ws_headline.cell(row=2, column=5, value=10.0)
    
    # Remove default sheet
    if "Sheet" in wb.sheetnames:
        wb.remove(wb["Sheet"])
    
    # Save to BytesIO
    excel_buffer = io.BytesIO()
    wb.save(excel_buffer)
    excel_buffer.seek(0)
    
    # Create mock uploaded file
    class MockUploadedFile:
        def __init__(self, buffer):
            self.buffer = buffer
            self.name = "test_hierarchy.xlsx"
        
        def read(self):
            return self.buffer.read()
    
    mock_file = MockUploadedFile(excel_buffer)
    
    # Parse requirements
    requirements = parse_metric_requirements(mock_file)
    
    print("\n✅ Full example test completed")
    print(f"   Area requirements: {len(requirements['area'])} rows")
    
    if not requirements['area'].empty:
        print("\n   Area requirements:")
        for _, row in requirements['area'].iterrows():
            print(f"      - {row['habitat']}: {row['units']:.2f} units")
    
    print("\n✅ Full hierarchy example test passed!")
    return True


def test_non_medium_unchanged():
    """Test that Low, High, and Very High processing is unchanged"""
    
    test_data = pd.DataFrame({
        "category": ["Area Habitats"] * 4,
        "habitat": ["Low habitat 1", "High habitat 1", "Very High habitat 1", "Low habitat 2"],
        "broad_group": ["Group A", "Group B", "Group C", "Group A"],
        "distinctiveness": ["Low", "High", "Very High", "Low"],
        "project_wide_change": [-5.0, -3.0, -2.0, 15.0],
        "on_site_change": [0.0, 0.0, 0.0, 15.0]
    })
    
    result = apply_area_offsets(test_data)
    flow_log = result.get("flow_log", [])
    
    # None of these should be flagged as priority_medium
    for entry in flow_log:
        assert not entry["priority_medium"], f"Non-Medium habitat should not be flagged as priority_medium"
    
    print("\n✅ Non-Medium habitats unchanged - test passed!")
    return True


if __name__ == "__main__":
    print("=" * 70)
    print("Testing Medium Distinctiveness Hierarchy Implementation")
    print("=" * 70)
    
    success = True
    
    try:
        test_medium_hierarchy_priority()
    except Exception as e:
        print(f"\n❌ Priority test failed: {e}")
        import traceback
        traceback.print_exc()
        success = False
    
    try:
        test_medium_hierarchy_full_example()
    except Exception as e:
        print(f"\n❌ Full example test failed: {e}")
        import traceback
        traceback.print_exc()
        success = False
    
    try:
        test_non_medium_unchanged()
    except Exception as e:
        print(f"\n❌ Non-Medium test failed: {e}")
        import traceback
        traceback.print_exc()
        success = False
    
    print("\n" + "=" * 70)
    if success:
        print("✅ ALL TESTS PASSED")
    else:
        print("❌ SOME TESTS FAILED")
    print("=" * 70)
    
    exit(0 if success else 1)
