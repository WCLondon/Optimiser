"""
Test for watercourse trading rules implementation

Tests the following watercourse trading rules:
- Very High: Not eligible for normal trading (bespoke compensation required)
- High: Same habitat required (like-for-like)
- Medium: Same habitat required (like-for-like)
- Low: Must trade to better distinctiveness habitat
"""

import io
import pandas as pd
import openpyxl
from metric_reader import parse_metric_requirements, can_offset_watercourse


class MockUploadedFile:
    """Mock file object for testing"""
    def __init__(self, buffer):
        self.buffer = buffer
        self.name = "test_metric.xlsx"
    
    def read(self):
        return self.buffer.read()


def test_can_offset_watercourse_very_high():
    """Test that Very High distinctiveness watercourses cannot offset anything (bespoke compensation required)"""
    
    # Very High deficits should not be offsettable by any surplus
    assert not can_offset_watercourse("Very High", "Other rivers and streams", "Very High", "Other rivers and streams"), \
        "Very High should not be offsettable (bespoke compensation required)"
    assert not can_offset_watercourse("Very High", "Priority habitat", "High", "Priority habitat"), \
        "Very High should not be offsettable even by higher distinctiveness"
    
    print("✅ test_can_offset_watercourse_very_high passed")


def test_can_offset_watercourse_high():
    """Test that High distinctiveness requires like-for-like (same habitat)"""
    
    # High requires same habitat
    assert can_offset_watercourse("High", "Other rivers and streams", "High", "Other rivers and streams"), \
        "High should be offsettable by High of same habitat"
    assert can_offset_watercourse("High", "Other rivers and streams", "High", "rivers and streams"), \
        "High should be offsettable by High of same habitat (normalized)"
    
    # High cannot be offset by different habitat, even if higher distinctiveness
    assert not can_offset_watercourse("High", "Other rivers and streams", "High", "Canals"), \
        "High cannot be offset by different habitat"
    assert not can_offset_watercourse("High", "Ditches", "Very High", "Canals"), \
        "High cannot be offset by different habitat even with higher distinctiveness"
    
    # High cannot be offset by lower distinctiveness
    assert not can_offset_watercourse("High", "Other rivers and streams", "Medium", "Other rivers and streams"), \
        "High cannot be offset by lower distinctiveness"
    
    print("✅ test_can_offset_watercourse_high passed")


def test_can_offset_watercourse_medium():
    """Test that Medium distinctiveness requires like-for-like (same habitat)"""
    
    # Medium requires same habitat
    assert can_offset_watercourse("Medium", "Ditches", "Medium", "Ditches"), \
        "Medium should be offsettable by Medium of same habitat"
    assert can_offset_watercourse("Medium", "Ditches", "High", "Ditches"), \
        "Medium should be offsettable by High of same habitat"
    
    # Medium cannot be offset by different habitat
    assert not can_offset_watercourse("Medium", "Ditches", "Medium", "Canals"), \
        "Medium cannot be offset by different habitat"
    assert not can_offset_watercourse("Medium", "Ditches", "High", "Canals"), \
        "Medium cannot be offset by different habitat even with higher distinctiveness"
    
    # Medium cannot be offset by lower distinctiveness
    assert not can_offset_watercourse("Medium", "Ditches", "Low", "Ditches"), \
        "Medium cannot be offset by lower distinctiveness"
    
    print("✅ test_can_offset_watercourse_medium passed")


def test_can_offset_watercourse_low():
    """Test that Low distinctiveness must trade to better distinctiveness"""
    
    # Low cannot be offset by Low (must trade up)
    assert not can_offset_watercourse("Low", "Culvert", "Low", "Culvert"), \
        "Low cannot be offset by Low (must trade to better distinctiveness)"
    
    # Low can be offset by higher distinctiveness of same habitat
    assert can_offset_watercourse("Low", "Culvert", "Medium", "Culvert"), \
        "Low should be offsettable by Medium of same habitat"
    assert can_offset_watercourse("Low", "Culvert", "High", "Culvert"), \
        "Low should be offsettable by High of same habitat"
    
    # Low cannot be offset by higher distinctiveness of different habitat
    assert not can_offset_watercourse("Low", "Culvert", "Medium", "Ditches"), \
        "Low cannot be offset by different habitat even with higher distinctiveness"
    
    print("✅ test_can_offset_watercourse_low passed")


def test_watercourse_habitat_normalization():
    """Test that habitat names are properly normalized for matching"""
    
    # Rivers and streams variations
    assert can_offset_watercourse("High", "Other rivers and streams", "High", "rivers and streams"), \
        "Rivers/streams variations should match"
    assert can_offset_watercourse("High", "Rivers and streams", "High", "Other rivers and streams"), \
        "Rivers/streams variations should match (reverse)"
    
    # Canals
    assert can_offset_watercourse("Medium", "Canals", "Medium", "canals"), \
        "Canals case variations should match"
    
    # Ditches
    assert can_offset_watercourse("Medium", "Ditches", "Medium", "ditches"), \
        "Ditches case variations should match"
    
    # Different habitats should not match
    assert not can_offset_watercourse("High", "Canals", "High", "Ditches"), \
        "Different habitats should not match"
    assert not can_offset_watercourse("High", "Rivers and streams", "High", "Ditches"), \
        "Different habitats should not match"
    
    print("✅ test_watercourse_habitat_normalization passed")


def test_watercourse_trading_with_surpluses():
    """Test that watercourse deficits are properly offset by eligible surpluses"""
    
    wb = openpyxl.Workbook()
    
    # Create Headline Results sheet
    ws_headline = wb.create_sheet("Headline Results")
    
    headers = ["Unit Type", "Target", "Baseline Units", "Units Required", "Unit Deficit"]
    for col, header in enumerate(headers, start=1):
        ws_headline.cell(row=5, column=col, value=header)
    
    data = [
        ["Habitat units", "10.00%", 0.00, 0.00, 0.00],
        ["Hedgerow units", "10.00%", 0.00, 0.00, 0.00],
        ["Watercourse units", "10.00%", 5.00, 5.50, 0.00],
    ]
    
    for row_idx, row_data in enumerate(data, start=6):
        for col_idx, value in enumerate(row_data, start=1):
            ws_headline.cell(row=row_idx, column=col_idx, value=value)
    
    # Create Trading Summary WaterCs sheet
    ws_water = wb.create_sheet("Trading Summary WaterCs")
    
    # Add headers
    water_headers = ["Habitat", "Distinctiveness", "Project-wide unit change"]
    for col, header in enumerate(water_headers, start=1):
        ws_water.cell(row=1, column=col, value=header)
    
    # Test data: deficit in Medium Ditches, surplus in Medium Ditches (should offset)
    # and deficit in High rivers, surplus in Medium rivers (should NOT offset)
    # Make sure surpluses don't fully cover net gain requirement
    water_data = [
        ["Ditches", "Medium", -2.0],
        ["Ditches", "Medium", 2.0],  # Exactly offsets the above deficit (no surplus left)
        ["Other rivers and streams", "High", -1.5],
        ["Other rivers and streams", "Medium", 0.3],  # Should NOT offset (lower distinctiveness), but < net gain
    ]
    
    for row_idx, row_data in enumerate(water_data, start=2):
        for col_idx, value in enumerate(row_data, start=1):
            ws_water.cell(row=row_idx, column=col_idx, value=value)
    
    # Remove default sheet
    if "Sheet" in wb.sheetnames:
        wb.remove(wb["Sheet"])
    
    # Save to BytesIO
    excel_buffer = io.BytesIO()
    wb.save(excel_buffer)
    excel_buffer.seek(0)
    
    # Create mock uploaded file
    mock_file = MockUploadedFile(excel_buffer)
    
    # Parse
    requirements = parse_metric_requirements(mock_file)
    
    # Check watercourse requirements
    water_req_df = requirements["watercourses"]
    assert not water_req_df.empty, "Watercourse requirements should not be empty"
    
    # The Medium Ditches deficit (-2.0) should be fully offset by Medium Ditches surplus (+2.5)
    # So it should not appear in requirements
    ditches_rows = water_req_df[water_req_df["habitat"].str.contains("Ditches", case=False, na=False)]
    ditches_rows = ditches_rows[~ditches_rows["habitat"].str.contains("Net Gain", case=False, na=False)]
    assert ditches_rows.empty, "Medium Ditches deficit should be fully offset by Medium Ditches surplus"
    
    # The High rivers deficit (-1.5) should NOT be offset by Medium rivers surplus
    # So it should appear in requirements
    rivers_rows = water_req_df[water_req_df["habitat"].str.contains("rivers", case=False, na=False)]
    rivers_rows = rivers_rows[~rivers_rows["habitat"].str.contains("Net Gain", case=False, na=False)]
    assert not rivers_rows.empty, "High rivers deficit should remain (cannot be offset by Medium)"
    
    rivers_units = rivers_rows.iloc[0]["units"]
    assert abs(rivers_units - 1.5) < 0.0001, \
        f"Expected 1.5 units for High rivers deficit, got {rivers_units}"
    
    # Net gain should also be present
    net_gain_rows = water_req_df[water_req_df["habitat"].str.contains("Net Gain", case=False, na=False)]
    assert not net_gain_rows.empty, "Net Gain (Watercourses) should be in requirements"
    
    print("✅ test_watercourse_trading_with_surpluses passed")


def test_watercourse_low_must_trade_up():
    """Test that Low distinctiveness watercourses must trade to better distinctiveness"""
    
    wb = openpyxl.Workbook()
    
    # Create Headline Results sheet
    ws_headline = wb.create_sheet("Headline Results")
    
    headers = ["Unit Type", "Target", "Baseline Units", "Units Required", "Unit Deficit"]
    for col, header in enumerate(headers, start=1):
        ws_headline.cell(row=5, column=col, value=header)
    
    data = [
        ["Habitat units", "10.00%", 0.00, 0.00, 0.00],
        ["Hedgerow units", "10.00%", 0.00, 0.00, 0.00],
        ["Watercourse units", "10.00%", 0.00, 0.00, 0.00],
    ]
    
    for row_idx, row_data in enumerate(data, start=6):
        for col_idx, value in enumerate(row_data, start=1):
            ws_headline.cell(row=row_idx, column=col_idx, value=value)
    
    # Create Trading Summary WaterCs sheet
    ws_water = wb.create_sheet("Trading Summary WaterCs")
    
    water_headers = ["Habitat", "Distinctiveness", "Project-wide unit change"]
    for col, header in enumerate(water_headers, start=1):
        ws_water.cell(row=1, column=col, value=header)
    
    # Test data: Low deficit with Low surplus (should NOT offset) and Medium surplus (should offset)
    water_data = [
        ["Culvert", "Low", -3.0],
        ["Culvert", "Low", 1.0],     # Should NOT offset (Low cannot offset Low)
        ["Culvert", "Medium", 1.5],   # Should offset
    ]
    
    for row_idx, row_data in enumerate(water_data, start=2):
        for col_idx, value in enumerate(row_data, start=1):
            ws_water.cell(row=row_idx, column=col_idx, value=value)
    
    # Remove default sheet
    if "Sheet" in wb.sheetnames:
        wb.remove(wb["Sheet"])
    
    # Save to BytesIO
    excel_buffer = io.BytesIO()
    wb.save(excel_buffer)
    excel_buffer.seek(0)
    
    # Create mock uploaded file
    mock_file = MockUploadedFile(excel_buffer)
    
    # Parse
    requirements = parse_metric_requirements(mock_file)
    
    # Check watercourse requirements
    water_req_df = requirements["watercourses"]
    
    # The Low deficit (-3.0) should only be partially offset by Medium surplus (+1.5)
    # Low surplus (+1.0) should NOT be used
    # Remaining deficit should be 3.0 - 1.5 = 1.5
    culvert_rows = water_req_df[water_req_df["habitat"].str.contains("Culvert", case=False, na=False)]
    culvert_rows = culvert_rows[~culvert_rows["habitat"].str.contains("Net Gain", case=False, na=False)]
    
    assert not culvert_rows.empty, "Low Culvert deficit should remain after partial offsetting"
    
    culvert_units = culvert_rows.iloc[0]["units"]
    expected_remaining = 3.0 - 1.5  # Low surplus not used, only Medium surplus
    assert abs(culvert_units - expected_remaining) < 0.0001, \
        f"Expected {expected_remaining} remaining units for Low Culvert deficit, got {culvert_units}"
    
    print("✅ test_watercourse_low_must_trade_up passed")


def test_watercourse_very_high_no_trading():
    """Test that Very High distinctiveness watercourses are not offset (bespoke compensation required)"""
    
    wb = openpyxl.Workbook()
    
    # Create Headline Results sheet
    ws_headline = wb.create_sheet("Headline Results")
    
    headers = ["Unit Type", "Target", "Baseline Units", "Units Required", "Unit Deficit"]
    for col, header in enumerate(headers, start=1):
        ws_headline.cell(row=5, column=col, value=header)
    
    data = [
        ["Habitat units", "10.00%", 0.00, 0.00, 0.00],
        ["Hedgerow units", "10.00%", 0.00, 0.00, 0.00],
        ["Watercourse units", "10.00%", 0.00, 0.00, 0.00],
    ]
    
    for row_idx, row_data in enumerate(data, start=6):
        for col_idx, value in enumerate(row_data, start=1):
            ws_headline.cell(row=row_idx, column=col_idx, value=value)
    
    # Create Trading Summary WaterCs sheet
    ws_water = wb.create_sheet("Trading Summary WaterCs")
    
    water_headers = ["Habitat", "Distinctiveness", "Project-wide unit change"]
    for col, header in enumerate(water_headers, start=1):
        ws_water.cell(row=1, column=col, value=header)
    
    # Test data: Very High deficit with Very High surplus (should NOT offset - bespoke compensation required)
    water_data = [
        ["Priority habitat", "Very High", -2.0],
        ["Priority habitat", "Very High", 3.0],  # Should NOT offset (bespoke compensation required)
    ]
    
    for row_idx, row_data in enumerate(water_data, start=2):
        for col_idx, value in enumerate(row_data, start=1):
            ws_water.cell(row=row_idx, column=col_idx, value=value)
    
    # Remove default sheet
    if "Sheet" in wb.sheetnames:
        wb.remove(wb["Sheet"])
    
    # Save to BytesIO
    excel_buffer = io.BytesIO()
    wb.save(excel_buffer)
    excel_buffer.seek(0)
    
    # Create mock uploaded file
    mock_file = MockUploadedFile(excel_buffer)
    
    # Parse
    requirements = parse_metric_requirements(mock_file)
    
    # Check watercourse requirements
    water_req_df = requirements["watercourses"]
    
    # Very High deficit should remain fully unmet (bespoke compensation required)
    priority_rows = water_req_df[water_req_df["habitat"].str.contains("Priority", case=False, na=False)]
    priority_rows = priority_rows[~priority_rows["habitat"].str.contains("Net Gain", case=False, na=False)]
    
    assert not priority_rows.empty, "Very High Priority habitat deficit should remain (bespoke compensation required)"
    
    priority_units = priority_rows.iloc[0]["units"]
    assert abs(priority_units - 2.0) < 0.0001, \
        f"Expected 2.0 units for Very High deficit (no offsetting), got {priority_units}"
    
    print("✅ test_watercourse_very_high_no_trading passed")


if __name__ == "__main__":
    print("=" * 60)
    print("Testing Watercourse Trading Rules")
    print("=" * 60)
    
    # Unit tests for can_offset_watercourse function
    test_can_offset_watercourse_very_high()
    test_can_offset_watercourse_high()
    test_can_offset_watercourse_medium()
    test_can_offset_watercourse_low()
    test_watercourse_habitat_normalization()
    
    # Integration tests with metric parsing
    test_watercourse_trading_with_surpluses()
    test_watercourse_low_must_trade_up()
    test_watercourse_very_high_no_trading()
    
    print("=" * 60)
    print("✅ ALL TESTS PASSED!")
    print("=" * 60)
