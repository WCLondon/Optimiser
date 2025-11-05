"""
Test that low and medium distinctiveness surpluses can offset Net Gain (Watercourses).

According to trading rules:
- Net Gain: Anything

This means ANY distinctiveness level should be able to offset Net Gain.
"""

import io
import pandas as pd
import openpyxl
from metric_reader import parse_metric_requirements


class MockUploadedFile:
    """Mock file object for testing"""
    def __init__(self, buffer):
        self.buffer = buffer
        self.name = "test_metric.xlsx"
    
    def read(self):
        return self.buffer.read()


def test_low_medium_surplus_offsets_net_gain():
    """
    Test that Low and Medium distinctiveness surpluses can offset Net Gain.
    
    Scenario:
    - Low surplus: 0.50 units
    - Medium surplus: 0.80 units
    - No deficits
    - Net Gain requirement: 10% of 2.50 = 0.25 units
    
    Expected:
    - Net Gain should be fully covered by Low/Medium surplus
    - Requirements should be empty
    """
    
    wb = openpyxl.Workbook()
    
    # Create Headline Results sheet
    ws_headline = wb.create_sheet("Headline Results")
    
    headers = ["Unit Type", "Target", "Baseline Units", "Units Required", "Unit Deficit"]
    for col, header in enumerate(headers, start=1):
        ws_headline.cell(row=5, column=col, value=header)
    
    data = [
        ["Habitat units", "10.00%", 0.00, 0.00, 0.00],
        ["Hedgerow units", "10.00%", 0.00, 0.00, 0.00],
        ["Watercourse units", "10.00%", 2.50, 2.75, 0.00],  # 10% Net Gain = 0.25 units
    ]
    
    for row_idx, row_data in enumerate(data, start=6):
        for col_idx, value in enumerate(row_data, start=1):
            ws_headline.cell(row=row_idx, column=col_idx, value=value)
    
    # Create Trading Summary Watercourses sheet
    ws_water = wb.create_sheet("Trading Summary Watercourses")
    
    water_headers = [
        "Feature",
        "Distinctiveness",
        "Project-wide unit change"
    ]
    
    for col, header in enumerate(water_headers, start=1):
        ws_water.cell(row=1, column=col, value=header)
    
    # Low distinctiveness surplus
    ws_water.cell(row=2, column=1, value="Ditches")
    ws_water.cell(row=2, column=2, value="Low")
    ws_water.cell(row=2, column=3, value=0.50)  # Low surplus
    
    # Medium distinctiveness surplus
    ws_water.cell(row=3, column=1, value="Other rivers and streams")
    ws_water.cell(row=3, column=2, value="Medium")
    ws_water.cell(row=3, column=3, value=0.80)  # Medium surplus
    
    # Remove default sheet
    if "Sheet" in wb.sheetnames:
        wb.remove(wb["Sheet"])
    
    # Save to BytesIO
    excel_buffer = io.BytesIO()
    wb.save(excel_buffer)
    excel_buffer.seek(0)
    
    # Create mock uploaded file
    mock_file = MockUploadedFile(excel_buffer)
    
    # Parse
    requirements = parse_metric_requirements(mock_file)
    
    print("=" * 90)
    print("TEST: LOW AND MEDIUM SURPLUS OFFSETS NET GAIN (WATERCOURSES)")
    print("=" * 90)
    print()
    print("📊 METRIC FILE STRUCTURE:")
    print("-" * 90)
    print("Watercourse habitats:")
    print("  - Low distinctiveness: Ditches (+0.50 units)")
    print("  - Medium distinctiveness: Other rivers and streams (+0.80 units)")
    print("  - Total surplus: 1.30 units")
    print()
    print("Net Gain requirement:")
    print("  - Baseline: 2.50 units")
    print("  - Target: 10%")
    print("  - Net Gain needed: 0.25 units")
    print()
    print("📋 EXPECTED BEHAVIOR:")
    print("-" * 90)
    print("Trading rule: Net Gain can be offset by ANYTHING")
    print("Expected: Net Gain (0.25) offset by Low/Medium surplus")
    print("Result: No requirements (all covered)")
    print()
    
    # Check watercourse requirements
    water_req_df = requirements["watercourses"]
    
    print("📤 ACTUAL RESULT:")
    print("-" * 90)
    print(f"Watercourse requirements count: {len(water_req_df)}")
    
    if not water_req_df.empty:
        print("\nRequirements:")
        total_req = 0.0
        for _, row in water_req_df.iterrows():
            print(f"  - {row['habitat']}: {row['units']:.4f} units")
            total_req += row['units']
        print(f"\nTotal requirements: {total_req:.4f} units")
        
        # Check if Net Gain is in requirements
        has_net_gain = any("Net Gain" in str(h) for h in water_req_df["habitat"])
        if has_net_gain:
            net_gain_row = water_req_df[water_req_df["habitat"].str.contains("Net Gain", na=False)]
            net_gain_units = net_gain_row["units"].iloc[0] if not net_gain_row.empty else 0.0
            print()
            print(f"❌ FAILURE: Net Gain requirement remains: {net_gain_units:.4f} units")
            print(f"   Expected: 0.00 units (should be offset by Low/Medium surplus)")
            print()
            print("   This indicates Low/Medium surplus is NOT offsetting Net Gain!")
            raise AssertionError(f"Net Gain not offset by Low/Medium surplus. Remaining: {net_gain_units:.4f}")
    else:
        print("  (empty - all covered)")
        print()
        print("✅ SUCCESS: Net Gain fully offset by Low/Medium surplus!")
    
    print()
    print("=" * 90)
    return True


def test_only_low_surplus_offsets_net_gain():
    """
    Test that ONLY Low distinctiveness surplus can offset Net Gain.
    """
    
    wb = openpyxl.Workbook()
    
    # Create Headline Results sheet
    ws_headline = wb.create_sheet("Headline Results")
    
    headers = ["Unit Type", "Target", "Baseline Units", "Units Required", "Unit Deficit"]
    for col, header in enumerate(headers, start=1):
        ws_headline.cell(row=5, column=col, value=header)
    
    data = [
        ["Habitat units", "10.00%", 0.00, 0.00, 0.00],
        ["Hedgerow units", "10.00%", 0.00, 0.00, 0.00],
        ["Watercourse units", "10.00%", 1.00, 1.10, 0.00],  # 10% Net Gain = 0.10 units
    ]
    
    for row_idx, row_data in enumerate(data, start=6):
        for col_idx, value in enumerate(row_data, start=1):
            ws_headline.cell(row=row_idx, column=col_idx, value=value)
    
    # Create Trading Summary Watercourses sheet
    ws_water = wb.create_sheet("Trading Summary Watercourses")
    
    water_headers = ["Feature", "Distinctiveness", "Project-wide unit change"]
    
    for col, header in enumerate(water_headers, start=1):
        ws_water.cell(row=1, column=col, value=header)
    
    # Only Low distinctiveness surplus
    ws_water.cell(row=2, column=1, value="Ditches")
    ws_water.cell(row=2, column=2, value="Low")
    ws_water.cell(row=2, column=3, value=0.20)  # Low surplus (more than needed)
    
    # Remove default sheet
    if "Sheet" in wb.sheetnames:
        wb.remove(wb["Sheet"])
    
    # Save to BytesIO
    excel_buffer = io.BytesIO()
    wb.save(excel_buffer)
    excel_buffer.seek(0)
    
    # Create mock uploaded file
    mock_file = MockUploadedFile(excel_buffer)
    
    # Parse
    requirements = parse_metric_requirements(mock_file)
    
    print()
    print("=" * 90)
    print("TEST: ONLY LOW SURPLUS OFFSETS NET GAIN")
    print("=" * 90)
    print()
    print("📊 SCENARIO:")
    print("-" * 90)
    print("  - Low surplus: 0.20 units")
    print("  - Net Gain needed: 0.10 units")
    print("  - Expected: Net Gain covered by Low surplus")
    print()
    
    water_req_df = requirements["watercourses"]
    
    print("📤 RESULT:")
    print("-" * 90)
    
    if not water_req_df.empty:
        print(f"Requirements count: {len(water_req_df)}")
        for _, row in water_req_df.iterrows():
            print(f"  - {row['habitat']}: {row['units']:.4f} units")
        
        has_net_gain = any("Net Gain" in str(h) for h in water_req_df["habitat"])
        if has_net_gain:
            print()
            print("❌ FAILURE: Low surplus did not offset Net Gain!")
            raise AssertionError("Low surplus should offset Net Gain")
    else:
        print("✅ SUCCESS: Net Gain covered by Low surplus")
    
    print()
    print("=" * 90)
    return True


if __name__ == "__main__":
    print()
    all_passed = True
    
    try:
        test_low_medium_surplus_offsets_net_gain()
    except Exception as e:
        print(f"\n❌ test_low_medium_surplus_offsets_net_gain failed: {e}\n")
        import traceback
        traceback.print_exc()
        all_passed = False
    
    try:
        test_only_low_surplus_offsets_net_gain()
    except Exception as e:
        print(f"\n❌ test_only_low_surplus_offsets_net_gain failed: {e}\n")
        import traceback
        traceback.print_exc()
        all_passed = False
    
    if all_passed:
        print()
        print("=" * 90)
        print("✅ ALL TESTS PASSED!")
        print("=" * 90)
        print()
    else:
        print()
        print("=" * 90)
        print("❌ SOME TESTS FAILED - Low/Medium surplus NOT offsetting Net Gain!")
        print("=" * 90)
        print()
        exit(1)
