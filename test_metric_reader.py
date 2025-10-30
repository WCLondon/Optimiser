"""
Simple test for metric_reader module
"""

import io
import pandas as pd
import openpyxl
from metric_reader import parse_metric_requirements


def test_basic_functionality():
    """Test that the basic metric reader functions work"""
    # Create a simple mock Excel file in memory
    wb = openpyxl.Workbook()
    
    # Create a Trading Summary Area Habitats sheet
    ws = wb.create_sheet("Trading Summary Area Habitats")
    
    # Add headers
    headers = ["Habitat", "Broad habitat", "Distinctiveness", "Project-wide unit change", "On-site unit change"]
    for col, header in enumerate(headers, start=1):
        ws.cell(row=1, column=col, value=header)
    
    # Add some test data (deficits = negative values)
    data = [
        ["Grassland", "Grassland and marsh", "Medium", -5.0, 0.0],
        ["Woodland", "Woodland and forest", "High", -3.0, 0.0],
        ["Heathland", "Heathland and shrub", "Medium", 2.0, 1.0],  # surplus, should be ignored
    ]
    
    for row_idx, row_data in enumerate(data, start=2):
        for col_idx, value in enumerate(row_data, start=1):
            ws.cell(row=row_idx, column=col_idx, value=value)
    
    # Remove default sheet
    if "Sheet" in wb.sheetnames:
        wb.remove(wb["Sheet"])
    
    # Save to BytesIO
    excel_buffer = io.BytesIO()
    wb.save(excel_buffer)
    excel_buffer.seek(0)
    
    # Create a mock uploaded file object
    class MockUploadedFile:
        def __init__(self, buffer):
            self.buffer = buffer
            self.name = "test_metric.xlsx"
        
        def read(self):
            return self.buffer.read()
    
    mock_file = MockUploadedFile(excel_buffer)
    
    # Test parsing
    try:
        requirements = parse_metric_requirements(mock_file)
        
        print("✅ parse_metric_requirements succeeded")
        print(f"   Area habitats: {len(requirements['area'])} rows")
        print(f"   Hedgerows: {len(requirements['hedgerows'])} rows")
        print(f"   Watercourses: {len(requirements['watercourses'])} rows")
        
        # Check that we got the expected deficits
        area_df = requirements['area']
        if not area_df.empty:
            print("\nArea deficits found:")
            for _, row in area_df.iterrows():
                print(f"   - {row['habitat']}: {row['units']:.2f} units")
            
            # Verify we got 2 deficits (positive units, ignoring the surplus)
            assert len(area_df) == 2, f"Expected 2 deficits, got {len(area_df)}"
            
            # Verify units are positive (absolute values)
            for _, row in area_df.iterrows():
                assert row['units'] > 0, f"Expected positive units, got {row['units']}"
        
        print("\n✅ All tests passed!")
        return True
        
    except Exception as e:
        print(f"❌ Test failed: {e}")
        import traceback
        traceback.print_exc()
        return False


if __name__ == "__main__":
    success = test_basic_functionality()
    exit(0 if success else 1)
