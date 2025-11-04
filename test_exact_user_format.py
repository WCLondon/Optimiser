"""
Test with EXACT format from user's metric file - section header format
The user reports that parsed requirements show:
- Species-rich native hedgerow: 0.554
- Native hedgerow with trees: 1.035  
- Non-native and ornamental hedgerow: 0.246

This means deficits are NOT being offset despite High surplus of 2.70
"""

import io
import pandas as pd
import openpyxl
from metric_reader import parse_metric_requirements, normalise_requirements, open_metric_workbook


class MockUploadedFile:
    """Mock file object for testing"""
    def __init__(self, buffer):
        self.buffer = buffer
        self.name = "test_metric.xlsx"
    
    def read(self):
        return self.buffer.read()


def test_exact_user_format_section_headers():
    """
    Reproduce the exact issue with section header format from user's metric file.
    
    The user's metric has:
    - Section headers like "High Distinctiveness" 
    - Sub-header row with "Habitat group", "Project-wide unit change"
    - No "Distinctiveness" column
    """
    
    wb = openpyxl.Workbook()
    
    # Create Headline Results sheet
    ws_headline = wb.create_sheet("Headline Results")
    
    headers = ["Unit Type", "Target", "Baseline Units", "Units Required", "Unit Deficit"]
    for col, header in enumerate(headers, start=1):
        ws_headline.cell(row=5, column=col, value=header)
    
    data = [
        ["Habitat units", "10.00%", 1.81, 2.00, 0.00],
        ["Hedgerow units", "10.00%", 1.84, 2.03, 0.00],
        ["Watercourse units", "10.00%", 0.00, 0.00, 0.00],
    ]
    
    for row_idx, row_data in enumerate(data, start=6):
        for col_idx, value in enumerate(row_data, start=1):
            ws_headline.cell(row=row_idx, column=col_idx, value=value)
    
    # Create Trading Summary Hedgerows sheet - EXACT structure from user
    ws_hedge = wb.create_sheet("Trading Summary Hedgerows")
    
    row_num = 1
    
    # Very High Distinctiveness section
    ws_hedge.cell(row=row_num, column=1, value="Very High Distinctiveness")
    row_num += 1
    ws_hedge.cell(row=row_num, column=1, value="Habitat group")
    ws_hedge.cell(row=row_num, column=2, value="On-site unit change")
    ws_hedge.cell(row=row_num, column=3, value="Off-site unit change")
    ws_hedge.cell(row=row_num, column=4, value="Project-wide unit change")
    row_num += 1
    ws_hedge.cell(row=row_num, column=1, value="Species-rich native hedgerow with trees - associated with bank or ditch")
    ws_hedge.cell(row=row_num, column=2, value=0.00)
    ws_hedge.cell(row=row_num, column=3, value=0.00)
    ws_hedge.cell(row=row_num, column=4, value=0.00)
    row_num += 1
    # Totals row
    ws_hedge.cell(row=row_num, column=2, value=0.00)
    ws_hedge.cell(row=row_num, column=3, value=0.00)
    ws_hedge.cell(row=row_num, column=4, value=0.00)
    row_num += 2  # Skip a row
    
    # High Distinctiveness section
    ws_hedge.cell(row=row_num, column=1, value="High Distinctiveness")
    row_num += 1
    ws_hedge.cell(row=row_num, column=1, value="Habitat group")
    ws_hedge.cell(row=row_num, column=2, value="On-site unit change")
    ws_hedge.cell(row=row_num, column=3, value="Off-site unit change")
    ws_hedge.cell(row=row_num, column=4, value="Project wide unit change")  # Note: "Project wide" not "Project-wide"
    row_num += 1
    ws_hedge.cell(row=row_num, column=1, value="Species-rich native hedgerow with trees")
    ws_hedge.cell(row=row_num, column=2, value=2.70)
    ws_hedge.cell(row=row_num, column=3, value=0.00)
    ws_hedge.cell(row=row_num, column=4, value=2.70)
    row_num += 1
    # More habitat rows
    ws_hedge.cell(row=row_num, column=1, value="Species-rich native hedgerow - associated with bank or ditch")
    ws_hedge.cell(row=row_num, column=2, value=0.00)
    ws_hedge.cell(row=row_num, column=3, value=0.00)
    ws_hedge.cell(row=row_num, column=4, value=0.00)
    row_num += 1
    ws_hedge.cell(row=row_num, column=1, value="Native hedgerow with trees - associated with bank or ditch")
    ws_hedge.cell(row=row_num, column=2, value=0.00)
    ws_hedge.cell(row=row_num, column=3, value=0.00)
    ws_hedge.cell(row=row_num, column=4, value=0.00)
    row_num += 1
    # Totals row
    ws_hedge.cell(row=row_num, column=2, value=2.70)
    ws_hedge.cell(row=row_num, column=3, value=0.00)
    ws_hedge.cell(row=row_num, column=4, value=2.70)
    row_num += 2
    
    # Medium Distinctiveness section
    ws_hedge.cell(row=row_num, column=1, value="Medium Distinctiveness")
    row_num += 1
    ws_hedge.cell(row=row_num, column=1, value="Habitat group")
    ws_hedge.cell(row=row_num, column=2, value="On-site unit change")
    ws_hedge.cell(row=row_num, column=3, value="Off-site unit change")
    ws_hedge.cell(row=row_num, column=4, value="Project wide unit change")
    row_num += 1
    ws_hedge.cell(row=row_num, column=1, value="Species-rich native hedgerow")
    ws_hedge.cell(row=row_num, column=2, value=-0.55)
    ws_hedge.cell(row=row_num, column=3, value=0.00)
    ws_hedge.cell(row=row_num, column=4, value=-0.55)
    row_num += 1
    ws_hedge.cell(row=row_num, column=1, value="Native hedgerow - associated with bank or ditch")
    ws_hedge.cell(row=row_num, column=2, value=0.00)
    ws_hedge.cell(row=row_num, column=3, value=0.00)
    ws_hedge.cell(row=row_num, column=4, value=0.00)
    row_num += 1
    ws_hedge.cell(row=row_num, column=1, value="Native hedgerow with trees")
    ws_hedge.cell(row=row_num, column=2, value=-1.04)
    ws_hedge.cell(row=row_num, column=3, value=0.00)
    ws_hedge.cell(row=row_num, column=4, value=-1.04)
    row_num += 1
    # More zero rows
    ws_hedge.cell(row=row_num, column=1, value="Ecologically valuable line of trees")
    ws_hedge.cell(row=row_num, column=4, value=0.00)
    row_num += 1
    ws_hedge.cell(row=row_num, column=1, value="Ecologically valuable line of trees - associated with bank or ditch")
    ws_hedge.cell(row=row_num, column=4, value=0.00)
    row_num += 1
    # Totals row
    ws_hedge.cell(row=row_num, column=2, value=-1.59)
    ws_hedge.cell(row=row_num, column=3, value=0.00)
    ws_hedge.cell(row=row_num, column=4, value=-1.59)
    row_num += 2
    
    # Low Distinctiveness section
    ws_hedge.cell(row=row_num, column=1, value="Low Distinctiveness")
    row_num += 1
    ws_hedge.cell(row=row_num, column=1, value="Habitat group")
    ws_hedge.cell(row=row_num, column=4, value="Project wide unit change")
    row_num += 1
    ws_hedge.cell(row=row_num, column=1, value="Native hedgerow")
    ws_hedge.cell(row=row_num, column=4, value=0.00)
    row_num += 1
    ws_hedge.cell(row=row_num, column=1, value="Line of trees")
    ws_hedge.cell(row=row_num, column=4, value=0.00)
    row_num += 1
    ws_hedge.cell(row=row_num, column=1, value="Line of trees - associated with bank or ditch")
    ws_hedge.cell(row=row_num, column=4, value=0.00)
    row_num += 1
    # Totals row
    ws_hedge.cell(row=row_num, column=4, value=0.00)
    row_num += 2
    
    # Very Low Distinctiveness section
    ws_hedge.cell(row=row_num, column=1, value="Very Low Distinctiveness")
    row_num += 1
    ws_hedge.cell(row=row_num, column=1, value="Habitat group")
    ws_hedge.cell(row=row_num, column=2, value="On-site unit change")
    ws_hedge.cell(row=row_num, column=3, value="Off-site unit change")
    ws_hedge.cell(row=row_num, column=4, value="Project wide unit change")
    row_num += 1
    ws_hedge.cell(row=row_num, column=1, value="Non-native and ornamental hedgerow")
    ws_hedge.cell(row=row_num, column=2, value=-0.25)
    ws_hedge.cell(row=row_num, column=3, value=0.00)
    ws_hedge.cell(row=row_num, column=4, value=-0.25)
    row_num += 1
    # Totals row
    ws_hedge.cell(row=row_num, column=2, value=-0.25)
    ws_hedge.cell(row=row_num, column=3, value=0.00)
    ws_hedge.cell(row=row_num, column=4, value=-0.25)
    
    # Remove default sheet
    if "Sheet" in wb.sheetnames:
        wb.remove(wb["Sheet"])
    
    # Save to BytesIO
    excel_buffer = io.BytesIO()
    wb.save(excel_buffer)
    excel_buffer.seek(0)
    
    # Create mock uploaded file
    mock_file = MockUploadedFile(excel_buffer)
    
    # First, debug what normalise_requirements extracts
    xls = open_metric_workbook(mock_file)
    HEDGE_SHEETS = ["Trading Summary Hedgerows"]
    hedge_norm, colmap, sheet_name = normalise_requirements(xls, HEDGE_SHEETS, "Hedgerows")
    
    print("\n" + "="*90)
    print("DEBUG: DISTINCTIVENESS EXTRACTION FROM SECTION HEADERS")
    print("="*90)
    print("\nExtracted data:")
    print(hedge_norm[["habitat", "distinctiveness", "project_wide_change"]].to_string(index=False))
    
    na_count = hedge_norm["distinctiveness"].isna().sum()
    print(f"\nRows with NA distinctiveness: {na_count}")
    
    if na_count > 0:
        print("\n❌ PROBLEM FOUND: Some habitats have NA distinctiveness!")
        print("This would prevent apply_hedgerow_offsets() from working correctly.")
        print("\nHabitats with NA distinctiveness:")
        na_rows = hedge_norm[hedge_norm["distinctiveness"].isna()]
        for _, row in na_rows.iterrows():
            print(f"  - {row['habitat']}: {row['project_wide_change']}")
    
    # Now run full parse
    excel_buffer.seek(0)
    requirements = parse_metric_requirements(mock_file)
    
    print("\n" + "="*90)
    print("PARSED REQUIREMENTS FROM parse_metric_requirements()")
    print("="*90)
    
    hedge_req_df = requirements["hedgerows"]
    
    if hedge_req_df.empty:
        print("✅ requirements['hedgerows'] is EMPTY - deficits were offset correctly!")
    else:
        print("❌ requirements['hedgerows'] is NOT EMPTY:")
        print()
        print(hedge_req_df.to_string(index=False))
        print()
        print("USER REPORTED SEEING:")
        print("  Species-rich native hedgerow: 0.5540587249979952")
        print("  Native hedgerow with trees: 1.035")
        print("  Non-native and ornamental hedgerow: 0.246")
        print()
        print("This matches the deficits NOT being offset!")
    
    print("\n" + "="*90)
    
    return hedge_req_df.empty


if __name__ == "__main__":
    success = test_exact_user_format_section_headers()
    exit(0 if success else 1)
