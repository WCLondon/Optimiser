"""
Test with Trading Summary that includes Summary columns on the right.

The issue shows a table like this:
Habitat group | On-site unit change | Off-site unit change | Project-wide unit change | [blank] | Very High Distinctiveness Summary
                                                                                                  Very High Distinctiveness Units available...

These extra columns on the right might interfere with header detection or distinctiveness extraction.
"""

import io
import pandas as pd
import openpyxl
from metric_reader import parse_metric_requirements


class MockUploadedFile:
    """Mock file object for testing"""
    def __init__(self, buffer):
        self.buffer = buffer
        self.name = "test_metric.xlsx"
    
    def read(self):
        return self.buffer.read()


def test_with_summary_columns():
    """
    Test with summary columns on the right side of the table
    """
    wb = openpyxl.Workbook()
    
    # Create Headline Results sheet  
    ws_headline = wb.create_sheet("Headline Results")
    
    headers = ["Unit Type", "Target", "Baseline Units", "Units Required", "Unit Deficit"]
    for col, header in enumerate(headers, start=1):
        ws_headline.cell(row=5, column=col, value=header)
    
    data = [
        ["Habitat units", "10.00%", 10.00, 11.00, 0.00],
        ["Hedgerow units", "10.00%", 0.00, 0.00, 0.00],
        ["Watercourse units", "10.00%", 0.00, 0.00, 0.00],
    ]
    
    for row_idx, row_data in enumerate(data, start=6):
        for col_idx, value in enumerate(row_data, start=1):
            ws_headline.cell(row=row_idx, column=col_idx, value=value)
    
    # Create Trading Summary Hedgerows sheet with Summary columns
    ws_hedge = wb.create_sheet("Trading Summary Hedgerows")
    
    row = 1
    
    # Add Trading Rule at top
    ws_hedge.cell(row=row, column=1, value="Trading Rule")
    ws_hedge.cell(row=row, column=4, value="Trading Satisfied?")
    row += 2
    
    # Medium Distinctiveness section header WITH summary info on right
    ws_hedge.cell(row=row, column=1, value="Medium Distinctiveness")
    ws_hedge.cell(row=row, column=6, value="Medium Distinctiveness Summary")
    row += 1
    
    # Headers row
    ws_hedge.cell(row=row, column=1, value="Habitat group")
    ws_hedge.cell(row=row, column=2, value="On-site unit change")
    ws_hedge.cell(row=row, column=3, value="Off-site unit change")
    ws_hedge.cell(row=row, column=4, value="Project-wide unit change")
    ws_hedge.cell(row=row, column=5, value="")  # blank
    ws_hedge.cell(row=row, column=6, value="Units available from higher distinctiveness habitats")
    ws_hedge.cell(row=row, column=7, value=0.00)
    row += 1
    
    # Data rows
    ws_hedge.cell(row=row, column=1, value="Species-rich native hedgerow")
    ws_hedge.cell(row=row, column=2, value=1.47)
    ws_hedge.cell(row=row, column=3, value=0.00)
    ws_hedge.cell(row=row, column=4, value=1.47)
    ws_hedge.cell(row=row, column=5, value="✓")  # checkmark
    ws_hedge.cell(row=row, column=6, value="Medium Distinctiveness net change in units")
    ws_hedge.cell(row=row, column=7, value=1.47)
    row += 1
    
    ws_hedge.cell(row=row, column=1, value="Native hedgerow - associated with bank or ditch")
    ws_hedge.cell(row=row, column=2, value=0.00)
    ws_hedge.cell(row=row, column=3, value=0.00)
    ws_hedge.cell(row=row, column=4, value=0.00)
    ws_hedge.cell(row=row, column=6, value="Cumulative availability of units")
    ws_hedge.cell(row=row, column=7, value=1.47)
    row += 1
    
    # Total row
    ws_hedge.cell(row=row, column=1, value="")
    ws_hedge.cell(row=row, column=2, value=1.47)
    ws_hedge.cell(row=row, column=3, value=0.00)
    ws_hedge.cell(row=row, column=4, value=1.47)
    row += 3
    
    # Low Distinctiveness section header WITH summary info on right
    ws_hedge.cell(row=row, column=1, value="Low Distinctiveness")
    ws_hedge.cell(row=row, column=6, value="Low Distinctiveness Summary")
    row += 1
    
    # Headers row
    ws_hedge.cell(row=row, column=1, value="Habitat group")
    ws_hedge.cell(row=row, column=2, value="On-site unit change")
    ws_hedge.cell(row=row, column=3, value="Off-site unit change")
    ws_hedge.cell(row=row, column=4, value="Project-wide unit change")
    ws_hedge.cell(row=row, column=5, value="")  # blank
    ws_hedge.cell(row=row, column=6, value="Low Distinctiveness net change in units")
    ws_hedge.cell(row=row, column=7, value=-0.70)
    row += 1
    
    # Data rows
    ws_hedge.cell(row=row, column=1, value="Native hedgerow")
    ws_hedge.cell(row=row, column=2, value=-0.70)
    ws_hedge.cell(row=row, column=3, value=0.00)
    ws_hedge.cell(row=row, column=4, value=-0.70)
    ws_hedge.cell(row=row, column=5, value="⚠")  # warning symbol
    ws_hedge.cell(row=row, column=6, value="Cumulative availability of units")
    ws_hedge.cell(row=row, column=7, value=0.77)
    row += 1
    
    # Total row
    ws_hedge.cell(row=row, column=1, value="")
    ws_hedge.cell(row=row, column=2, value=-0.70)
    ws_hedge.cell(row=row, column=3, value=0.00)
    ws_hedge.cell(row=row, column=4, value=-0.70)
    
    # Remove default sheet
    if "Sheet" in wb.sheetnames:
        wb.remove(wb["Sheet"])
    
    # Save to BytesIO
    excel_buffer = io.BytesIO()
    wb.save(excel_buffer)
    excel_buffer.seek(0)
    
    # Create mock uploaded file
    mock_file = MockUploadedFile(excel_buffer)
    
    # Parse with debug
    from metric_reader import open_metric_workbook, normalise_requirements
    xls = open_metric_workbook(mock_file)
    HEDGE_SHEETS = ["Trading Summary Hedgerows"]
    hedge_norm, colmap, sheet = normalise_requirements(xls, HEDGE_SHEETS, "Hedgerows")
    
    print("\n" + "="*80)
    print("TEST: Trading Summary with Summary columns on the right")
    print("="*80)
    print("\nDEBUG - Parsed data:")
    print(hedge_norm[['habitat', 'distinctiveness', 'project_wide_change']])
    print("\nDEBUG - Distinctiveness values:")
    for idx, row in hedge_norm.iterrows():
        print(f"  {row['habitat']}: '{row['distinctiveness']}'")
    
    # Now test full parsing
    excel_buffer.seek(0)
    mock_file = MockUploadedFile(excel_buffer)
    print("\nThis mimics the actual format from the issue which has:")
    print("  - 'Medium Distinctiveness Summary' header in column 6")
    print("  - Summary metrics like 'Units available...' and 'Cumulative availability...'")
    print("  - These extra columns might interfere with parsing")
    print("\nExpected:")
    print("  - Medium: Species-rich native hedgerow: +1.47")
    print("  - Low: Native hedgerow: -0.70")
    print("  - Deficit should be offset by surplus")
    
    requirements = parse_metric_requirements(mock_file)
    
    # Verify results
    hedge_req_df = requirements["hedgerows"]
    
    print("\n" + "-"*80)
    print("RESULT:")
    print("-"*80)
    if hedge_req_df.empty:
        print("  (empty - all requirements covered)")
    else:
        print(hedge_req_df.to_string(index=False))
    
    # Check if Native hedgerow appears
    native_rows = hedge_req_df[
        hedge_req_df["habitat"].str.contains("Native hedgerow", case=False, na=False) & 
        ~hedge_req_df["habitat"].str.contains("Species-rich", case=False, na=False) &
        ~hedge_req_df["habitat"].str.contains("with trees", case=False, na=False) &
        ~hedge_req_df["habitat"].str.contains("associated", case=False, na=False)
    ]
    
    print("\n" + "="*80)
    if not native_rows.empty:
        print("❌ ISSUE REPRODUCED!")
        print(f"   'Native hedgerow' appears with {native_rows.iloc[0]['units']:.6f} units")
        return False
    else:
        print("✅ WORKING CORRECTLY!")
        return True


if __name__ == "__main__":
    success = test_with_summary_columns()
    exit(0 if success else 1)
