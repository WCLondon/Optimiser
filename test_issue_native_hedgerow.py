"""
Test for the specific issue reported: Native hedgerow deficit showing up 
despite Species-rich native hedgerow surplus.

Issue description:
- Medium: Species-rich native hedgerow: +1.47 units (surplus)
- Low: Native hedgerow: -0.70 units (deficit)

Expected: The Low deficit should be offset by the Medium surplus
Actual: Native hedgerow showing as 0.6996 units in requirements

This tests the hedgerow trading rule:
- Medium, Low, Very Low: Same distinctiveness or better can offset
- So Medium surplus SHOULD offset Low deficit
"""

import io
import pandas as pd
import openpyxl
from metric_reader import parse_metric_requirements


class MockUploadedFile:
    """Mock file object for testing"""
    def __init__(self, buffer):
        self.buffer = buffer
        self.name = "test_metric.xlsx"
    
    def read(self):
        return self.buffer.read()


def test_native_hedgerow_issue():
    """
    Test the exact scenario from the GitHub issue:
    - Medium: Species-rich native hedgerow: +1.47 units (surplus)
    - Low: Native hedgerow: -0.70 units (deficit)
    
    Expected: Deficit should be fully offset by surplus
    """
    
    wb = openpyxl.Workbook()
    
    # Create Headline Results sheet
    ws_headline = wb.create_sheet("Headline Results")
    
    headers = ["Unit Type", "Target", "Baseline Units", "Units Required", "Unit Deficit"]
    for col, header in enumerate(headers, start=1):
        ws_headline.cell(row=5, column=col, value=header)
    
    # Use similar values to the issue
    data = [
        ["Habitat units", "10.00%", 10.00, 11.00, 0.00],
        ["Hedgerow units", "10.00%", 0.70, 0.77, 0.00],
        ["Watercourse units", "10.00%", 0.00, 0.00, 0.00],
    ]
    
    for row_idx, row_data in enumerate(data, start=6):
        for col_idx, value in enumerate(row_data, start=1):
            ws_headline.cell(row=row_idx, column=col_idx, value=value)
    
    # Create Trading Summary Hedgerows sheet
    ws_hedge = wb.create_sheet("Trading Summary Hedgerows")
    
    hedge_headers = ["Habitat", "Distinctiveness", "Project-wide unit change"]
    for col, header in enumerate(hedge_headers, start=1):
        ws_hedge.cell(row=1, column=col, value=header)
    
    # Data from the GitHub issue
    hedge_data = [
        ["Species-rich native hedgerow", "Medium", 1.47],
        ["Native hedgerow", "Low", -0.70],
    ]
    
    for row_idx, row_data in enumerate(hedge_data, start=2):
        for col_idx, value in enumerate(row_data, start=1):
            ws_hedge.cell(row=row_idx, column=col_idx, value=value)
    
    # Remove default sheet
    if "Sheet" in wb.sheetnames:
        wb.remove(wb["Sheet"])
    
    # Save to BytesIO
    excel_buffer = io.BytesIO()
    wb.save(excel_buffer)
    excel_buffer.seek(0)
    
    # Create mock uploaded file
    mock_file = MockUploadedFile(excel_buffer)
    
    # Parse
    requirements = parse_metric_requirements(mock_file)
    
    # Verify results
    hedge_req_df = requirements["hedgerows"]
    
    print("\n" + "="*60)
    print("Test: Native hedgerow issue from GitHub")
    print("="*60)
    print("\nInput:")
    print("  Medium: Species-rich native hedgerow: +1.47 units")
    print("  Low: Native hedgerow: -0.70 units")
    print("  Net gain requirement: 0.70 × 10% = 0.07 units")
    print("\nExpected:")
    print("  Total surplus: 1.47 units")
    print("  Low deficit (0.70) should be offset by Medium surplus")
    print("  Remaining surplus: 1.47 - 0.70 = 0.77 units")
    print("  Net gain (0.07) also covered by remaining surplus")
    print("  No Native hedgerow in requirements!")
    
    print("\nActual requirements:")
    if hedge_req_df.empty:
        print("  (empty - all requirements covered by on-site surplus)")
    else:
        print(hedge_req_df)
        print(f"\n  Total units: {hedge_req_df['units'].sum():.4f}")
    
    # Check that Native hedgerow deficit is NOT in requirements
    native_rows = hedge_req_df[hedge_req_df["habitat"].str.contains("Native hedgerow", case=False, na=False) & 
                                ~hedge_req_df["habitat"].str.contains("Species-rich", case=False, na=False)]
    
    if not native_rows.empty:
        print(f"\n❌ ISSUE REPRODUCED: Native hedgerow showing {native_rows.iloc[0]['units']:.4f} units")
        print("   Medium surplus should have offset Low deficit!")
        return False
    
    print("\n✅ Native hedgerow deficit correctly offset by Species-rich surplus")
    print("✅ Issue resolved!")
    
    return True


if __name__ == "__main__":
    success = test_native_hedgerow_issue()
    exit(0 if success else 1)
