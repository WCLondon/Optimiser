"""
Test what happens when distinctiveness extraction fails (e.g., missing section headers)

This simulates the user's issue where deficits appear in requirements
because distinctiveness values are not being extracted correctly.
"""

import io
import pandas as pd
import openpyxl
import warnings
from metric_reader import parse_metric_requirements


class MockUploadedFile:
    """Mock file object for testing"""
    def __init__(self, buffer):
        self.buffer = buffer
        self.name = "test_metric.xlsx"
    
    def read(self):
        return self.buffer.read()


def test_failed_distinctiveness_extraction():
    """
    Test scenario where distinctiveness extraction fails.
    
    This happens when:
    - No "Distinctiveness" column exists
    - Section headers don't contain distinctiveness keywords
    - Section headers are formatted differently
    
    Expected behavior:
    - Deficits appear in requirements (not offset)
    - Warning is issued about undefined distinctiveness
    """
    
    wb = openpyxl.Workbook()
    
    # Create Headline Results sheet
    ws_headline = wb.create_sheet("Headline Results")
    
    headers = ["Unit Type", "Target", "Baseline Units", "Units Required", "Unit Deficit"]
    for col, header in enumerate(headers, start=1):
        ws_headline.cell(row=5, column=col, value=header)
    
    data = [
        ["Habitat units", "10.00%", 1.81, 2.00, 0.00],
        ["Hedgerow units", "10.00%", 1.84, 2.03, 0.00],
        ["Watercourse units", "10.00%", 0.00, 0.00, 0.00],
    ]
    
    for row_idx, row_data in enumerate(data, start=6):
        for col_idx, value in enumerate(row_data, start=1):
            ws_headline.cell(row=row_idx, column=col_idx, value=value)
    
    # Create Trading Summary Hedgerows sheet WITHOUT proper distinctiveness headers
    # This simulates a malformed or non-standard metric file
    ws_hedge = wb.create_sheet("Trading Summary Hedgerows")
    
    # Just a simple table without section headers for distinctiveness
    hedge_headers = ["Habitat group", "Project-wide unit change"]
    for col, header in enumerate(hedge_headers, start=1):
        ws_hedge.cell(row=1, column=col, value=header)
    
    hedge_data = [
        ["Species-rich native hedgerow with trees", 2.70],
        ["Species-rich native hedgerow", -0.55],
        ["Native hedgerow with trees", -1.04],
        ["Non-native and ornamental hedgerow", -0.25],
    ]
    
    for row_idx, row_data in enumerate(hedge_data, start=2):
        for col_idx, value in enumerate(row_data, start=1):
            ws_hedge.cell(row=row_idx, column=col_idx, value=value)
    
    # Remove default sheet
    if "Sheet" in wb.sheetnames:
        wb.remove(wb["Sheet"])
    
    # Save to BytesIO
    excel_buffer = io.BytesIO()
    wb.save(excel_buffer)
    excel_buffer.seek(0)
    
    # Create mock uploaded file
    mock_file = MockUploadedFile(excel_buffer)
    
    # Parse with warnings captured
    with warnings.catch_warnings(record=True) as w:
        warnings.simplefilter("always")
        requirements = parse_metric_requirements(mock_file)
        
        # Check if warning was issued
        hedgerow_warnings = [warning for warning in w if "Hedgerows" in str(warning.message) and "distinctiveness" in str(warning.message).lower()]
        
        if hedgerow_warnings:
            print("\n" + "="*90)
            print("✅ WARNING ISSUED AS EXPECTED:")
            print("="*90)
            for warning in hedgerow_warnings:
                print(f"  {warning.message}")
        else:
            print("\n" + "="*90)
            print("⚠️  NO WARNING ISSUED (unexpected)")
            print("="*90)
    
    print("\n" + "="*90)
    print("TEST: FAILED DISTINCTIVENESS EXTRACTION")
    print("="*90)
    
    print("\n📊 SCENARIO:")
    print("-" * 90)
    print("Metric file WITHOUT proper distinctiveness information:")
    print("  - No 'Distinctiveness' column")
    print("  - No section headers like 'High Distinctiveness'")
    print("  - Just raw habitat names and values")
    
    print("\n📋 INPUT:")
    print("-" * 90)
    print("Species-rich native hedgerow with trees: +2.70 (surplus)")
    print("Species-rich native hedgerow: -0.55 (deficit)")
    print("Native hedgerow with trees: -1.04 (deficit)")
    print("Non-native and ornamental hedgerow: -0.25 (deficit)")
    
    hedge_req_df = requirements["hedgerows"]
    
    print("\n📤 OUTPUT (requirements['hedgerows']):")
    print("-" * 90)
    if hedge_req_df.empty:
        print("(empty)")
        print()
        print("❌ UNEXPECTED: Requirements are empty despite missing distinctiveness")
        print("   This means deficits were somehow offset without distinctiveness info")
        success = False
    else:
        print(hedge_req_df.to_string(index=False))
        print()
        total_req = hedge_req_df["units"].sum()
        print(f"Total requirements: {total_req:.4f} units")
        print()
        print("✅ EXPECTED: Deficits appear in requirements because distinctiveness is unknown")
        print("   Without distinctiveness, trading rules cannot be applied")
        print("   This matches what the user reported seeing!")
        
        # Check if deficits are in requirements
        has_species_rich = not hedge_req_df[hedge_req_df["habitat"].str.contains("Species-rich native hedgerow", na=False) & 
                                            ~hedge_req_df["habitat"].str.contains("trees", na=False)].empty
        has_native_trees = not hedge_req_df[hedge_req_df["habitat"].str.contains("Native hedgerow with trees", na=False)].empty
        has_non_native = not hedge_req_df[hedge_req_df["habitat"].str.contains("Non-native", na=False)].empty
        
        if has_species_rich and has_native_trees and has_non_native:
            print()
            print("✅ All three deficits present (as expected when distinctiveness extraction fails)")
            success = True
        else:
            print()
            print("⚠️  Not all deficits present (unexpected)")
            success = False
    
    print("\n" + "="*90)
    print("CONCLUSION:")
    print("="*90)
    if success:
        print("This test demonstrates what happens when distinctiveness extraction fails.")
        print("Deficits cannot be offset without distinctiveness information.")
        print("This explains why the user sees deficits in their parsed requirements!")
    else:
        print("Test result was unexpected - needs further investigation")
    
    print("="*90 + "\n")
    
    return success


if __name__ == "__main__":
    success = test_failed_distinctiveness_extraction()
    exit(0 if success else 1)
