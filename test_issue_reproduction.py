"""
Test to reproduce the exact issue reported in the GitHub issue.

From the issue:
Very High Distinctiveness: Species-rich native hedgerow with trees - associated with bank or ditch: 0.00
High Distinctiveness: Species-rich native hedgerow with trees: 2.70 ✓
Medium Distinctiveness: 
  - Species-rich native hedgerow: -0.55 ⚠
  - Native hedgerow with trees: -1.04 ⚠
Very Low Distinctiveness: Non-native and ornamental hedgerow: -0.25 ⚠

Expected: High surplus (2.70) should offset all Medium and Very Low deficits (1.84 total)
Actual (from issue): The deficits appear in requirements with these values:
  - Species-rich native hedgerow: 0.554
  - Native hedgerow with trees: 1.035
  - Non-native and ornamental hedgerow: 0.246
"""

import io
import pandas as pd
import openpyxl
from metric_reader import parse_metric_requirements


class MockUploadedFile:
    """Mock file object for testing"""
    def __init__(self, buffer):
        self.buffer = buffer
        self.name = "test_metric.xlsx"
    
    def read(self):
        return self.buffer.read()


def test_exact_issue_reproduction():
    """
    Reproduce the exact issue scenario with precise values
    """
    
    wb = openpyxl.Workbook()
    
    # Create Headline Results sheet
    ws_headline = wb.create_sheet("Headline Results")
    
    headers = ["Unit Type", "Target", "Baseline Units", "Units Required", "Unit Deficit"]
    for col, header in enumerate(headers, start=1):
        ws_headline.cell(row=5, column=col, value=header)
    
    # From the issue, net gain is 10% and there's a need
    # Let's use baseline that would create similar requirements
    data = [
        ["Habitat units", "10.00%", 1.81, 2.00, 0.00],
        ["Hedgerow units", "10.00%", 1.84, 2.03, 0.00],  # 1.84 baseline, 10% = 0.184 net gain
        ["Watercourse units", "10.00%", 0.00, 0.00, 0.00],
    ]
    
    for row_idx, row_data in enumerate(data, start=6):
        for col_idx, value in enumerate(row_data, start=1):
            ws_headline.cell(row=row_idx, column=col_idx, value=value)
    
    # Create Trading Summary Hedgerows sheet with exact data from issue
    ws_hedge = wb.create_sheet("Trading Summary Hedgerows")
    
    hedge_headers = ["Habitat", "Distinctiveness", "Project-wide unit change"]
    for col, header in enumerate(hedge_headers, start=1):
        ws_hedge.cell(row=1, column=col, value=header)
    
    # Exact data from the issue
    hedge_data = [
        ["Species-rich native hedgerow with trees - associated with bank or ditch", "Very High", 0.00],
        ["Species-rich native hedgerow with trees", "High", 2.70],
        ["Species-rich native hedgerow", "Medium", -0.55],
        ["Native hedgerow with trees", "Medium", -1.04],
        ["Non-native and ornamental hedgerow", "Very Low", -0.25],
    ]
    
    for row_idx, row_data in enumerate(hedge_data, start=2):
        for col_idx, value in enumerate(row_data, start=1):
            ws_hedge.cell(row=row_idx, column=col_idx, value=value)
    
    # Remove default sheet
    if "Sheet" in wb.sheetnames:
        wb.remove(wb["Sheet"])
    
    # Save to BytesIO
    excel_buffer = io.BytesIO()
    wb.save(excel_buffer)
    excel_buffer.seek(0)
    
    # Create mock uploaded file
    mock_file = MockUploadedFile(excel_buffer)
    
    # Parse
    requirements = parse_metric_requirements(mock_file)
    
    # Display results
    print("\n" + "="*80)
    print("EXACT ISSUE REPRODUCTION TEST")
    print("="*80)
    
    print("\n📊 INPUT DATA (from GitHub issue):")
    print("-" * 80)
    print("Very High: Species-rich native hedgerow with trees - bank/ditch: 0.00")
    print("High:      Species-rich native hedgerow with trees: +2.70 ✓")
    print("Medium:    Species-rich native hedgerow: -0.55 ⚠")
    print("Medium:    Native hedgerow with trees: -1.04 ⚠")
    print("Very Low:  Non-native and ornamental hedgerow: -0.25 ⚠")
    print()
    print("Total surplus:  2.70 units")
    print("Total deficits: 0.55 + 1.04 + 0.25 = 1.84 units")
    print("Net gain:       1.84 × 10% = 0.184 units")
    print("Total need:     1.84 + 0.184 = 2.024 units")
    
    print("\n📋 TRADING RULES FOR HEDGEROWS:")
    print("-" * 80)
    print("✓ High surplus (2.70) CAN offset Medium deficits (High > Medium)")
    print("✓ High surplus (2.70) CAN offset Very Low deficits (High > Very Low)")
    print()
    print("Expected outcome:")
    print("  - All deficits (1.84) should be offset by High surplus (2.70)")
    print("  - Net gain (0.184) should also be covered by remaining surplus (0.86)")
    print("  - NO off-site requirements needed!")
    
    hedge_req_df = requirements["hedgerows"]
    
    print("\n🔍 ACTUAL REQUIREMENTS FROM PARSER:")
    print("-" * 80)
    if hedge_req_df.empty:
        print("  (empty - all requirements covered by on-site surplus)")
        print()
        print("✅ PASS: Issue is FIXED - surpluses correctly offsetting deficits!")
    else:
        print(hedge_req_df.to_string(index=False))
        print()
        
        # Check what appeared in requirements
        has_species_rich = hedge_req_df[hedge_req_df["habitat"].str.contains("Species-rich native hedgerow", case=False, na=False) & 
                                         ~hedge_req_df["habitat"].str.contains("trees", case=False, na=False)]
        has_native_with_trees = hedge_req_df[hedge_req_df["habitat"].str.contains("Native hedgerow with trees", case=False, na=False)]
        has_non_native = hedge_req_df[hedge_req_df["habitat"].str.contains("Non-native", case=False, na=False)]
        
        print("❌ FAIL: Issue is NOT fixed - deficits still appearing in requirements:")
        if not has_species_rich.empty:
            print(f"   • Species-rich native hedgerow: {has_species_rich.iloc[0]['units']:.3f} units")
        if not has_native_with_trees.empty:
            print(f"   • Native hedgerow with trees: {has_native_with_trees.iloc[0]['units']:.3f} units")
        if not has_non_native.empty:
            print(f"   • Non-native and ornamental hedgerow: {has_non_native.iloc[0]['units']:.3f} units")
        print()
        print("These should have been offset by the High surplus (2.70 units)")
    
    print("=" * 80)
    
    return hedge_req_df.empty


if __name__ == "__main__":
    success = test_exact_issue_reproduction()
    exit(0 if success else 1)
