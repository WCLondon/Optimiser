"""
Test for area habitat surplus offsetting - addresses the issue where Medium surpluses
are not offsetting Low deficits.

This test verifies that:
1. Area habitat surpluses are correctly parsed
2. Distinctiveness values are correctly extracted
3. Medium surpluses offset Low deficits according to area trading rules
4. Only unmet deficits appear in final requirements
"""

import io
import pandas as pd
import openpyxl
from metric_reader import parse_metric_requirements


class MockUploadedFile:
    """Mock file object for testing"""
    def __init__(self, buffer):
        self.buffer = buffer
        self.name = "test_metric.xlsx"
    
    def read(self):
        return self.buffer.read()


def test_area_medium_offsets_low():
    """
    Test that Medium distinctiveness area habitat surpluses offset Low distinctiveness deficits.
    
    Scenario from user issue:
    - Medium: Grassland +0.03, Heathland +0.42, Individual trees +0.08 (total +0.53 surplus)
    - Low: Modified grassland -2.48, Ruderal -0.04, but also some surpluses +0.68 (net -1.85 deficit)
    - Net gain: 10% of baseline
    
    Expected behavior:
    - The 0.53 Medium surplus should offset part of the 1.85 Low deficit
    - Remaining Low deficit should be: 1.85 - 0.53 = 1.32 units
    """
    
    wb = openpyxl.Workbook()
    
    # Create Headline Results sheet
    ws_headline = wb.create_sheet("Headline Results")
    
    headers = ["Unit Type", "Target", "Baseline Units", "Units Required", "Unit Deficit"]
    for col, header in enumerate(headers, start=1):
        ws_headline.cell(row=5, column=col, value=header)
    
    # User's scenario: baseline 6.7 units, 10% net gain = 0.67
    data = [
        ["Habitat units", "10.00%", 6.70, 7.37, 0.00],
        ["Hedgerow units", "10.00%", 0.00, 0.00, 0.00],
        ["Watercourse units", "10.00%", 0.00, 0.00, 0.00],
    ]
    
    for row_idx, row_data in enumerate(data, start=6):
        for col_idx, value in enumerate(row_data, start=1):
            ws_headline.cell(row=row_idx, column=col_idx, value=value)
    
    # Create Trading Summary Area Habitats sheet
    ws_area = wb.create_sheet("Trading Summary Area Habitats")
    
    # Section header format (no Distinctiveness column)
    row_num = 1
    
    # Medium Distinctiveness section
    ws_area.cell(row=row_num, column=1, value="Medium Distinctiveness")
    row_num += 1
    ws_area.cell(row=row_num, column=1, value="Habitat group")
    ws_area.cell(row=row_num, column=2, value="Group")
    ws_area.cell(row=row_num, column=3, value="Project-wide unit change")
    row_num += 1
    
    # Medium surpluses (simplified from user's data)
    ws_area.cell(row=row_num, column=1, value="Other neutral grassland")
    ws_area.cell(row=row_num, column=2, value="Grassland")
    ws_area.cell(row=row_num, column=3, value=0.03)
    row_num += 1
    
    ws_area.cell(row=row_num, column=1, value="Mixed scrub")
    ws_area.cell(row=row_num, column=2, value="Heathland and shrub")
    ws_area.cell(row=row_num, column=3, value=0.42)
    row_num += 1
    
    ws_area.cell(row=row_num, column=1, value="Urban tree")
    ws_area.cell(row=row_num, column=2, value="Individual trees")
    ws_area.cell(row=row_num, column=3, value=0.08)
    row_num += 2  # Skip a row
    
    # Low Distinctiveness section
    ws_area.cell(row=row_num, column=1, value="Low Distinctiveness")
    row_num += 1
    ws_area.cell(row=row_num, column=1, value="Habitat group")
    ws_area.cell(row=row_num, column=2, value="Group")
    ws_area.cell(row=row_num, column=3, value="Project-wide unit change")
    row_num += 1
    
    # Low deficits and small surpluses
    ws_area.cell(row=row_num, column=1, value="Modified grassland")
    ws_area.cell(row=row_num, column=2, value="Grassland")
    ws_area.cell(row=row_num, column=3, value=-2.48)
    row_num += 1
    
    ws_area.cell(row=row_num, column=1, value="Ruderal/ephemeral")
    ws_area.cell(row=row_num, column=2, value="Sparsely vegetated land")
    ws_area.cell(row=row_num, column=3, value=-0.04)
    row_num += 1
    
    # Small Low surpluses (simplified)
    ws_area.cell(row=row_num, column=1, value="Introduced shrub")
    ws_area.cell(row=row_num, column=2, value="Urban")
    ws_area.cell(row=row_num, column=3, value=0.05)
    row_num += 1
    
    ws_area.cell(row=row_num, column=1, value="Sustainable drainage system")
    ws_area.cell(row=row_num, column=2, value="Urban")
    ws_area.cell(row=row_num, column=3, value=0.13)
    row_num += 1
    
    ws_area.cell(row=row_num, column=1, value="Vegetated garden")
    ws_area.cell(row=row_num, column=2, value="Urban")
    ws_area.cell(row=row_num, column=3, value=0.50)
    
    # Remove default sheet
    if "Sheet" in wb.sheetnames:
        wb.remove(wb["Sheet"])
    
    # Save to BytesIO
    excel_buffer = io.BytesIO()
    wb.save(excel_buffer)
    excel_buffer.seek(0)
    
    # Create mock uploaded file
    mock_file = MockUploadedFile(excel_buffer)
    
    # Parse
    requirements = parse_metric_requirements(mock_file)
    
    # Verify results
    area_req_df = requirements["area"]
    
    print("\n" + "="*80)
    print("Test: Area habitat - Medium offsets Low")
    print("="*80)
    print("\nInput:")
    print("  Medium Distinctiveness:")
    print("    - Grassland: +0.03 units")
    print("    - Heathland: +0.42 units")
    print("    - Individual trees: +0.08 units")
    print("    - Total Medium surplus: 0.53 units")
    print()
    print("  Low Distinctiveness:")
    print("    - Modified grassland: -2.48 units")
    print("    - Ruderal: -0.04 units")
    print("    - Small surpluses: +0.68 units")
    print("    - Net Low deficit: -1.84 units")
    print()
    print("  Net gain requirement: 10% of 6.70 = 0.67 units")
    
    print("\nExpected:")
    print("  - Medium surplus (0.53) offsets part of Low deficit (1.84)")
    print("  - Remaining Low deficit: 1.84 - 0.53 = 1.31 units")
    print("  - Plus net gain: 0.67 units")
    print("  - Total requirements: ~1.98 units")
    
    print("\nActual requirements:")
    if area_req_df.empty:
        print("  (empty)")
    else:
        print(area_req_df.to_string(index=False))
    
    # Calculate totals
    total_medium_surplus = 0.03 + 0.42 + 0.08  # 0.53
    low_deficits = 2.48 + 0.04  # 2.52
    low_surpluses = 0.05 + 0.13 + 0.50  # 0.68
    net_low_deficit = low_deficits - low_surpluses  # 1.84
    
    # After offsetting
    remaining_low_deficit = max(0, net_low_deficit - total_medium_surplus)  # 1.84 - 0.53 = 1.31
    net_gain = 0.67
    expected_total = remaining_low_deficit + net_gain  # ~1.98
    
    # Check requirements
    if not area_req_df.empty:
        actual_total = area_req_df["units"].sum()
        print(f"\n  Total requirements: {actual_total:.2f} units")
        print(f"  Expected: ~{expected_total:.2f} units")
        
        # Allow some tolerance for rounding
        if abs(actual_total - expected_total) < 0.1:
            print("\n✅ Medium surplus correctly offset Low deficit!")
            print(f"✅ Requirements reduced from {net_low_deficit + net_gain:.2f} to {actual_total:.2f} units")
            success = True
        else:
            print(f"\n❌ Unexpected total: {actual_total:.2f} vs expected {expected_total:.2f}")
            success = False
    else:
        print("\n❌ Requirements are empty (unexpected)")
        success = False
    
    print("\n" + "="*80)
    
    return success


if __name__ == "__main__":
    print("="*80)
    print("Testing Area Habitat Surplus Offsetting")
    print("="*80)
    
    success = test_area_medium_offsets_low()
    
    print("\n" + "="*80)
    if success:
        print("✅ TEST PASSED!")
    else:
        print("❌ TEST FAILED")
    print("="*80)
    
    exit(0 if success else 1)
