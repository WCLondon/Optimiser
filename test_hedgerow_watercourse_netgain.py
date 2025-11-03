"""
Test for hedgerow and watercourse net gain calculation from Headline Results
"""

import io
import pandas as pd
import openpyxl
from metric_reader import parse_metric_requirements


class MockUploadedFile:
    """Mock file object for testing"""
    def __init__(self, buffer):
        self.buffer = buffer
        self.name = "test_metric.xlsx"
    
    def read(self):
        return self.buffer.read()


def test_hedgerow_netgain_with_deficits():
    """Test that hedgerow net gain is calculated and added when there are deficits"""
    
    wb = openpyxl.Workbook()
    
    # Create Headline Results sheet
    ws_headline = wb.create_sheet("Headline Results")
    
    headers = ["Unit Type", "Target", "Baseline Units", "Units Required", "Unit Deficit"]
    for col, header in enumerate(headers, start=1):
        ws_headline.cell(row=5, column=col, value=header)
    
    # Test data matching the issue example
    data = [
        ["Habitat units", "10.00%", 1.81, 2.00, 1.44],
        ["Hedgerow units", "10.00%", 0.94, 1.04, 1.04],
        ["Watercourse units", "10.00%", 0.00, 0.00, 0.00],
    ]
    
    for row_idx, row_data in enumerate(data, start=6):
        for col_idx, value in enumerate(row_data, start=1):
            ws_headline.cell(row=row_idx, column=col_idx, value=value)
    
    # Create Trading Summary Hedgerows sheet
    ws_hedge = wb.create_sheet("Trading Summary Hedgerows")
    
    hedge_headers = ["Habitat", "Project-wide unit change"]
    for col, header in enumerate(hedge_headers, start=1):
        ws_hedge.cell(row=1, column=col, value=header)
    
    hedge_data = [
        ["Native hedgerow", -1.04],
    ]
    
    for row_idx, row_data in enumerate(hedge_data, start=2):
        for col_idx, value in enumerate(row_data, start=1):
            ws_hedge.cell(row=row_idx, column=col_idx, value=value)
    
    # Remove default sheet
    if "Sheet" in wb.sheetnames:
        wb.remove(wb["Sheet"])
    
    # Save to BytesIO
    excel_buffer = io.BytesIO()
    wb.save(excel_buffer)
    excel_buffer.seek(0)
    
    # Create mock uploaded file
    mock_file = MockUploadedFile(excel_buffer)
    
    # Parse
    requirements = parse_metric_requirements(mock_file)
    
    # Verify baseline_info
    baseline_info = requirements["baseline_info"]
    assert baseline_info['hedgerow']['baseline_units'] == 0.94, \
        f"Expected 0.94, got {baseline_info['hedgerow']['baseline_units']}"
    assert baseline_info['hedgerow']['target_percent'] == 0.10, \
        f"Expected 0.10, got {baseline_info['hedgerow']['target_percent']}"
    
    # Calculate expected net gain
    expected_net_gain = 0.94 * 0.10
    
    # Check hedgerow requirements
    hedge_req_df = requirements["hedgerows"]
    assert not hedge_req_df.empty, "Hedgerow requirements should not be empty"
    
    # Find Net Gain row
    net_gain_rows = hedge_req_df[hedge_req_df["habitat"].str.contains("Net Gain", case=False, na=False)]
    assert not net_gain_rows.empty, "Net Gain (Hedgerows) should be in requirements"
    
    net_gain_units = net_gain_rows.iloc[0]["units"]
    assert abs(net_gain_units - expected_net_gain) < 0.0001, \
        f"Expected {expected_net_gain:.4f} net gain units, got {net_gain_units:.4f}"
    
    # Verify total includes both deficit and net gain
    total_units = hedge_req_df["units"].sum()
    expected_total = 1.04 + expected_net_gain
    assert abs(total_units - expected_total) < 0.0001, \
        f"Expected total {expected_total:.4f}, got {total_units:.4f}"
    
    print("✅ test_hedgerow_netgain_with_deficits passed")
    return True


def test_watercourse_netgain():
    """Test that watercourse net gain is calculated and added"""
    
    wb = openpyxl.Workbook()
    
    # Create Headline Results sheet
    ws_headline = wb.create_sheet("Headline Results")
    
    headers = ["Unit Type", "Target", "Baseline Units", "Units Required", "Unit Deficit"]
    for col, header in enumerate(headers, start=1):
        ws_headline.cell(row=5, column=col, value=header)
    
    # Test data with watercourse baseline
    data = [
        ["Habitat units", "10.00%", 1.00, 1.10, 0.10],
        ["Hedgerow units", "10.00%", 0.50, 0.55, 0.05],
        ["Watercourse units", "10.00%", 2.50, 2.75, 0.25],
    ]
    
    for row_idx, row_data in enumerate(data, start=6):
        for col_idx, value in enumerate(row_data, start=1):
            ws_headline.cell(row=row_idx, column=col_idx, value=value)
    
    # Create Trading Summary Watercourses sheet
    ws_water = wb.create_sheet("Trading Summary Watercourses")
    
    water_headers = ["Feature", "Project-wide unit change"]
    for col, header in enumerate(water_headers, start=1):
        ws_water.cell(row=1, column=col, value=header)
    
    water_data = [
        ["Ditch", -0.25],
    ]
    
    for row_idx, row_data in enumerate(water_data, start=2):
        for col_idx, value in enumerate(row_data, start=1):
            ws_water.cell(row=row_idx, column=col_idx, value=value)
    
    # Remove default sheet
    if "Sheet" in wb.sheetnames:
        wb.remove(wb["Sheet"])
    
    # Save to BytesIO
    excel_buffer = io.BytesIO()
    wb.save(excel_buffer)
    excel_buffer.seek(0)
    
    # Create mock uploaded file
    mock_file = MockUploadedFile(excel_buffer)
    
    # Parse
    requirements = parse_metric_requirements(mock_file)
    
    # Verify baseline_info
    baseline_info = requirements["baseline_info"]
    assert baseline_info['watercourse']['baseline_units'] == 2.50, \
        f"Expected 2.50, got {baseline_info['watercourse']['baseline_units']}"
    assert baseline_info['watercourse']['target_percent'] == 0.10, \
        f"Expected 0.10, got {baseline_info['watercourse']['target_percent']}"
    
    # Calculate expected net gain
    expected_net_gain = 2.50 * 0.10
    
    # Check watercourse requirements
    water_req_df = requirements["watercourses"]
    assert not water_req_df.empty, "Watercourse requirements should not be empty"
    
    # Find Net Gain row
    net_gain_rows = water_req_df[water_req_df["habitat"].str.contains("Net Gain", case=False, na=False)]
    assert not net_gain_rows.empty, "Net Gain (Watercourses) should be in requirements"
    
    net_gain_units = net_gain_rows.iloc[0]["units"]
    assert abs(net_gain_units - expected_net_gain) < 0.0001, \
        f"Expected {expected_net_gain:.4f} net gain units, got {net_gain_units:.4f}"
    
    # Verify total includes both deficit and net gain
    total_units = water_req_df["units"].sum()
    expected_total = 0.25 + expected_net_gain
    assert abs(total_units - expected_total) < 0.0001, \
        f"Expected total {expected_total:.4f}, got {total_units:.4f}"
    
    print("✅ test_watercourse_netgain passed")
    return True


def test_netgain_only_when_baseline_positive():
    """Test that net gain is not added when baseline is zero"""
    
    wb = openpyxl.Workbook()
    
    # Create Headline Results sheet
    ws_headline = wb.create_sheet("Headline Results")
    
    headers = ["Unit Type", "Target", "Baseline Units", "Units Required", "Unit Deficit"]
    for col, header in enumerate(headers, start=1):
        ws_headline.cell(row=5, column=col, value=header)
    
    # Test data - zero baseline for hedgerow and watercourse
    data = [
        ["Habitat units", "10.00%", 1.00, 1.10, 0.10],
        ["Hedgerow units", "10.00%", 0.00, 0.00, 0.00],
        ["Watercourse units", "10.00%", 0.00, 0.00, 0.00],
    ]
    
    for row_idx, row_data in enumerate(data, start=6):
        for col_idx, value in enumerate(row_data, start=1):
            ws_headline.cell(row=row_idx, column=col_idx, value=value)
    
    # Remove default sheet
    if "Sheet" in wb.sheetnames:
        wb.remove(wb["Sheet"])
    
    # Save to BytesIO
    excel_buffer = io.BytesIO()
    wb.save(excel_buffer)
    excel_buffer.seek(0)
    
    # Create mock uploaded file
    mock_file = MockUploadedFile(excel_buffer)
    
    # Parse
    requirements = parse_metric_requirements(mock_file)
    
    # Verify no net gain entries when baseline is zero
    hedge_req_df = requirements["hedgerows"]
    water_req_df = requirements["watercourses"]
    
    if not hedge_req_df.empty:
        has_net_gain = hedge_req_df["habitat"].str.contains("Net Gain", case=False, na=False).any()
        assert not has_net_gain, "Should not have Net Gain (Hedgerows) when baseline is 0"
    
    if not water_req_df.empty:
        has_net_gain = water_req_df["habitat"].str.contains("Net Gain", case=False, na=False).any()
        assert not has_net_gain, "Should not have Net Gain (Watercourses) when baseline is 0"
    
    print("✅ test_netgain_only_when_baseline_positive passed")
    return True


if __name__ == "__main__":
    print("=" * 60)
    print("Testing Hedgerow and Watercourse Net Gain Calculations")
    print("=" * 60)
    
    success1 = test_hedgerow_netgain_with_deficits()
    print()
    success2 = test_watercourse_netgain()
    print()
    success3 = test_netgain_only_when_baseline_positive()
    
    print("\n" + "=" * 60)
    if success1 and success2 and success3:
        print("✅ ALL TESTS PASSED!")
    else:
        print("❌ SOME TESTS FAILED")
    print("=" * 60)
    
    exit(0 if (success1 and success2 and success3) else 1)
