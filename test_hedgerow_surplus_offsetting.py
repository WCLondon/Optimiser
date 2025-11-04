"""
Test for hedgerow surplus offsetting - addresses GitHub issue about hedgerow surpluses
not being read by the metric reader and therefore not mitigating downstream deficits.

This test verifies that:
1. Hedgerow surpluses are correctly parsed from Trading Summary sheets
2. Distinctiveness values are correctly extracted from Distinctiveness column
3. Surpluses offset deficits according to hedgerow trading rules
4. Only unmet deficits appear in final requirements
"""

import io
import pandas as pd
import openpyxl
from metric_reader import parse_metric_requirements


class MockUploadedFile:
    """Mock file object for testing"""
    def __init__(self, buffer):
        self.buffer = buffer
        self.name = "test_metric.xlsx"
    
    def read(self):
        return self.buffer.read()


def test_hedgerow_surplus_offsetting_with_distinctiveness_column():
    """
    Test that hedgerow surpluses offset deficits when distinctiveness is in a column.
    
    Scenario from GitHub issue:
    - High: Species-rich native hedgerow with trees: +0.37 units (surplus)
    - Medium: Species-rich native hedgerow: +0.13 units (surplus)
    - Very Low: Non-native and ornamental hedgerow: -0.03 units (deficit)
    
    Expected behavior:
    - The 0.03 Very Low deficit should be offset by the 0.50 total surplus
    - Only net gain requirement should appear in final requirements
    """
    
    wb = openpyxl.Workbook()
    
    # Create Headline Results sheet
    ws_headline = wb.create_sheet("Headline Results")
    
    headers = ["Unit Type", "Target", "Baseline Units", "Units Required", "Unit Deficit"]
    for col, header in enumerate(headers, start=1):
        ws_headline.cell(row=5, column=col, value=header)
    
    data = [
        ["Habitat units", "10.00%", 1.00, 1.10, 0.00],
        ["Hedgerow units", "10.00%", 0.94, 1.04, 0.00],
        ["Watercourse units", "10.00%", 0.00, 0.00, 0.00],
    ]
    
    for row_idx, row_data in enumerate(data, start=6):
        for col_idx, value in enumerate(row_data, start=1):
            ws_headline.cell(row=row_idx, column=col_idx, value=value)
    
    # Create Trading Summary Hedgerows sheet with Distinctiveness column
    ws_hedge = wb.create_sheet("Trading Summary Hedgerows")
    
    hedge_headers = ["Habitat", "Distinctiveness", "Project-wide unit change"]
    for col, header in enumerate(hedge_headers, start=1):
        ws_hedge.cell(row=1, column=col, value=header)
    
    # Data from the GitHub issue
    hedge_data = [
        ["Species-rich native hedgerow with trees", "High", 0.37],
        ["Species-rich native hedgerow", "Medium", 0.13],
        ["Non-native and ornamental hedgerow", "Very Low", -0.03],
    ]
    
    for row_idx, row_data in enumerate(hedge_data, start=2):
        for col_idx, value in enumerate(row_data, start=1):
            ws_hedge.cell(row=row_idx, column=col_idx, value=value)
    
    # Remove default sheet
    if "Sheet" in wb.sheetnames:
        wb.remove(wb["Sheet"])
    
    # Save to BytesIO
    excel_buffer = io.BytesIO()
    wb.save(excel_buffer)
    excel_buffer.seek(0)
    
    # Create mock uploaded file
    mock_file = MockUploadedFile(excel_buffer)
    
    # Parse
    requirements = parse_metric_requirements(mock_file)
    
    # Verify results
    hedge_req_df = requirements["hedgerows"]
    
    print("\n" + "="*60)
    print("Test: Hedgerow surplus offsetting with Distinctiveness column")
    print("="*60)
    print("\nInput:")
    print("  High: Species-rich native hedgerow with trees: +0.37 units")
    print("  Medium: Species-rich native hedgerow: +0.13 units")
    print("  Very Low: Non-native and ornamental hedgerow: -0.03 units")
    print("  Net gain requirement: 0.94 × 10% = 0.094 units")
    print("\nExpected:")
    print("  Total surplus: 0.50 units")
    print("  Very Low deficit (0.03) offset by surplus")
    print("  Net gain (0.094) also covered by remaining surplus")
    print("  No off-site requirements needed!")
    
    print("\nActual requirements:")
    if hedge_req_df.empty:
        print("  (empty - all requirements covered by on-site surplus)")
    else:
        print(hedge_req_df)
    
    # Check that Non-native hedgerow deficit is NOT in requirements
    deficit_rows = hedge_req_df[hedge_req_df["habitat"].str.contains("Non-native", case=False, na=False)]
    assert deficit_rows.empty, "❌ Non-native hedgerow deficit should have been offset by surplus"
    
    # In this case, surplus covers BOTH deficit and net gain
    # So requirements should be empty or only have remaining unmet items
    total_surplus = 0.37 + 0.13  # 0.50
    total_need = 0.03 + 0.094    # 0.124
    
    if total_surplus >= total_need:
        # Surplus covers everything
        assert hedge_req_df.empty or hedge_req_df["units"].sum() < 0.001, \
            "❌ All requirements should be covered by surplus"
        print("\n✅ Non-native hedgerow deficit correctly offset by surplus")
        print("✅ Net gain also covered by remaining surplus (0.47 units)")
        print("✅ No off-site requirements needed!")
    else:
        # Some requirements remain
        net_gain_rows = hedge_req_df[hedge_req_df["habitat"].str.contains("Net Gain", case=False, na=False)]
        assert not net_gain_rows.empty, "❌ Net Gain (Hedgerows) should be in requirements"
        expected_remaining = total_need - total_surplus
        total_units = hedge_req_df["units"].sum()
        assert abs(total_units - expected_remaining) < 0.001, \
            f"❌ Remaining requirement should be {expected_remaining:.4f}, got {total_units:.4f}"
        print(f"\n✅ Requirements correctly calculated: {total_units:.4f} units")
    
    return True


def test_hedgerow_offsetting_with_section_headers():
    """
    Test that hedgerow surpluses offset deficits when distinctiveness is in section headers.
    
    This tests backward compatibility with the old format where distinctiveness
    is extracted from section headers like "High Distinctiveness" instead of a column.
    """
    
    wb = openpyxl.Workbook()
    
    # Create Headline Results sheet
    ws_headline = wb.create_sheet("Headline Results")
    
    headers = ["Unit Type", "Target", "Baseline Units", "Units Required", "Unit Deficit"]
    for col, header in enumerate(headers, start=1):
        ws_headline.cell(row=5, column=col, value=header)
    
    data = [
        ["Habitat units", "10.00%", 1.00, 1.10, 0.00],
        ["Hedgerow units", "10.00%", 0.94, 1.04, 0.00],
        ["Watercourse units", "10.00%", 0.00, 0.00, 0.00],
    ]
    
    for row_idx, row_data in enumerate(data, start=6):
        for col_idx, value in enumerate(row_data, start=1):
            ws_headline.cell(row=row_idx, column=col_idx, value=value)
    
    # Create Trading Summary Hedgerows sheet with section headers
    ws_hedge = wb.create_sheet("Trading Summary Hedgerows")
    
    # High Distinctiveness section
    ws_hedge.cell(row=1, column=1, value="High Distinctiveness")
    ws_hedge.cell(row=2, column=1, value="Habitat")
    ws_hedge.cell(row=2, column=2, value="Project-wide unit change")
    ws_hedge.cell(row=3, column=1, value="Species-rich native hedgerow with trees")
    ws_hedge.cell(row=3, column=2, value=0.37)
    
    # Medium Distinctiveness section
    ws_hedge.cell(row=5, column=1, value="Medium Distinctiveness")
    ws_hedge.cell(row=6, column=1, value="Habitat")
    ws_hedge.cell(row=6, column=2, value="Project-wide unit change")
    ws_hedge.cell(row=7, column=1, value="Species-rich native hedgerow")
    ws_hedge.cell(row=7, column=2, value=0.13)
    
    # Very Low Distinctiveness section
    ws_hedge.cell(row=9, column=1, value="Very Low Distinctiveness")
    ws_hedge.cell(row=10, column=1, value="Habitat")
    ws_hedge.cell(row=10, column=2, value="Project-wide unit change")
    ws_hedge.cell(row=11, column=1, value="Non-native and ornamental hedgerow")
    ws_hedge.cell(row=11, column=2, value=-0.03)
    
    # Remove default sheet
    if "Sheet" in wb.sheetnames:
        wb.remove(wb["Sheet"])
    
    # Save to BytesIO
    excel_buffer = io.BytesIO()
    wb.save(excel_buffer)
    excel_buffer.seek(0)
    
    # Create mock uploaded file
    mock_file = MockUploadedFile(excel_buffer)
    
    # Parse
    requirements = parse_metric_requirements(mock_file)
    
    # Verify results
    hedge_req_df = requirements["hedgerows"]
    
    print("\n" + "="*60)
    print("Test: Hedgerow surplus offsetting with section headers")
    print("="*60)
    print("\nInput (with section headers format):")
    print("  High: Species-rich native hedgerow with trees: +0.37 units")
    print("  Medium: Species-rich native hedgerow: +0.13 units")
    print("  Very Low: Non-native and ornamental hedgerow: -0.03 units")
    print("  Net gain requirement: 0.94 × 10% = 0.094 units")
    
    print("\nActual requirements:")
    if hedge_req_df.empty:
        print("  (empty - all requirements covered by on-site surplus)")
    else:
        print(hedge_req_df)
    
    # Check that Non-native hedgerow deficit is NOT in requirements
    deficit_rows = hedge_req_df[hedge_req_df["habitat"].str.contains("Non-native", case=False, na=False)]
    assert deficit_rows.empty, "❌ Non-native hedgerow deficit should have been offset by surplus"
    
    # In this case, surplus covers BOTH deficit and net gain
    total_surplus = 0.37 + 0.13  # 0.50
    total_need = 0.03 + 0.094    # 0.124
    
    if total_surplus >= total_need:
        # Surplus covers everything
        assert hedge_req_df.empty or hedge_req_df["units"].sum() < 0.001, \
            "❌ All requirements should be covered by surplus"
        print("\n✅ Section header format also works correctly")
        print("✅ Deficit correctly offset by surplus")
        print("✅ Net gain also covered by surplus")
        print("✅ No off-site requirements needed!")
    else:
        # Some requirements remain
        net_gain_rows = hedge_req_df[hedge_req_df["habitat"].str.contains("Net Gain", case=False, na=False)]
        assert not net_gain_rows.empty, "❌ Net Gain (Hedgerows) should be in requirements"
        expected_remaining = total_need - total_surplus
        total_units = hedge_req_df["units"].sum()
        assert abs(total_units - expected_remaining) < 0.001, \
            f"❌ Remaining requirement should be {expected_remaining:.4f}, got {total_units:.4f}"
        print(f"\n✅ Requirements correctly calculated: {total_units:.4f} units")
    
    return True


def test_hedgerow_trading_rules():
    """
    Test that hedgerow trading rules are applied correctly:
    - Very High & High: Like-for-like only (same habitat)
    - Medium, Low, Very Low: Same distinctiveness or better
    """
    
    wb = openpyxl.Workbook()
    
    # Create Headline Results sheet
    ws_headline = wb.create_sheet("Headline Results")
    headers = ["Unit Type", "Target", "Baseline Units", "Units Required", "Unit Deficit"]
    for col, header in enumerate(headers, start=1):
        ws_headline.cell(row=5, column=col, value=header)
    
    data = [
        ["Habitat units", "10.00%", 0.00, 0.00, 0.00],
        ["Hedgerow units", "10.00%", 0.00, 0.00, 0.00],
        ["Watercourse units", "10.00%", 0.00, 0.00, 0.00],
    ]
    
    for row_idx, row_data in enumerate(data, start=6):
        for col_idx, value in enumerate(row_data, start=1):
            ws_headline.cell(row=row_idx, column=col_idx, value=value)
    
    # Create Trading Summary Hedgerows sheet
    ws_hedge = wb.create_sheet("Trading Summary Hedgerows")
    
    hedge_headers = ["Habitat", "Distinctiveness", "Project-wide unit change"]
    for col, header in enumerate(hedge_headers, start=1):
        ws_hedge.cell(row=1, column=col, value=header)
    
    # Test scenarios:
    # 1. Medium surplus (0.5) should offset Low deficit (0.2)
    # 2. High surplus (0.3) should offset Medium deficit (0.1)
    # 3. Very High deficit (0.5) should NOT be offset (no matching surplus)
    hedge_data = [
        ["Native hedgerow A", "Medium", 0.5],   # Surplus
        ["Native hedgerow B", "High", 0.3],     # Surplus
        ["Native hedgerow C", "Low", -0.2],     # Deficit - should be offset by Medium
        ["Native hedgerow D", "Medium", -0.1],  # Deficit - should be offset by High
        ["Native hedgerow E", "Very High", -0.5],  # Deficit - cannot be offset (like-for-like required)
    ]
    
    for row_idx, row_data in enumerate(hedge_data, start=2):
        for col_idx, value in enumerate(row_data, start=1):
            ws_hedge.cell(row=row_idx, column=col_idx, value=value)
    
    # Remove default sheet
    if "Sheet" in wb.sheetnames:
        wb.remove(wb["Sheet"])
    
    # Save to BytesIO
    excel_buffer = io.BytesIO()
    wb.save(excel_buffer)
    excel_buffer.seek(0)
    
    # Create mock uploaded file
    mock_file = MockUploadedFile(excel_buffer)
    
    # Parse
    requirements = parse_metric_requirements(mock_file)
    
    hedge_req_df = requirements["hedgerows"]
    
    print("\n" + "="*60)
    print("Test: Hedgerow trading rules")
    print("="*60)
    print("\nInput:")
    print("  Medium surplus: Native hedgerow A: +0.5 units")
    print("  High surplus: Native hedgerow B: +0.3 units")
    print("  Low deficit: Native hedgerow C: -0.2 units")
    print("  Medium deficit: Native hedgerow D: -0.1 units")
    print("  Very High deficit: Native hedgerow E: -0.5 units")
    
    print("\nExpected:")
    print("  Low deficit offset by Medium surplus ✓")
    print("  Medium deficit offset by High surplus ✓")
    print("  Very High deficit NOT offset (like-for-like required)")
    
    print("\nActual requirements:")
    print(hedge_req_df)
    
    # Check that Low and Medium deficits are NOT in requirements
    low_deficit = hedge_req_df[hedge_req_df["habitat"] == "Native hedgerow C"]
    medium_deficit = hedge_req_df[hedge_req_df["habitat"] == "Native hedgerow D"]
    assert low_deficit.empty, "❌ Low deficit should have been offset by Medium surplus"
    assert medium_deficit.empty, "❌ Medium deficit should have been offset by High surplus"
    
    # Check that Very High deficit IS in requirements
    vh_deficit = hedge_req_df[hedge_req_df["habitat"] == "Native hedgerow E"]
    assert not vh_deficit.empty, "❌ Very High deficit should remain (like-for-like required)"
    assert abs(vh_deficit.iloc[0]["units"] - 0.5) < 0.001, \
        f"❌ Very High deficit should be 0.5, got {vh_deficit.iloc[0]['units']}"
    
    print("\n✅ Low deficit correctly offset by Medium surplus")
    print("✅ Medium deficit correctly offset by High surplus")
    print("✅ Very High deficit correctly remains (like-for-like required)")
    
    return True


if __name__ == "__main__":
    print("=" * 60)
    print("Testing Hedgerow Surplus Offsetting")
    print("=" * 60)
    
    success1 = test_hedgerow_surplus_offsetting_with_distinctiveness_column()
    print()
    success2 = test_hedgerow_offsetting_with_section_headers()
    print()
    success3 = test_hedgerow_trading_rules()
    
    print("\n" + "=" * 60)
    if success1 and success2 and success3:
        print("✅ ALL TESTS PASSED!")
    else:
        print("❌ SOME TESTS FAILED")
    print("=" * 60)
    
    exit(0 if (success1 and success2 and success3) else 1)
